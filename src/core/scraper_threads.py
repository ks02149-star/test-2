from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
import pandas as pd
import concurrent.futures
import traceback
from PyQt5.QtCore import QDate
from src.utils.spell_checker_core import check_text
import os
import sys
import json
import time
import datetime
import urllib.parse
import re
import random
import hashlib
import tempfile
import threading
import requests
from bs4 import BeautifulSoup
from bs4 import BeautifulSoup
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QObject
from src.config import SESSION, WORKSPACE_DIR, ASSETS_DIR, DATA_DIR, SETTINGS_PATH, CREDENTIALS_PATH

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from webdriver_manager.chrome import ChromeDriverManager
import gspread
from oauth2client.service_account import ServiceAccountCredentials

class WorkerSignals(QObject):
    log = pyqtSignal(str)
    finished = pyqtSignal()
    error = pyqtSignal(str)
    match_found = pyqtSignal(dict)


class DriverInitWorker(QThread):
    finished = pyqtSignal(str)
    error = pyqtSignal(str)

    def run(self):
        try:
            import os, re
            from webdriver_manager.chrome import ChromeDriverManager
            original_path = ChromeDriverManager().install()
            
            driver_dir = os.path.dirname(original_path)
            patched_path = os.path.join(driver_dir, "chromedriver_patched.exe")
            
            if not os.path.exists(patched_path):
                with open(original_path, 'rb') as f:
                    binary_data = f.read()
                
                patched_binary = re.sub(b'cdc_[a-zA-Z0-9]{22}_', b'rnd_' + b'X' * 22 + b'_', binary_data)
                
                with open(patched_path, 'wb') as f:
                    f.write(patched_binary)
                    
            self.finished.emit(patched_path)
        except Exception as e:
            self.error.emit(str(e))


class ScraperWorker(QThread):
    def __init__(self, excel_path, selected_sheet, display_count, capture_screenshot, global_driver_path, is_custom_excel):
        super().__init__()
        self.excel_path = excel_path
        self.selected_sheet = selected_sheet
        self.display_count = display_count
        self.capture_screenshot = capture_screenshot
        self.global_driver_path = global_driver_path
        self.is_custom_excel = is_custom_excel
        self.signals = WorkerSignals()
        
        self.driver_pool_lock = threading.Lock()
        self.active_drivers = []
        self.thread_local = threading.local()

    def log(self, message):
        self.signals.log.emit(message)

    def get_random_android_ua_and_size(self):
        import random
        pool = [
            ("Mozilla/5.0 (Linux; Android 14; SM-S928N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.6478.122 Mobile Safari/537.36", "412,915"),
            ("Mozilla/5.0 (Linux; Android 13; SM-A536N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.6422.113 Mobile Safari/537.36", "412,915"),
            ("Mozilla/5.0 (Linux; Android 14; SM-F731N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.6367.113 Mobile Safari/537.36", "360,844"),
            ("Mozilla/5.0 (Linux; Android 13; SM-G998N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.6312.118 Mobile Safari/537.36", "384,853"),
            ("Mozilla/5.0 (Linux; Android 14; SM-S918N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.6478.122 Mobile Safari/537.36", "412,915")
        ]
        return random.choice(pool)

    def get_thread_driver(self):
        if not hasattr(self.thread_local, 'driver'):
            from src.utils.driver_manager import setup_chrome_options
            from selenium.webdriver.chrome.options import Options
            base_options = setup_chrome_options()
            options = Options()
            
            ua, size = self.get_random_android_ua_and_size()
            
            for arg in base_options.arguments:
                if not arg.startswith('--window-size=') and not arg.startswith('user-agent='):
                    options.add_argument(arg)
                    
            for key, val in base_options.experimental_options.items():
                options.add_experimental_option(key, val)
                
            options.add_argument(f"user-agent={ua}")
            options.add_argument(f"--window-size={size}")
            options.page_load_strategy = 'eager'
            
            service = Service(self.global_driver_path)
            driver = webdriver.Chrome(service=service, options=options)
            driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {"source": "Object.defineProperty(navigator, \'webdriver\', { get: () => undefined })"})
            self.thread_local.driver = driver
            
            with self.driver_pool_lock:
                self.active_drivers.append(driver)
                
        return self.thread_local.driver

    def cleanup_drivers(self):
        with self.driver_pool_lock:
            for d in self.active_drivers:
                try:
                    d.quit()
                except:
                    pass
            self.active_drivers.clear()

    def get_blog_data_by_crawling(self, driver, keyword, target_blogs):
        if target_blogs is None: target_blogs = []
        titles, blog_names = [], []
        sort_name = "관련도순"
        screenshot_filepaths = []
        target_folder_path = ""
        
        self.log(f"▶ '{keyword}' - 데이터 탐색 중... ({self.display_count}개 목표)")
        
        try:
            url = f"https://m.search.naver.com/search.naver?ssc=tab.m_blog.all&sm=mtb_jum&query={urllib.parse.quote(keyword)}"
            driver.get(url)
            
            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, ".lst_total > li.bx, div[data-template-id='ugcItem']"))
                )
            except TimeoutException:
                self.log(f"[안내] '{keyword}' - 요소를 찾지 못해 대기 시간이 초과되었습니다.")
                return pd.DataFrame({'블로그명': pd.Series(dtype='str'), '제목': pd.Series(dtype='str')}), [], ""

            last_height = driver.execute_script("return document.body.scrollHeight")
            
            while True:
                js_count = """
                var els = document.querySelectorAll(".lst_total > li.bx");
                if (els.length === 0) els = document.querySelectorAll("div[data-template-id='ugcItem']");
                var organicCount = 0;
                for(var i=0; i<els.length; i++) {
                    if (!els[i].querySelector(".sp_ad, .ad_ico, a[data-heatmap-target='articleSourceJSX_adtag']")) {
                        organicCount++;
                    }
                }
                return organicCount;
                """
                current_count = driver.execute_script(js_count)
                
                if current_count >= self.display_count:
                    break
                    
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                
                waited = 0
                new_height = last_height
                while waited < 2.0:
                    time.sleep(0.2)
                    waited += 0.2
                    new_height = driver.execute_script("return document.body.scrollHeight")
                    if new_height > last_height:
                        break
                        
                if new_height == last_height:
                    break
                last_height = new_height

            js_extract = """
            var results = [];
            var containers = document.querySelectorAll(".lst_total > li.bx");
            if (containers.length === 0) containers = document.querySelectorAll("div[data-template-id='ugcItem']");
            
            for(var i=0; i<containers.length; i++) {
                if (containers[i].querySelector(".sp_ad, .ad_ico, a[data-heatmap-target='articleSourceJSX_adtag']")) {
                    continue;
                }
                
                var t = containers[i].querySelector(".api_txt_lines.total_tit, .title_link, a[data-heatmap-target='.nblg']");
                var b = containers[i].querySelector(".name, .sub_txt, a[data-heatmap-target='articleSourceJSX_title']");
                
                var titleText = t ? t.innerText.trim() : "제목 추출 불가";
                var blogText = "";
                if (b) {
                    blogText = b.innerText.trim();
                } else {
                    blogText = containers[i].innerText.replace(/\\n/g, ' ').substring(0, 50);
                }
                
                results.push({title: titleText, blog: blogText});
            }
            return results;
            """
            extracted_data = driver.execute_script(js_extract)
            
            for data in extracted_data[:self.display_count]:
                titles.append(data['title'])
                blog_names.append(data['blog'])
                    
            if self.capture_screenshot and target_blogs and titles:
                js_highlight_and_count = """
                var targets = arguments[0];
                var containers = document.querySelectorAll(".lst_total > li.bx");
                if (containers.length === 0) containers = document.querySelectorAll("div[data-template-id='ugcItem']");
                
                var matchRanks = [];
                var matchCount = 0;
                var actualRank = 1;
                
                for (var i = 0; i < containers.length; i++) {
                    if (containers[i].querySelector(".sp_ad, .ad_ico, a[data-heatmap-target='articleSourceJSX_adtag']")) {
                        continue;
                    }
                    
                    var b = containers[i].querySelector(".name, .sub_txt, a[data-heatmap-target='articleSourceJSX_title']");
                    var blogNameText = b ? b.innerText.trim() : "";
                    
                    var isMatch = false;
                    for (var j = 0; j < targets.length; j++) {
                        if (targets[j].trim() !== "" && blogNameText.includes(targets[j].trim())) {
                            isMatch = true;
                            break;
                        }
                    }
                    
                    if (isMatch) {
                        var targetElement = b || containers[i];
                        targetElement.style.backgroundColor = 'yellow';
                        targetElement.style.color = 'black';
                        containers[i].setAttribute('data-target-match', matchCount);
                        matchRanks.push(actualRank);
                        matchCount++;
                    }
                    
                    actualRank++;
                }
                return matchRanks;
                """
                match_ranks = driver.execute_script(js_highlight_and_count, target_blogs)
                
                if match_ranks:
                    base_dir = os.path.dirname(os.path.abspath(__file__)) if not getattr(sys, 'frozen', False) else os.path.dirname(sys.executable)
                    workspace_dir = os.path.join(base_dir, "Workspace")
                    
                    safe_keyword = re.sub(r'[\\/*?:"<>|]', "", keyword)
                    representative_target = target_blogs[0]
                    safe_target = re.sub(r'[\\/*?:"<>|]', "", representative_target)
                    timestamp = time.strftime("%Y%m%d_%H%M%S")
                    
                    target_folder_path = os.path.join(workspace_dir, safe_target)
                    
                    if self.is_custom_excel:
                        target_folder_path = os.path.join(target_folder_path, safe_keyword)
                        
                    os.makedirs(target_folder_path, exist_ok=True)
                    
                    for i, rank in enumerate(match_ranks):
                        js_scroll = f"""
                        var el = document.querySelector("[data-target-match='{i}']");
                        if (el) {{ el.scrollIntoView({{behavior: 'instant', block: 'center'}}); }}
                        """
                        driver.execute_script(js_scroll)
                        time.sleep(1)
                        
                        screenshot_filename = f"{safe_keyword}_{sort_name}_{rank}위노출_{timestamp}.png"
                        screenshot_filepath = os.path.join(target_folder_path, screenshot_filename)
                        driver.save_screenshot(screenshot_filepath)
                        screenshot_filepaths.append((rank, screenshot_filepath))
                        
                        if self.is_custom_excel:
                            self.log(f"[캡처] 목표({representative_target}) {rank}위 저장 완료: '{safe_target}/{safe_keyword}'")
                        else:
                            self.log(f"[캡처] 목표({representative_target}) {rank}위 저장 완료: '{safe_target}'")
                else:
                    self.log(f"[안내] 목표 블로그 '{', '.join(target_blogs)}'가 존재하지 않아 캡처 생략.")
            
            if not titles:
                self.log(f"[안내] '{keyword}' - 데이터를 추출하지 못했습니다.")
                        
        except Exception as e:
            self.log(f"[오류] '{keyword}' - 크롤링 실패: {e}")
                
        df = pd.DataFrame({'블로그명': blog_names, '제목': titles})
        df['제목'] = df['제목'].apply(lambda x: x[:20] + '...' if len(x) > 20 else x)
        df.index = df.index + 1
        return df, screenshot_filepaths, target_folder_path

    def get_rank_string(self, df_result, target_list):
        if df_result.empty or not target_list:
            return "검색/분석 불가"
        
        valid_targets = [t.strip() for t in target_list if t.strip()]
        if not valid_targets:
            return "미노출 (범위 밖)"
            
        pattern = "|".join([re.escape(t) for t in valid_targets])
        mask = df_result['블로그명'].str.contains(pattern, na=False, regex=True)
        found_indices = df_result.index[mask].tolist()
        if found_indices:
            return ", ".join([f"{i}위" for i in found_indices])
        return "미노출 (범위 밖)"

    def run(self):

        try:
            df_input = pd.read_excel(self.excel_path, sheet_name=self.selected_sheet)
            required_columns = ['키워드', '목표블로그', '관련도순_순위', '최근조회일시']
            for col in required_columns:
                if col not in df_input.columns:
                    df_input[col] = None
                    
            df_input['관련도순_순위'] = df_input['관련도순_순위'].astype(object)
            total_rows = len(df_input)
            
            valid_keywords = df_input['키워드'].dropna().astype(str).str.strip()
            valid_keywords = valid_keywords[(valid_keywords.str.lower() != 'nan') & (valid_keywords != '') & (valid_keywords.str.lower() != 'none')]
            
            if total_rows == 0 or valid_keywords.empty:
                self.signals.error.emit(f"'{self.selected_sheet}' 탭의 엑셀에 유효한 키워드가 비어있습니다.")
                return

            self.log(f"\n총 {total_rows}개의 작업 목록 발견. 병렬 조회를 시작합니다.\n" + "-"*40)
            
            def process_task(index, row_dict):
                raw_keyword = str(row_dict['키워드']).strip()
                raw_target = str(row_dict['목표블로그']).strip()
                keyword = raw_keyword if pd.notna(row_dict['키워드']) and raw_keyword.lower() != 'nan' else ""
                target_blogs = [t.strip() for t in raw_target.split(',') if t.strip()] if pd.notna(row_dict['목표블로그']) and raw_target.lower() != 'nan' else []
                display_targets = ", ".join(target_blogs)
                
                if not keyword: return None
                    
                self.log(f"\n=== [{index + 1}/{total_rows}] 키워드: '{keyword}', 목표: '{display_targets}' ===")
                
                delay_time = random.uniform(2.5, 4.5)
                self.log(f"※ 봇 탐지 우회를 위해 {delay_time:.1f}초 대기 중...")
                time.sleep(delay_time)
                
                driver = self.get_thread_driver()
                df_sim, screenshots, folder_path = self.get_blog_data_by_crawling(driver, keyword, target_blogs)
                sim_rank = self.get_rank_string(df_sim, target_blogs)
                
                ranks = []
                if sim_rank and "위" in sim_rank:
                    for part in sim_rank.split(","):
                        if "위" in part:
                            try:
                                r_val = int(part.replace("위", "").strip())
                                ranks.append(r_val)
                            except:
                                pass
                                
                if ranks:
                    representative_target = target_blogs[0]
                    if not screenshots:
                        screenshots = [(r, "") for r in ranks]
                    self.signals.match_found.emit({
                        'company': representative_target,
                        'keyword': keyword,
                        'ranks': ranks,
                        'screenshots': screenshots,
                        'folder': folder_path
                    })
                
                result_text = f"<< '{keyword}' 검색 결과 >>\n"
                for idx, row in df_sim.iterrows():
                    result_text += f"[{idx}위] {row['블로그명']} | {row['제목']}\n"
                self.log(result_text)
                self.log(f"▶ '{keyword}' 기록 준비 완료: {sim_rank}")
                
                return {'keyword': keyword, 'target_blog': raw_target, 'sim_rank': sim_rank}

            results_buffer = [] 
            
            with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
                futures = []
                for index, row in df_input.iterrows():
                    futures.append(executor.submit(process_task, index, row.to_dict()))
                    
                for future in concurrent.futures.as_completed(futures):
                    try:
                        result = future.result() 
                        if result: results_buffer.append(result)
                    except Exception as e:
                        self.log(f"[치명적 오류] 작업 중단: {e}")
                            
            if results_buffer:
                self.log("\n모든 탐색 완료. 엑셀 파일에 결과를 저장합니다...")
                while True:
                    try:
                        wb = openpyxl.load_workbook(self.excel_path)
                        ws = wb[self.selected_sheet] if self.selected_sheet in wb.sheetnames else wb.active
                        
                        header_map = {cell.value: cell.column for cell in ws[1]}
                        for req_col in required_columns:
                            if req_col not in header_map:
                                new_col_idx = ws.max_column + 1
                                ws.cell(row=1, column=new_col_idx, value=req_col)
                                header_map[req_col] = new_col_idx
                                
                        keyword_col_idx = header_map['키워드']
                        target_col_idx = header_map['목표블로그']
                        
                        for data in results_buffer:
                            current_row = None
                            for r in range(2, ws.max_row + 1):
                                cell_keyword = str(ws.cell(row=r, column=keyword_col_idx).value or "").strip()
                                cell_target = str(ws.cell(row=r, column=target_col_idx).value or "").strip()
                                if cell_keyword == data['keyword'] and cell_target == data['target_blog']:
                                    current_row = r
                                    break
                                    
                            if current_row is None:
                                current_row = ws.max_row + 1
                                ws.cell(row=current_row, column=keyword_col_idx, value=data['keyword'])
                                ws.cell(row=current_row, column=target_col_idx, value=data['target_blog'])
                            
                            current_time = time.strftime("%Y-%m-%d %H:%M:%S")
                            ws.cell(row=current_row, column=header_map['관련도순_순위'], value=data['sim_rank'])
                            ws.cell(row=current_row, column=header_map['최근조회일시'], value=current_time)
                            
                        wb.save(self.excel_path)
                        wb.close()
                        self.log("엑셀 업데이트가 성공적으로 완료되었습니다.")
                        break 
                    except PermissionError:
                        self.log("[경고] 엑셀 파일이 열려있습니다. 창을 닫아주세요. 5초 뒤 재시도...")
                        time.sleep(5.0)
                    except Exception as e:
                        self.log(f"[오류] 엑셀 저장 실패: {e}")
                        break

        except Exception as e:
            self.log(f"[메인 오류] {e}")
            traceback.print_exc()
        finally:
            self.cleanup_drivers()
            self.signals.finished.emit()


class NewsWorker(QThread):
    finished = pyqtSignal(list)
    error = pyqtSignal(str)

    def run(self):
        url = "https://www.dmktnews.com/kwa-home"
        try:
            headers = {"User-Agent": "Mozilla/5.0"}
            response = requests.get(url, headers=headers, timeout=10)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, 'html.parser')
            
            news_items = []
            for a in soup.find_all('a'):
                href = a.get('href', '')
                title_elem = a.find('div', class_='tit') or a.find('h3') or a.find('strong')
                title = title_elem.text.strip() if title_elem else a.text.strip()
                
                if "뉴스202" in title:
                    title = title.split("뉴스202")[0].strip()
                    
                if href.startswith('/news/') and title and len(title) > 5:
                    full_url = f"https://www.dmktnews.com{href}"
                    if not any(item['url'] == full_url for item in news_items):
                        news_items.append({'title': title, 'url': full_url})
                    if len(news_items) >= 9:
                        break
                        
            self.finished.emit(news_items)
        except Exception as e:
            self.error.emit(str(e))


class StatsScraperWorker(QThread):
    finished = pyqtSignal(dict)  # Emits parsed stats dict
    error = pyqtSignal(str)      # Emits error message
    status = pyqtSignal(str)     # Emits status log message

    def __init__(self, blog_id, company_name, global_driver_path):
        super().__init__()
        self.blog_id = blog_id
        self.company_name = company_name
        self.global_driver_path = global_driver_path
        self.driver = None

    def run(self):

        try:
            self.status.emit("크롬 드라이버 준비 중...")
            if not self.global_driver_path:
                try:
                    self.global_driver_path = ChromeDriverManager().install()
                except Exception as e:
                    self.error.emit(f"드라이버 설치 실패: {e}")
                    return
            
            self.status.emit("크롬 브라우저를 구동 중입니다...")
            from src.utils.driver_manager import setup_chrome_options
            options = setup_chrome_options()
            
            # Setup Chrome User Data Profile Directory for session persistence
            base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
            profile_dir = os.path.join(base_dir, "Data", "ChromeProfiles", self.blog_id)
            os.makedirs(profile_dir, exist_ok=True)
            
            
            service = Service(self.global_driver_path)
            self.driver = webdriver.Chrome(service=service, options=options)
            self.driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {"source": "Object.defineProperty(navigator, \'webdriver\', { get: () => undefined })"})
            
            self.status.emit("네이버 블로그 통계 페이지로 이동 중...")
            target_url = f"https://blog.stat.naver.com/blog/visitor/daily?blogId={self.blog_id}"
            self.driver.get(target_url)
            
            # Detect if we need to login
            login_wait_start = time.time()
            logged_in = False
            
            while time.time() - login_wait_start < 120:  # Wait up to 2 minutes
                try:
                    curr_url = self.driver.current_url
                except Exception:
                    # Browser closed by user
                    self.error.emit("사용자가 브라우저를 종료했거나 통신이 끊겼습니다.")
                    return
                
                if "nid.naver.com" in curr_url:
                    self.status.emit("로그인 및 2차 인증 대기 중... (브라우저에서 로그인해 주세요)")
                elif "blog.stat.naver.com" in curr_url:
                    # Logged in, wait for page loaded
                    try:
                        WebDriverWait(self.driver, 5).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "body"))
                        )
                        logged_in = True
                        break
                    except Exception:
                        pass
                
                time.sleep(2)
                
            if not logged_in:
                self.error.emit("로그인 대기 시간이 초과되었습니다.")
                self.cleanup()
                return
                
            self.status.emit("로그인 완료 감지! 블로그 통계 데이터를 수집 중...")
            
            # Fetch visitor, views, inflow data using JS fetch on the same origin
            end_date = QDate.currentDate().toString("yyyy-MM-dd")
            start_date = QDate.currentDate().addDays(-30).toString("yyyy-MM-dd")
            
            # 1. Daily Visitors
            js_visitor = f"""
            var callback = arguments[arguments.length - 1];
            fetch('/api/visitor/daily?blogId={self.blog_id}&startDate={start_date}&endDate={end_date}')
                .then(res => res.json())
                .then(data => callback(data))
                .catch(err => callback(null));
            """
            self.driver.set_script_timeout(10)
            res_visitor = self.driver.execute_async_script(js_visitor)
            
            # 2. Daily Views
            js_views = f"""
            var callback = arguments[arguments.length - 1];
            fetch('/api/views/daily?blogId={self.blog_id}&startDate={start_date}&endDate={end_date}')
                .then(res => res.json())
                .then(data => callback(data))
                .catch(err => callback(null));
            """
            res_views = self.driver.execute_async_script(js_views)
            
            # 3. Referrer / Inflow
            js_inflow = f"""
            var callback = arguments[arguments.length - 1];
            fetch('/api/inflow/daily?blogId={self.blog_id}&startDate={start_date}&endDate={end_date}')
                .then(res => res.json())
                .then(data => callback(data))
                .catch(err => callback(null));
            """
            res_inflow = self.driver.execute_async_script(js_inflow)
            
            # Parsing daily stats
            dates = []
            visitor_cnts = []
            views_cnts = []
            inflow_paths = []
            
            # Process visitor
            try:
                raw_list = res_visitor.get('result', {}).get('dailyVisitor', [])
                for item in raw_list:
                    dt = str(item.get('date', ''))
                    if len(dt) == 8:
                        formatted_dt = f"{dt[4:6]}/{dt[6:8]}"
                    else:
                        formatted_dt = dt
                    dates.append(formatted_dt)
                    visitor_cnts.append(item.get('visitor', 0))
            except Exception:
                pass
                
            # Process views
            try:
                raw_list = res_views.get('result', {}).get('dailyViews', [])
                for item in raw_list:
                    views_cnts.append(item.get('views', 0))
            except Exception:
                pass
                
            # Adjust arrays size to match
            min_len = min(len(dates), len(visitor_cnts), len(views_cnts))
            if min_len > 0:
                dates = dates[:min_len]
                visitor_cnts = visitor_cnts[:min_len]
                views_cnts = views_cnts[:min_len]
            
            # Process inflow paths
            try:
                raw_list = res_inflow.get('result', {}).get('inflow', [])
                for item in raw_list[:5]:  # Top 5
                    inflow_paths.append({
                        'name': item.get('source', '알 수 없음'),
                        'value': item.get('count', 0)
                    })
            except Exception:
                pass
                
            if not dates:
                self.error.emit("데이터를 파싱하지 못했거나 권한이 없습니다.")
                self.cleanup()
                return
                
            stats_data = {
                'last_updated': time.strftime("%Y-%m-%d %H:%M:%S"),
                'dates': dates,
                'visitor': visitor_cnts,
                'views': views_cnts,
                'inflow': inflow_paths
            }
            
            # Save to JSON file
            data_dir = os.path.join(base_dir, "Data")
            stats_path = os.path.join(data_dir, f"stats_{self.company_name}.json")
            with open(stats_path, 'w', encoding='utf-8') as f:
                json.dump(stats_data, f, ensure_ascii=False, indent=4)
                
            self.finished.emit(stats_data)
            
        except Exception as e:
            self.error.emit(f"오류가 발생했습니다: {e}")
        finally:
            self.cleanup()

    def cleanup(self):
        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass
            self.driver = None


class SpellCheckWorker(QThread):
    finished = pyqtSignal(dict)
    error = pyqtSignal(str)

    def __init__(self, text, parent=None):
        super().__init__(parent)
        self.text = text

    def run(self):

        try:
            result = check_text(self.text)
            self.finished.emit(result)
        except Exception as e:
            self.error.emit(str(e))


class PlaceScraperWorker(QThread):
    def __init__(self, keyword, target_company, display_count, global_driver_path, record_mode=False):
        super().__init__()
        self.keyword = keyword
        self.target_company = target_company
        self.display_count = display_count
        self.global_driver_path = global_driver_path
        self.record_mode = record_mode
        self.signals = WorkerSignals()
        self.driver = None
        self._is_stopped = False

    def stop(self):
        self._is_stopped = True
        self.cleanup_drivers()

    def log(self, message):
        self.signals.log.emit(message)

    def get_patched_chromedriver(self):
        import os, re
        
        if not self.global_driver_path or not os.path.exists(self.global_driver_path):
            return None
            
        driver_dir = os.path.dirname(self.global_driver_path)
        patched_path = os.path.join(driver_dir, "chromedriver_patched.exe")
        
        if os.path.exists(patched_path):
            return patched_path
            
        try:
            with open(self.global_driver_path, 'rb') as f:
                binary_data = f.read()
                
            patched_binary = re.sub(b'cdc_[a-zA-Z0-9]{22}_', b'rnd_' + b'X' * 22 + b'_', binary_data)
            
            with open(patched_path, 'wb') as f:
                f.write(patched_binary)
                
            return patched_path
        except Exception as e:
            self.log(f"[경고] 드라이버 패치 실패: {e} (원본 드라이버를 사용합니다.)")
            return self.global_driver_path

    def get_random_android_ua_and_size(self):
        import random
        pool = [
            ("Mozilla/5.0 (Linux; Android 14; SM-S928N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.6478.122 Mobile Safari/537.36", "412,915"),
            ("Mozilla/5.0 (Linux; Android 13; SM-A536N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.6422.113 Mobile Safari/537.36", "412,915"),
            ("Mozilla/5.0 (Linux; Android 14; SM-F731N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.6367.113 Mobile Safari/537.36", "360,844"),
            ("Mozilla/5.0 (Linux; Android 13; SM-G998N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.6312.118 Mobile Safari/537.36", "384,853"),
            ("Mozilla/5.0 (Linux; Android 14; SM-S918N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.6478.122 Mobile Safari/537.36", "412,915")
        ]
        return random.choice(pool)

    def cleanup_drivers(self):
        if hasattr(self, 'driver_hospital') and self.driver_hospital:
            try:
                self.driver_hospital.quit()
            except:
                pass
            self.driver_hospital = None
            
        if hasattr(self, 'driver_place') and self.driver_place:
            try:
                self.driver_place.quit()
            except:
                pass
            self.driver_place = None

    def get_place_data_by_url(self, driver, url, label, delay=0):
        if delay > 0:
            time.sleep(delay)
            
        titles = []
        
        self.log(f"▶ '{self.keyword}' - {label} 카테고리 데이터 탐색 중... ({self.display_count}개 목표)")
        
        try:
            driver.get(url)
            
            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "li"))
                )
            except TimeoutException:
                self.log(f"[안내] '{self.keyword}' - {label} 목록을 찾을 수 없거나 시간이 초과되었습니다.")
                return pd.DataFrame({'제목': pd.Series(dtype='str')})

            last_height = driver.execute_script("return document.body.scrollHeight")
            retries = 0
            while True:
                js_count = """
                var els = document.querySelectorAll("li");
                var organicCount = 0;
                for(var i=0; i<els.length; i++) {
                    var isAd = false;
                    
                    var adLinks = els[i].querySelectorAll("a");
                    for(var j=0; j<adLinks.length; j++) {
                        var href = adLinks[j].getAttribute("href") || "";
                        if(href.indexOf("ader.naver.com") !== -1 || href.indexOf("adcr.naver.com") !== -1 || href.indexOf("NSPM_70") !== -1) {
                            isAd = true;
                            break;
                        }
                    }
                    if(!isAd) {
                        var svgs = els[i].querySelectorAll("svg");
                        for(var j=0; j<svgs.length; j++) {
                            var vb = svgs[j].getAttribute("viewBox");
                            if(vb === "0 0 39 16" || vb === "0 0 38 16" || vb === "0 0 40 16") {
                                isAd = true;
                                break;
                            }
                        }
                    }
                    if(!isAd) {
                        var blinds = els[i].querySelectorAll(".place_blind");
                        for(var j=0; j<blinds.length; j++) {
                            if(blinds[j].textContent.trim() === "광고") {
                                isAd = true;
                                break;
                            }
                        }
                    }
                    
                    if (!isAd) {
                        var links = els[i].querySelectorAll("a");
                        var foundTitle = false;
                        for(var k=0; k<links.length; k++) {
                            var a_text = links[k].innerText || "";
                            var first_line = a_text.trim().split('\\n')[0].trim();
                            if(first_line.length > 0 && first_line.indexOf("리뷰 ") === -1 && isNaN(first_line)) {
                                foundTitle = true;
                                break;
                            }
                        }
                        if(foundTitle) {
                            organicCount++;
                        }
                    }
                }
                return organicCount;
                """
                current_count = driver.execute_script(js_count)
                
                if current_count >= self.display_count:
                    break
                    
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                
                # Smart Polling: 기다리지 않고 페이지 높이가 변할 때까지 최대 2초간 0.2초 간격으로 확인
                waited = 0
                new_height = last_height
                while waited < 2.0:
                    time.sleep(0.2)
                    waited += 0.2
                    new_height = driver.execute_script("return document.body.scrollHeight")
                    if new_height > last_height:
                        break
                
                if new_height == last_height:
                    retries += 1
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight - 500);")
                    time.sleep(0.1)
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(0.5)
                    new_height = driver.execute_script("return document.body.scrollHeight")
                    
                    if new_height == last_height:
                        if retries >= 2:  # 최대 재시도 2회로 단축
                            break
                else:
                    retries = 0
                last_height = new_height

            js_extract = """
            var results = [];
            var els = document.querySelectorAll("li");
            for(var i=0; i<els.length; i++) {
                var isAd = false;
                
                var adLinks = els[i].querySelectorAll("a");
                for(var j=0; j<adLinks.length; j++) {
                    var href = adLinks[j].getAttribute("href") || "";
                    if(href.indexOf("ader.naver.com") !== -1 || href.indexOf("adcr.naver.com") !== -1 || href.indexOf("NSPM_70") !== -1) {
                        isAd = true;
                        break;
                    }
                }
                if(!isAd) {
                    var svgs = els[i].querySelectorAll("svg");
                    for(var j=0; j<svgs.length; j++) {
                        var vb = svgs[j].getAttribute("viewBox");
                        if(vb === "0 0 39 16" || vb === "0 0 38 16" || vb === "0 0 40 16") {
                            isAd = true;
                            break;
                        }
                    }
                }
                if(!isAd) {
                    var blinds = els[i].querySelectorAll(".place_blind");
                    for(var j=0; j<blinds.length; j++) {
                        if(blinds[j].textContent.trim() === "광고") {
                            isAd = true;
                            break;
                        }
                    }
                }
                
                if(isAd) continue;
                
                var links = els[i].querySelectorAll("a");
                var title = "";
                for(var k=0; k<links.length; k++) {
                    var a_text = links[k].innerText || "";
                    var first_line = a_text.trim().split('\\n')[0].trim();
                    if(first_line.length > 0 && first_line.indexOf("이미지 수") === -1 && isNaN(first_line)) {
                        title = first_line;
                        break;
                    }
                }
                if(title.length > 0) {
                    results.push(title);
                }
            }
            return results;
            """
            extracted_data = driver.execute_script(js_extract)
            
            for data in extracted_data[:self.display_count]:
                titles.append(data)
                
            if not titles:
                self.log(f"[안내] '{self.keyword}' - 데이터를 추출하지 못했습니다.")
                        
        except Exception as e:
            self.log(f"[오류] '{self.keyword}' - 크롤링 실패: {e}")
                
        df = pd.DataFrame({'제목': titles})
        df.index = df.index + 1
        return df

    def get_rank_string(self, df_result, target_company):
        if df_result.empty or not target_company:
            return "검색/분석 불가"
        
        valid_target = target_company.strip()
        if not valid_target:
            return "미노출 (범위 밖)"
            
        pattern = re.escape(valid_target)
        mask = df_result['제목'].str.contains(pattern, na=False, regex=True)
        found_indices = df_result.index[mask].tolist()
        if found_indices:
            return ", ".join([f"{i}위" for i in found_indices])
        return "미노출 (범위 밖)"

    def write_to_sheet(self, ranks):
        if not self.record_mode:
            return
            
        try:
            import gspread
            from oauth2client.service_account import ServiceAccountCredentials
            import os
            import datetime
            import time
            from src.config import CREDENTIALS_PATH
            from gspread.utils import rowcol_to_a1
            
            if not os.path.exists(CREDENTIALS_PATH):
                self.log("[기록 모드 오류] 자격증명 파일(credentials.json)이 존재하지 않아 기록을 생략합니다.")
                return
                
            self.log("[기록 모드] 구글 스프레드시트 업데이트를 시작합니다...")
            
            scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
            creds = ServiceAccountCredentials.from_json_keyfile_name(CREDENTIALS_PATH, scope)
            client = gspread.authorize(creds)
            
            sheet = client.open_by_key('17V0qlJg4hmpo1rLS5oAq9UPkFO6kQ1iMnovERV9PTOs').worksheet('Raw data')
            
            row1 = sheet.row_values(1)
            target_col = -1
            
            clean_target = self.target_company.replace(" ", "")
            for i, val in enumerate(row1):
                if val:
                    clean_val = val.replace(" ", "")
                    if clean_target in clean_val or clean_val in clean_target:
                        target_col = i + 1
                        break
                        
            if target_col == -1:
                target_col = ((len(row1) + 3) // 4) * 4 + 1
                if len(row1) == 0:
                    target_col = 1
                    
                title_range = f"{rowcol_to_a1(1, target_col)}:{rowcol_to_a1(1, target_col+2)}"
                date_cell = rowcol_to_a1(2, target_col)
                keyword_cell = rowcol_to_a1(2, target_col+1)
                rank_cell = rowcol_to_a1(2, target_col+2)
                
                sheet.update_cell(1, target_col, self.target_company)
                sheet.update_cell(2, target_col, "날짜")
                sheet.update_cell(2, target_col+1, "키워드")
                sheet.update_cell(2, target_col+2, "순위")
                
                try:
                    sheet.format(title_range, {
                        "textFormat": {"bold": True, "fontSize": 10},
                        "horizontalAlignment": "CENTER",
                    })
                    sheet.format(date_cell, {
                        "backgroundColor": {"red": 0.85, "green": 0.95, "blue": 0.85},
                        "horizontalAlignment": "CENTER",
                    })
                    sheet.format(keyword_cell, {
                        "backgroundColor": {"red": 0.85, "green": 0.9, "blue": 1.0},
                        "horizontalAlignment": "CENTER",
                    })
                    sheet.format(rank_cell, {
                        "backgroundColor": {"red": 1.0, "green": 0.85, "blue": 0.85},
                        "horizontalAlignment": "CENTER",
                    })
                    sheet.merge_cells(title_range)
                except Exception as fe:
                    self.log(f"[기록 모드 서식 오류] {fe}")
            
            today_str = datetime.date.today().strftime("%Y-%m-%d")
            
            if ranks and len(ranks) > 0:
                rank_val = str(ranks[0])
            else:
                rank_val = "미노출"
                
            data_to_write = [today_str, self.keyword, rank_val]
            max_retries = 5
            
            for attempt in range(max_retries):
                col_vals = sheet.col_values(target_col)
                next_row = len(col_vals) + 1
                if next_row < 3:
                    next_row = 3
                
                range_to_update = f"{rowcol_to_a1(next_row, target_col)}:{rowcol_to_a1(next_row, target_col+2)}"
                sheet.update([data_to_write], range_to_update)
                
                time.sleep(1)
                verify_vals = sheet.get(range_to_update)
                
                if verify_vals and len(verify_vals) > 0 and verify_vals[0] == data_to_write:
                    self.log(f"[기록 모드 완료] 행 {next_row}에 안전하게 기록되었습니다.")
                    return
                else:
                    self.log(f"[기록 모드 재시도] 행 {next_row}에 동시 기입 충돌 감지! 재시도 중... ({attempt+1}/{max_retries})")
                    
            self.log("[기록 모드 실패] 동시 기입 충돌로 인해 기록에 실패했습니다.")
            
        except Exception as e:
            self.log(f"[기록 모드 치명적 오류] {e}")



    def run(self):

        has_error = False
        error_msg = ""
        is_early_return = False
        try:
            if not self.keyword.strip() or not self.target_company.strip():
                error_msg = "키워드와 목표 업체명을 모두 입력해주세요."
                has_error = True
                is_early_return = True
                return

            self.log(f"=== 키워드: '{self.keyword}', 목표 업체명: '{self.target_company}' ===")
            
            delay_time = random.uniform(0.3, 0.7)
            self.log(f"※ 봇 탐지 우회를 위해 {delay_time:.1f}초 대기 중...")
            time.sleep(delay_time)
            
            if getattr(self, '_is_stopped', False):
                raise Exception("작업이 사용자에 의해 취소되었습니다.")
                
            patched_driver_path = self.get_patched_chromedriver()
            if not patched_driver_path:
                error_msg = "크롬 드라이버를 찾을 수 없거나 패치할 수 없습니다."
                has_error = True
                is_early_return = True
                return

            # 공통 크롬 옵션 (봇 탐지 회피용)
            from src.utils.driver_manager import setup_chrome_options
            options = setup_chrome_options()
            options.page_load_strategy = 'eager'
            # 백그라운드 숨김 처리 옵션 복원 (테스트 완료)
            
            # CDP를 이용해 자바스크립트 변수 은닉 (navigator.webdriver 지우기)
            cdp_script = "Object.defineProperty(navigator, 'webdriver', { get: () => undefined })"

            # ---------------------------------------------
            # 첫 번째 드라이버 (병원) 세팅
            # ---------------------------------------------
            ua_h, size_h = self.get_random_android_ua_and_size()
            options_h = Options()
            for arg in options.arguments:
                options_h.add_argument(arg)
            for key, val in options.experimental_options.items():
                options_h.add_experimental_option(key, val)
                
            options_h.add_argument(f"user-agent={ua_h}")
            options_h.add_argument(f"--window-size={size_h}")
            
            service_h = Service(patched_driver_path)
            self.driver_hospital = webdriver.Chrome(service=service_h, options=options_h)
            self.driver_hospital.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {'source': cdp_script})
            
            # 봇 탐지 방지 지연
            delay_between_browsers = random.uniform(0.2, 0.5)
            self.log(f"※ 브라우저 연속 실행에 따른 봇 탐지 방지를 위해 {delay_between_browsers:.1f}초 대기 중...")
            time.sleep(delay_between_browsers)
            
            # ---------------------------------------------
            # 두 번째 드라이버 (일반 플레이스) 세팅
            # ---------------------------------------------
            ua_p, size_p = self.get_random_android_ua_and_size()
            options_p = Options()
            for arg in options.arguments:
                options_p.add_argument(arg)
            for key, val in options.experimental_options.items():
                options_p.add_experimental_option(key, val)
                
            options_p.add_argument(f"user-agent={ua_p}")
            options_p.add_argument(f"--window-size={size_p}")
            
            service_p = Service(patched_driver_path)
            self.driver_place = webdriver.Chrome(service=service_p, options=options_p)
            self.driver_place.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {'source': cdp_script})
            
            import urllib.parse
            import concurrent.futures
            
            departments = [
                "이비인후과", "마취통증의학과", "정신건강의학과", "소아청소년과", 
                "재활의학과", "가정의학과", "성형외과", "정형외과", "신경외과", 
                "피부과", "산부인과", "비뇨기과", "통증의학과", "신경과", 
                "안과", "치과", "내과", "외과", "소아과", "한의원", "의원", "한방병원", "병원"
            ]
            dept = self.keyword
            for d in departments:
                if d in self.keyword:
                    dept = d
                    break
            
            q = urllib.parse.quote(self.keyword)
            d_q = urllib.parse.quote(dept)
            
            url_hospital = f"https://m.place.naver.com/hospital/list?query={q}&department={d_q}&x=126.9783882&y=37.5666103&level=top&entry=pll&originalQuery={q}"
            url_place = f"https://m.place.naver.com/place/list?query={q}&x=126.9783882&y=37.5666103&start=1&display=100&adult=false&deviceType=mobile&sessionId=dxF1wii6ikGt4t96FvdHJS4k&level=top&entry=pll"

            df_sim = pd.DataFrame({'제목': pd.Series(dtype='str')})
            
            with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
                future_to_label = {
                    executor.submit(self.get_place_data_by_url, self.driver_hospital, url_hospital, "병원", 0): 'hospital',
                    executor.submit(self.get_place_data_by_url, self.driver_place, url_place, "일반 플레이스", 0.5): 'place'
                }
                
                results = {}
                for future in concurrent.futures.as_completed(future_to_label):
                    label = future_to_label[future]
                    try:
                        res_df = future.result()
                        if not res_df.empty:
                            results[label] = res_df
                    except Exception as e:
                        pass
                
                if 'hospital' in results and not results['hospital'].empty:
                    df_sim = results['hospital']
                    self.log(f"[안내] 병원 카테고리 검색 결과를 최종 채택합니다.")
                elif 'place' in results and not results['place'].empty:
                    df_sim = results['place']
                    self.log(f"[안내] 병원 카테고리 누락으로, 일반 플레이스 검색 결과를 최종 채택합니다.")
                else:
                    self.log(f"[안내] '{self.keyword}' - 병원 및 일반 장소 모두에서 검색 결과를 찾지 못했습니다.")

            sim_rank = self.get_rank_string(df_sim, self.target_company)
            
            ranks = []
            if sim_rank and "위" in sim_rank:
                for part in sim_rank.split(","):
                    if "위" in part:
                        try:
                            r_val = int(part.replace("위", "").strip())
                            ranks.append(r_val)
                        except:
                            pass
                            
            if ranks:
                self.signals.match_found.emit({
                    'company': self.target_company,
                    'keyword': self.keyword,
                    'ranks': ranks,
                    'screenshots': [],
                    'folder': ''
                })
            
            result_text = f"<< '{self.keyword}' 검색 결과 >>\n"
            for idx, row in df_sim.iterrows():
                result_text += f"[{idx}위] {row['제목']}\n"
            self.log(result_text)
            self.log(f"▶ 기록 준비 완료: {sim_rank}")
            
            if self.record_mode:
                self.write_to_sheet(ranks)

        except Exception as e:
            if getattr(self, '_is_stopped', False):
                error_msg = "작업이 사용자에 의해 취소되었습니다."
                self.log(f"[안내] {error_msg}")
            else:
                self.log(f"[메인 오류] {e}")
                traceback.print_exc()
                error_msg = str(e)
            has_error = True
        finally:
            self.cleanup_drivers()
            if has_error:
                self.signals.error.emit(error_msg)
            elif not is_early_return:
                self.signals.finished.emit()

