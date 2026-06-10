import subprocess
import shutil
from openpyxl.styles import PatternFill, Font, Alignment
import openpyxl
import pandas as pd
import configparser

import os
import sys
import json
import time
import datetime
from datetime import date
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QObject, QUrl, QDate, QPropertyAnimation, QEasingCurve, QRect, QPoint, QMargins, QTimer
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QMessageBox, QDialog, QFrame, QLabel, QStackedWidget, QGraphicsDropShadowEffect, QCalendarWidget, QPushButton, QFileDialog, QSizePolicy, QGridLayout
from PyQt5.QtGui import QFont, QFontDatabase, QDesktopServices, QTextCharFormat, QColor, QBrush, QPainter, QCursor, QPixmap
from qfluentwidgets import (PushButton, PrimaryPushButton, ComboBox, SpinBox, SwitchButton, TextEdit, 
                            setTheme, Theme, TitleLabel, SubtitleLabel, InfoBar, InfoBarPosition,
                            IndeterminateProgressRing, FluentWindow, FluentIcon, LineEdit,
                            TransparentToolButton, ScrollArea, CardWidget, MessageBox,
                            setThemeColor, NavigationItemPosition, qconfig, isDarkTheme,
                            BodyLabel, IconWidget, HyperlinkButton, PasswordLineEdit, CheckBox, NavigationPushButton, RoundMenu, Action)
from qfluentwidgets.common.icon import drawIcon
from src.core.scraper_threads import ScraperWorker, DriverInitWorker
from src.ui.components.cards import ExposureCard
from src.config import SESSION, WORKSPACE_DIR, ASSETS_DIR, DATA_DIR, FONT_DIR, SETTINGS_PATH, CREDENTIALS_PATH

class ScraperInterface(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent=parent)
        self.setObjectName("ScraperInterface")
        
        self.base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
        
        self.workspace_dir = os.path.join(self.base_dir, 'Workspace')
        os.makedirs(self.workspace_dir, exist_ok=True)
        
        self.excel_filename = "키워드_순위_작업표.xlsx"
        self.excel_path = os.path.join(self.workspace_dir, self.excel_filename)
        
        self.data_dir = os.path.join(self.base_dir, 'Data')
        self.config_file = os.path.join(self.data_dir, 'config.ini')
        
        self.config = configparser.ConfigParser()
        self.global_driver_path = ""
        
        self.init_ui()
        self.check_environment()

    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(36, 36, 36, 36)
        main_layout.setSpacing(16)

        self.title_label = TitleLabel("블로그 순위 체크")
        main_layout.addWidget(self.title_label)

        # Horizontal Split Layout
        split_layout = QHBoxLayout()
        split_layout.setSpacing(24)
        
        # Left Panel (Existing UI controls)
        self.left_panel = QWidget(self)
        left_layout = QVBoxLayout(self.left_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(16)

        control_layout = QHBoxLayout()
        
        self.sheet_combo = ComboBox()
        self.sheet_combo.setMinimumWidth(200)
        control_layout.addWidget(SubtitleLabel("작업 시트 선택:"))
        control_layout.addWidget(self.sheet_combo)
        
        self.count_spinbox = SpinBox()
        self.count_spinbox.setRange(1, 100)
        control_layout.addStretch(1)
        control_layout.addWidget(SubtitleLabel("탐색 목표 개수:"))
        control_layout.addWidget(self.count_spinbox)
        
        self.screenshot_switch = SwitchButton()
        self.screenshot_switch.setOnText("캡처 켜짐")
        self.screenshot_switch.setOffText("캡처 꺼짐")
        font = QFont("SUIT", 10, QFont.Bold)
        self.screenshot_switch.setFont(font)
        if hasattr(self.screenshot_switch, 'label'):
            self.screenshot_switch.label.setFont(font)
            self.screenshot_switch.label.setStyleSheet("font-family: 'SUIT'; font-weight: bold;")
        else:
            self.screenshot_switch.setStyleSheet("QLabel { font-family: 'SUIT'; font-weight: bold; }")
        control_layout.addStretch(1)
        control_layout.addWidget(self.screenshot_switch)
        
        left_layout.addLayout(control_layout)

        self.loading_container = QWidget()
        loading_layout = QHBoxLayout(self.loading_container)
        loading_layout.setAlignment(Qt.AlignCenter)
        loading_layout.setContentsMargins(0, 0, 0, 0)
        
        self.loading_ring = IndeterminateProgressRing()
        self.loading_ring.setFixedSize(25, 25)
        self.loading_label = SubtitleLabel("크롬 드라이버를 백그라운드에서 점검/동기화 중입니다...")
        
        loading_layout.addWidget(self.loading_ring)
        loading_layout.addWidget(self.loading_label)
        left_layout.addWidget(self.loading_container)
        self.loading_container.hide()

        button_layout = QHBoxLayout()
        
        self.start_btn = PushButton("순위 체크 시작")
        self.start_btn.setStyleSheet("""
            QPushButton {
                background-color: #3CA0F0;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-family: 'SUIT';
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #59B4FF;
            }
            QPushButton:pressed {
                background-color: #268CD9;
            }
            QPushButton:disabled {
                background-color: #2C2C2C;
                color: #666666;
            }
        """)
        self.start_btn.clicked.connect(self.start_scraping)
        
        self.open_excel_btn = PushButton("엑셀 파일 열기")
        self.open_excel_btn.setStyleSheet("""
            QPushButton {
                background-color: #107C41;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-family: 'SUIT';
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1F9A55;
            }
            QPushButton:pressed {
                background-color: #0E6233;
            }
            QPushButton:disabled {
                background-color: #2C2C2C;
                color: #666666;
            }
        """)
        self.open_excel_btn.clicked.connect(self.open_excel_file)
        
        button_layout.addWidget(self.start_btn)
        button_layout.addWidget(self.open_excel_btn)
        
        left_layout.addLayout(button_layout)

        self.console_output = TextEdit()
        self.console_output.setReadOnly(True)
        left_layout.addWidget(self.console_output)

        split_layout.addWidget(self.left_panel, 3) # stretch factor 3

        # Right Panel (Exposure Status Panel)
        self.right_panel = QWidget(self)
        right_layout = QVBoxLayout(self.right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(12)
        
        right_title = SubtitleLabel("실시간 노출 현황")
        right_title.setFont(QFont("SUIT", 12, QFont.Bold))
        right_layout.addWidget(right_title)
        
        self.exposure_scroll = ScrollArea(self.right_panel)
        self.exposure_scroll.setWidgetResizable(True)
        
        self.exposure_content = QWidget()
        self.exposure_content.setObjectName("ExposureContent")
        self.exposure_content.setStyleSheet("QWidget#ExposureContent { background-color: transparent; }")
        self.exposure_layout = QVBoxLayout(self.exposure_content)
        self.exposure_layout.setContentsMargins(12, 12, 12, 12)
        self.exposure_layout.setSpacing(12)
        self.exposure_layout.addStretch(1)
        
        self.exposure_scroll.setWidget(self.exposure_content)
        right_layout.addWidget(self.exposure_scroll)
        
        split_layout.addWidget(self.right_panel, 2) # stretch factor 2

        main_layout.addLayout(split_layout)

        # Connect to theme changed for right panel styling
        self.update_right_panel_style()
        qconfig.themeChanged.connect(self.update_right_panel_style)

    def update_right_panel_style(self):
        is_dark = isDarkTheme()
        border_color = "#3A3A3A" if is_dark else "#E5E5E5"
        bg_color = "#202020" if is_dark else "#FFFFFF"
        self.exposure_scroll.setStyleSheet(f"QScrollArea {{ border: 1px solid {border_color}; border-radius: 8px; background-color: {bg_color}; }}")

    def clear_exposure_cards(self):
        while self.exposure_layout.count():
            child = self.exposure_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()
        self.exposure_layout.addStretch(1)

    def add_exposure_card(self, match_data):
        card = ExposureCard(match_data, self.exposure_content)
        insert_idx = max(0, self.exposure_layout.count() - 1)
        self.exposure_layout.insertWidget(insert_idx, card)

    def open_excel_file(self):
        if not os.path.exists(self.excel_path):
            InfoBar.error("오류", "엑셀 파일이 아직 생성되지 않았습니다.", duration=3000, position=InfoBarPosition.TOP, parent=self)
            return
            
        msg_box = QMessageBox(self)
        msg_box.setIcon(QMessageBox.Warning)
        msg_box.setWindowTitle("엑셀 파일 수정 주의사항")
        msg_box.setText("작업 시트를 수정하신 후, 반드시 엑셀 프로그램에서 '저장(Ctrl+S)'을 누른 다음 창을 닫고 크롤링을 시작해주세요.")
        msg_box.setInformativeText("저장하지 않고 닫거나 엑셀 창을 켜둔 채로 크롤링을 시작하면, 이전 데이터가 불러와지거나 저장 시 권한 오류가 발생합니다.")
        msg_box.setStandardButtons(QMessageBox.Ok)
        msg_box.exec_()
            
        try:
            if os.name == 'nt':
                os.startfile(self.excel_path)
            elif sys.platform == 'darwin':
                subprocess.call(['open', self.excel_path])
            else:
                subprocess.call(['xdg-open', self.excel_path])
            self.append_log("[안내] 엑셀 파일을 열었습니다. 작업 후 반드시 저장하고 엑셀 창을 닫아주세요.")
        except Exception as e:
            self.append_log(f"[오류] 엑셀 파일을 열 수 없습니다: {e}")

    def append_log(self, text):
        self.console_output.append(text)
        scrollbar = self.console_output.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    def check_environment(self):
        self.append_log("[시스템] 초기 환경을 점검합니다...")
        
        if os.path.exists(self.config_file):
            self.config.read(self.config_file, encoding='utf-8')
            display_count = int(self.config['SETTINGS'].get('DISPLAY_COUNT', 10))
            capture_str = self.config['SETTINGS'].get('CAPTURE_SCREENSHOT', 'False').strip().lower()
            capture_option = capture_str in ['true', '1', 'y', 'yes', 't']
        else:
            display_count = 10
            capture_option = False
            self.config['SETTINGS'] = {'DISPLAY_COUNT': '10', 'CAPTURE_SCREENSHOT': 'False'}
            os.makedirs(self.data_dir, exist_ok=True)
            with open(self.config_file, 'w', encoding='utf-8') as f:
                self.config.write(f)
                
        self.count_spinbox.setValue(display_count)
        self.screenshot_switch.setChecked(capture_option)

        if not os.path.exists(self.excel_path):
            df_template = pd.DataFrame(columns=['키워드', '목표블로그', '관련도순_순위', '최근조회일시'])
            with pd.ExcelWriter(self.excel_path, engine='openpyxl') as writer:
                df_template.to_excel(writer, index=False, sheet_name="기본작업표")
            self.format_excel_file(self.excel_path, "기본작업표")
            self.append_log(f"[시스템] 'Workspace/{self.excel_filename}' 템플릿 파일이 생성되었습니다.")
        
        self.load_excel_sheets()

        self.start_btn.setEnabled(False)
        self.open_excel_btn.setEnabled(False)
        self.loading_container.show()
        self.loading_ring.start()
        
        self.driver_worker = DriverInitWorker()
        self.driver_worker.finished.connect(self.on_driver_ready)
        self.driver_worker.error.connect(self.on_driver_error)
        self.driver_worker.start()

    def on_driver_ready(self, path):
        self.global_driver_path = path
        self.loading_ring.stop()
        self.loading_container.hide()
        self.append_log("[시스템] 크롬 드라이버 동기화 및 준비 완료.")
        self.start_btn.setEnabled(True)
        self.open_excel_btn.setEnabled(True)

    def on_driver_error(self, err_msg):
        self.loading_ring.stop()
        self.loading_container.hide()
        self.append_log(f"[치명적 오류] 크롬 드라이버 설치 실패: {err_msg}")
        InfoBar.error("오류", "크롬 드라이버를 설치할 수 없습니다.", duration=5000, position=InfoBarPosition.TOP, parent=self)

    def load_excel_sheets(self):
        try:
            wb = openpyxl.load_workbook(self.excel_path, read_only=True)
            for name in wb.sheetnames:
                self.sheet_combo.addItem(name)
            wb.close()
        except Exception as e:
            self.append_log(f"[오류] 엑셀 파일을 읽을 수 없습니다 (파일이 열려있는지 확인): {e}")

    def format_excel_file(self, file_path, sheet_name):
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 15
        wb.save(file_path)

    def backup_excel(self):
        try:
            backup_dir = os.path.join(self.workspace_dir, '.backup')
            os.makedirs(backup_dir, exist_ok=True)
            if os.name == 'nt': os.system(f'attrib +h "{backup_dir}"')
            
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            backup_path = os.path.join(backup_dir, f"키워드_순위_작업표_backup_{timestamp}.xlsx")
            shutil.copy2(self.excel_path, backup_path)
            
            now = time.time()
            deleted = 0
            for file_name in os.listdir(backup_dir):
                file_path = os.path.join(backup_dir, file_name)
                if os.path.isfile(file_path) and file_name.endswith('.xlsx'):
                    if now - os.path.getmtime(file_path) > 30 * 24 * 60 * 60:
                        os.remove(file_path)
                        deleted += 1
            if deleted > 0: self.append_log(f"[시스템] 30일 경과 백업 파일 {deleted}개 삭제 완료.")
        except Exception:
            pass

    def start_scraping(self):
        selected_sheet = self.sheet_combo.currentText()
        if not selected_sheet:
            InfoBar.error("오류", "작업할 시트를 선택해주세요.", duration=3000, position=InfoBarPosition.TOP, parent=self)
            return

        if os.path.exists(self.excel_path):
            try:
                with open(self.excel_path, 'a') as f:
                    pass
            except PermissionError:
                self.append_log("[오류] 엑셀 파일이 열려있어 작업을 시작할 수 없습니다. 창을 닫아주세요.")
                InfoBar.error("작업 거부", "엑셀 파일이 열려있습니다. 창을 닫은 후 다시 시도해주세요.", duration=5000, position=InfoBarPosition.TOP, parent=self)
                return
            except Exception as e:
                self.append_log(f"[오류] 엑셀 파일 접근 검사 실패: {e}")
                return

        self.backup_excel()

        self.start_btn.setEnabled(False)
        self.open_excel_btn.setEnabled(False)
        self.sheet_combo.setEnabled(False)
        self.console_output.clear()
        self.clear_exposure_cards()
        
        display_count = self.count_spinbox.value()
        capture_screenshot = self.screenshot_switch.isChecked()
        is_custom_excel = selected_sheet != "기본작업표"

        self.config['SETTINGS'] = {'DISPLAY_COUNT': str(display_count), 'CAPTURE_SCREENSHOT': str(capture_screenshot)}
        with open(self.config_file, 'w', encoding='utf-8') as f:
            self.config.write(f)

        self.worker = ScraperWorker(
            self.excel_path, selected_sheet, display_count, 
            capture_screenshot, self.global_driver_path, is_custom_excel
        )
        self.worker.signals.log.connect(self.append_log)
        self.worker.signals.error.connect(self.show_error)
        self.worker.signals.match_found.connect(self.add_exposure_card)
        self.worker.signals.finished.connect(self.on_scraping_finished)
        self.worker.start()

    def show_error(self, err_msg):
        InfoBar.error("작업 중단", err_msg, duration=5000, position=InfoBarPosition.TOP, parent=self)
        self.on_scraping_finished()

    def on_scraping_finished(self):
        self.start_btn.setEnabled(True)
        self.open_excel_btn.setEnabled(True)
        self.sheet_combo.setEnabled(True)
        self.append_log("\n[안내] 모든 루틴이 종료되었습니다.")
        InfoBar.success("완료", "크롤링 작업이 성공적으로 종료되었습니다.", duration=4000, position=InfoBarPosition.TOP, parent=self)

