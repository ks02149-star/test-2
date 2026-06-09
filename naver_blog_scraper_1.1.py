import sys
import traceback
import subprocess
import os

__version__ = "1.0"

# --- [추가] 시스템 입출력 스트림 및 Qt 내부 경고 강제 음소거 ---
# QFluentWidgets의 프로모션 텍스트 및 PyQt5의 폰트 열거 경고(qt.qpa.fonts) 출력 차단
# --noconsole 빌드 시 출력 스트림 부재로 인한 강제 종료(Crash) 방어
try:
    if sys.stdout is None or sys.stderr is None:
        pass
    else:
        sys.stdout = open(os.devnull, 'w')
        sys.stderr = open(os.devnull, 'w')
except Exception:
    pass

os.environ['QT_LOGGING_RULES'] = '*.debug=false;qt.*.warning=false'

# --- GUI 환경 전용 치명적 오류 추적기 ---
def custom_excepthook(exc_type, exc_value, exc_tb):
    error_msg = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
    
    # 1. 오류 내용을 Workspace 디렉토리 내에 강제 저장
    try:
        base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
        workspace_dir = os.path.join(base_dir, "Workspace")
        os.makedirs(workspace_dir, exist_ok=True)
        log_path = os.path.join(workspace_dir, "error_log.txt")
        
        with open(log_path, "w", encoding="utf-8") as f:
            f.write(error_msg)
    except:
        pass
        
    # 2. 콘솔의 input() 대신 GUI 팝업창(QMessageBox)으로 오류 출력
    try:
        from PyQt5.QtWidgets import QApplication, QMessageBox
        if QApplication.instance():
            msg_box = QMessageBox()
            msg_box.setIcon(QMessageBox.Critical)
            msg_box.setWindowTitle("치명적 오류 발생")
            msg_box.setText("프로그램 실행 중 오류가 발생하여 종료됩니다.\n'Workspace' 폴더에 생성된 'error_log.txt'를 확인해주세요.")
            msg_box.setDetailedText(error_msg)
            msg_box.exec_()
    except:
        pass

sys.excepthook = custom_excepthook

def install_required_packages():
    if getattr(sys, 'frozen', False):
        return
        
    packages = {
        "pandas": "pandas", 
        "selenium": "selenium", 
        "webdriver-manager": "webdriver_manager", 
        "openpyxl": "openpyxl",
        "PyQt5": "PyQt5",
        "PyQtWebEngine": "PyQt5.QtWebEngineWidgets",
        "PyQt-Fluent-Widgets": "qfluentwidgets"
    }
    if os.name == 'nt':
        packages["pywin32"] = "win32api"
        
    for pip_name, import_name in packages.items():
        try:
            __import__(import_name)
        except ImportError:
            subprocess.check_call([sys.executable, "-m", "pip", "install", pip_name])

install_required_packages()

import urllib.parse
import pandas as pd
import shutil
import configparser
import time
import re
import random
import hashlib
import ctypes
import tempfile
import concurrent.futures
import threading

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from selenium import webdriver
import selenium.webdriver.chrome.webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager

os.environ['WDM_LOG'] = '0'
os.environ['WDM_LOG_LEVEL'] = '0'

import json
import winreg
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QObject, QUrl, QDate
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QMessageBox, QDialog, QFrame, QLabel, QCalendarWidget, QPushButton, QFileDialog, QSizePolicy
from PyQt5.QtGui import QFont, QFontDatabase, QDesktopServices, QTextCharFormat, QColor, QBrush
try:
    from PyQt5.QtWebEngineWidgets import QWebEngineView
except ImportError:
    QWebEngineView = None
from qfluentwidgets import (PushButton, PrimaryPushButton, ComboBox, SpinBox, SwitchButton, TextEdit, 
                            setTheme, Theme, TitleLabel, SubtitleLabel, InfoBar, InfoBarPosition,
                            IndeterminateProgressRing, FluentWindow, FluentIcon, LineEdit,
                            TransparentToolButton, ScrollArea, CardWidget, MessageBox,
                            setThemeColor, NavigationItemPosition, qconfig, isDarkTheme,
                            BodyLabel, IconWidget, HyperlinkButton)
import datetime
import requests
from bs4 import BeautifulSoup
from PyQt5.QtWidgets import QGridLayout

# --- [신설] macOS 스타일 내비게이션 바 하이라이트 패치 ---
from PyQt5.QtCore import QRect, QRectF, QPoint, QMargins
from PyQt5.QtGui import QPainter, QColor, QCursor, QPixmap
from qfluentwidgets import NavigationPushButton
from qfluentwidgets.common.icon import drawIcon

def patched_navigation_paintEvent(self, e):
    painter = QPainter(self)
    painter.setRenderHints(QPainter.Antialiasing | QPainter.TextAntialiasing | QPainter.SmoothPixmapTransform)
    painter.setPen(Qt.NoPen)

    if self.isPressed:
        painter.setOpacity(0.7)
    if not self.isEnabled():
        painter.setOpacity(0.4)

    # 배경 그리기
    c = 255 if isDarkTheme() else 0
    m = self._margins()
    pl, pr = m.left(), m.right()
    globalRect = QRect(self.mapToGlobal(QPoint()), self.size())

    # macOS 스타일로 선택 시 파란색 단색 배경 적용, 미선택 마우스 호버 시 옅은 흰색 배경 적용
    if self._canDrawIndicator():
        # 선택 상태: 선명한 파란색 둥근 배경 (모서리 반경 6px) 및 좌우/상하 여백 조정
        painter.setBrush(QColor(10, 132, 255))  # macOS Active Blue
        painter.drawRoundedRect(self.rect().adjusted(6, 2, -6, -2), 6, 6)
    elif ((self.isEnter and globalRect.contains(QCursor.pos())) or self.isAboutSelected) and self.isEnabled():
        # 호버/진입 상태: 옅은 반투명 회색 배경
        painter.setBrush(QColor(c, c, c, 12))
        painter.drawRoundedRect(self.rect().adjusted(6, 2, -6, -2), 6, 6)

    # 아이콘 그리기
    drawIcon(self._icon, painter, QRectF(11.5 + pl, 10, 16, 16))

    # 텍스트 그리기
    if self.isCompacted:
        return

    painter.setFont(self.font())
    
    # 선택된 상태에서는 텍스트를 순수 흰색(#FFFFFF)으로 강제 적용, 미선택 상태에서는 기존 텍스트 컬러
    if self._canDrawIndicator():
        painter.setPen(QColor(255, 255, 255))
    else:
        painter.setPen(self.textColor())

    left = 44 + pl if not self.icon().isNull() else pl + 16
    painter.drawText(QRect(left, 0, self.width() - left - pr, self.height()), Qt.AlignVCenter | Qt.AlignLeft, self.text())

# 패치 적용
NavigationPushButton.paintEvent = patched_navigation_paintEvent


# --- [신설] 부산대 한국어 맞춤법 검사기 API 연동 모듈 ---
from dataclasses import asdict, dataclass
from html import unescape
from typing import Callable
import urllib.error
import urllib.request

DEFAULT_RESULTS_URL = "https://nara-speller.co.kr/old_speller/results"
DEFAULT_MAX_CHARS = 1500
DEFAULT_TIMEOUT = 30
DEFAULT_THROTTLE_SECONDS = 1.2
RESULT_PAYLOAD_PATTERN = re.compile(r"data\s*=\s*(\[[\s\S]*?\]);\s*pageIdx\s*=")
NO_ISSUES_PATTERN = re.compile(r"맞춤법과\s*문법\s*오류를\s*찾지\s*못했습니다", re.MULTILINE)
TAG_PATTERN = re.compile(r"<[^>]+>")
LINE_BREAK_PATTERN = re.compile(r"<br\s*/?>", re.IGNORECASE)
SENTENCE_BOUNDARY_PATTERN = re.compile(r"(?<=[.!?。！？])\s+")
PARAGRAPH_SEPARATOR_PATTERN = re.compile(r"\n(?:[ \t]*\n)+")

DEFAULT_HEADERS = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ko,en-US;q=0.9,en;q=0.8",
    "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
    "Origin": "https://nara-speller.co.kr",
    "Referer": "https://nara-speller.co.kr/old_speller/",
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36"
    ),
}

@dataclass(frozen=True)
class SpellCheckIssue:
    chunk_index: int
    page_index: int
    issue_index: int
    sentence: str
    original: str
    suggestions: list[str]
    reason: str
    start: int | None
    end: int | None
    correct_method: int | None
    error_message: str

def strip_html(value: str | None) -> str:
    text = LINE_BREAK_PATTERN.sub("\n", value or "")
    text = TAG_PATTERN.sub("", text)
    return unescape(text).strip()

def split_candidates(value: str | None) -> list[str]:
    return [candidate.strip() for candidate in str(value or "").split("|") if candidate.strip()]

def split_text_into_chunks(text: str, max_chars: int = DEFAULT_MAX_CHARS) -> list[str]:
    original = str(text or "")
    if not original.strip():
        return []

    units = split_paragraph_units(original)
    chunks = []
    current = ""

    for unit in units:
        candidate = unit if not current else f"{current}{unit}"

        if len(candidate) <= max_chars:
            current = candidate
            continue

        if current:
            chunks.append(current)
            current = ""

        if len(unit) <= max_chars:
            current = unit
            continue

        separator = ""
        body = unit
        separator_match = PARAGRAPH_SEPARATOR_PATTERN.search(unit)

        if separator_match and separator_match.end() == len(unit):
            separator = separator_match.group(0)
            body = unit[: separator_match.start()]

        for sentence in split_long_paragraph(body, max_chars=max_chars):
            if len(sentence) <= max_chars:
                chunks.append(sentence)
                continue

            start = 0
            while start < len(sentence):
                chunks.append(sentence[start : start + max_chars])
                start += max_chars

        if separator:
            if chunks and len(chunks[-1]) + len(separator) <= max_chars:
                chunks[-1] += separator
            else:
                current = separator

    if current:
        chunks.append(current)

    return chunks

def split_paragraph_units(text: str) -> list[str]:
    units = []
    start = 0

    for match in PARAGRAPH_SEPARATOR_PATTERN.finditer(text):
        paragraph = text[start : match.start()]
        separator = match.group(0)

        if paragraph:
            units.append(paragraph + separator)
        elif units:
            units[-1] += separator
        else:
            units.append(separator)

        start = match.end()

    tail = text[start:]
    if tail:
        units.append(tail)

    return units

def split_long_paragraph(paragraph: str, *, max_chars: int) -> list[str]:
    sentence_boundaries = list(SENTENCE_BOUNDARY_PATTERN.finditer(paragraph))

    if not sentence_boundaries:
        return [paragraph]

    sentences = []
    start = 0

    for boundary in sentence_boundaries:
        sentences.append(paragraph[start : boundary.end()])
        start = boundary.end()

    if start < len(paragraph):
        sentences.append(paragraph[start:])

    groups = []
    current = ""

    for sentence in sentences:
        candidate = sentence if not current else f"{current}{sentence}"

        if len(candidate) <= max_chars:
            current = candidate
            continue

        if current:
            groups.append(current)
        current = sentence

    if current:
        groups.append(current)

    return groups

def fetch_spell_check_html(
    text: str,
    *,
    strong_rules: bool = True,
    timeout: int = DEFAULT_TIMEOUT,
    url: str = DEFAULT_RESULTS_URL,
) -> str:
    body = {
        "text1": text,
        "chkKey": "",
    }

    if strong_rules:
        body["btnModeChange"] = "on"

    request = urllib.request.Request(
        url,
        data=urllib.parse.urlencode(body).encode("utf-8"),
        headers=DEFAULT_HEADERS,
        method="POST",
    )

    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            return response.read().decode("utf-8", "ignore")
    except urllib.error.HTTPError as error:
        if error.code == 403:
            raise RuntimeError(
                "The spell-check service returned HTTP 403. "
                "This environment may be hitting a Cloudflare/browser challenge. "
                "Retry later with lower request volume or from a browser-friendly network."
            ) from error

        raise RuntimeError(f"The spell-check service returned HTTP {error.code}.") from error

def extract_result_payload(html: str) -> list[dict]:
    match = RESULT_PAYLOAD_PATTERN.search(html)

    if not match:
        if NO_ISSUES_PATTERN.search(html):
            return []
        raise ValueError("Unable to find the spell-check payload in the returned HTML.")

    payload = json.loads(match.group(1))

    if not isinstance(payload, list):
        raise ValueError("The extracted spell-check payload was not a list.")

    return payload

def apply_page_corrections(page: dict) -> str:
    source = str(page.get("str", ""))
    corrected = source

    for error in sorted(page.get("errInfo", []), key=lambda item: int(item.get("start", -1)), reverse=True):
        suggestions = split_candidates(error.get("candWord"))
        original = str(error.get("orgStr", ""))

        if not suggestions:
            continue

        start = int(error.get("start", -1))
        end = int(error.get("end", -1))

        if start < 0 or end < start or end >= len(source):
            continue

        slice_end = end + 1
        if original:
            while slice_end > start and source[start:slice_end] != original and source[start : slice_end - 1] == original:
                slice_end -= 1

        corrected = f"{corrected[:start]}{suggestions[0]}{corrected[slice_end:]}"

    return corrected

def build_visible_text_index(text: str) -> tuple[str, list[int], list[int | None]]:
    visible_chars = []
    visible_indices = []
    visible_lookup = []

    for index, char in enumerate(text):
        if char.isspace():
            visible_lookup.append(None)
            continue

        visible_lookup.append(len(visible_indices))
        visible_chars.append(char)
        visible_indices.append(index)

    return "".join(visible_chars), visible_indices, visible_lookup

def preserve_original_layout(original: str, suggestion: str) -> str:
    if "\n" not in original:
        return suggestion

    original_visible, original_visible_indices, _ = build_visible_text_index(original)
    suggestion_visible, suggestion_visible_indices, _ = build_visible_text_index(suggestion)

    if original_visible != suggestion_visible:
        return suggestion

    if not original_visible_indices or not suggestion_visible_indices:
        return original if original.strip() else suggestion

    merged = []
    leading_original = original[: original_visible_indices[0]]
    leading_suggestion = suggestion[: suggestion_visible_indices[0]]
    merged.append(leading_original if leading_original.isspace() else leading_suggestion)

    for ordinal, suggestion_index in enumerate(suggestion_visible_indices):
        merged.append(suggestion[suggestion_index])

        next_original_index = original_visible_indices[ordinal + 1] if ordinal + 1 < len(original_visible_indices) else None
        next_suggestion_index = (
            suggestion_visible_indices[ordinal + 1] if ordinal + 1 < len(suggestion_visible_indices) else None
        )

        original_gap = (
            original[original_visible_indices[ordinal] + 1 : next_original_index]
            if next_original_index is not None
            else original[original_visible_indices[ordinal] + 1 :]
        )
        suggestion_gap = (
            suggestion[suggestion_index + 1 : next_suggestion_index]
            if next_suggestion_index is not None
            else suggestion[suggestion_index + 1 :]
        )

        merged.append(original_gap if "\n" in original_gap else suggestion_gap)

    return "".join(merged)

def apply_chunk_corrections(chunk: str, pages: list[dict]) -> str:
    combined_source = "".join(str(page.get("str", "")) for page in pages)
    fallback = "".join(apply_page_corrections(page) for page in pages) or chunk

    if not combined_source:
        return fallback

    chunk_visible, chunk_visible_indices, _ = build_visible_text_index(chunk)
    source_visible, _, source_visible_lookup = build_visible_text_index(combined_source)

    if chunk_visible != source_visible:
        return fallback

    replacements = []
    page_offset = 0

    for page in pages:
        for error in page.get("errInfo", []):
            suggestions = split_candidates(error.get("candWord"))
            if not suggestions:
                continue

            start = int(error.get("start", -1))
            end = int(error.get("end", -1))

            if start < 0 or end < start:
                continue

            start += page_offset
            end += page_offset

            visible_ordinals = [
                source_visible_lookup[index]
                for index in range(start, min(end + 1, len(source_visible_lookup)))
                if source_visible_lookup[index] is not None
            ]

            if not visible_ordinals:
                continue

            original_start = chunk_visible_indices[visible_ordinals[0]]
            original_end = chunk_visible_indices[visible_ordinals[-1]]
            replacements.append((original_start, original_end, suggestions[0], str(error.get("orgStr", ""))))

        page_offset += len(str(page.get("str", "")))

    if not replacements:
        return chunk

    corrected = chunk

    for start, end, suggestion, original in sorted(replacements, key=lambda item: item[0], reverse=True):
        slice_end = end + 1
        if original:
            while (
                slice_end > start
                and corrected[start:slice_end] != original
                and corrected[start : slice_end - 1] == original
            ):
                slice_end -= 1

        original_slice = corrected[start:slice_end]
        replacement = preserve_original_layout(original_slice, suggestion)
        corrected = f"{corrected[:start]}{replacement}{corrected[slice_end:]}"

    return corrected

def build_issue(chunk_index: int, page_index: int, issue_index: int, page: dict, error: dict) -> SpellCheckIssue:
    return SpellCheckIssue(
        chunk_index=chunk_index,
        page_index=page_index,
        issue_index=issue_index,
        sentence=str(page.get("str", "")),
        original=str(error.get("orgStr", "")),
        suggestions=split_candidates(error.get("candWord")),
        reason=strip_html(error.get("help")) or strip_html(error.get("errMsg")),
        start=int(error["start"]) if str(error.get("start", "")).strip() else None,
        end=int(error["end"]) if str(error.get("end", "")).strip() else None,
        correct_method=int(error["correctMethod"])
        if str(error.get("correctMethod", "")).strip()
        else None,
        error_message=strip_html(error.get("errMsg")),
    )

def check_text(
    text: str,
    *,
    max_chars: int = DEFAULT_MAX_CHARS,
    strong_rules: bool = True,
    timeout: int = DEFAULT_TIMEOUT,
    throttle_seconds: float = DEFAULT_THROTTLE_SECONDS,
    requester: Callable[..., str] = fetch_spell_check_html,
    sleep_fn: Callable[[float], None] = time.sleep,
) -> dict:
    chunks = split_text_into_chunks(text, max_chars=max_chars)
    corrected_chunks = []
    issues = []
    chunk_reports = []

    for chunk_index, chunk in enumerate(chunks):
        if chunk_index > 0 and throttle_seconds > 0:
            sleep_fn(throttle_seconds)

        html = requester(chunk, strong_rules=strong_rules, timeout=timeout)
        pages = extract_result_payload(html)
        corrected_chunk = apply_chunk_corrections(chunk, pages)

        corrected_chunks.append(corrected_chunk)
        chunk_reports.append(
            {
                "chunk_index": chunk_index,
                "original_text": chunk,
                "corrected_text": corrected_chunk,
                "page_count": len(pages),
            }
        )

        for page_index, page in enumerate(pages):
            for issue_index, error in enumerate(page.get("errInfo", [])):
                issues.append(build_issue(chunk_index, page_index, issue_index, page, error))

    return {
        "original_text": str(text or ""),
        "corrected_text": "".join(corrected_chunks),
        "chunks": chunk_reports,
        "issues": issues,
        "meta": {
            "chunk_count": len(chunks),
            "strong_rules": strong_rules,
            "max_chars": max_chars,
        },
    }

class WorkerSignals(QObject):
    log = pyqtSignal(str)
    finished = pyqtSignal()
    error = pyqtSignal(str)
    match_found = pyqtSignal(dict)

class DriverInitWorker(QThread):
    finished = pyqtSignal(str)
    error = pyqtSignal(str)

    def run(self):
        try:
            path = ChromeDriverManager().install()
            self.finished.emit(path)
        except Exception as e:
            self.error.emit(str(e))

class ScraperWorker(QThread):
    def __init__(self, excel_path, selected_sheet, display_count, capture_screenshot, global_driver_path, is_custom_excel):
        super().__init__()
        self.excel_path = excel_path
        self.selected_sheet = selected_sheet
        self.display_count = display_count
        self.capture_screenshot = capture_screenshot
        self.global_driver_path = global_driver_path
        self.is_custom_excel = is_custom_excel
        self.signals = WorkerSignals()
        
        self.driver_pool_lock = threading.Lock()
        self.active_drivers = []
        self.thread_local = threading.local()

    def log(self, message):
        self.signals.log.emit(message)

    def get_thread_driver(self):
        if not hasattr(self.thread_local, 'driver'):
            options = Options()
            options.page_load_strategy = 'eager'
            options.add_argument('--headless')
            options.add_argument('--disable-gpu')
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--window-size=390,844')
            options.add_argument('user-agent=Mozilla/5.0 (iPhone; CPU iPhone OS 16_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Mobile/15E148 Safari/604.1')
            options.add_argument('--disable-blink-features=AutomationControlled')
            
            service = Service(self.global_driver_path)
            driver = webdriver.Chrome(service=service, options=options)
            self.thread_local.driver = driver
            
            with self.driver_pool_lock:
                self.active_drivers.append(driver)
                
        return self.thread_local.driver

    def cleanup_drivers(self):
        with self.driver_pool_lock:
            for d in self.active_drivers:
                try:
                    d.quit()
                except:
                    pass
            self.active_drivers.clear()

    def get_blog_data_by_crawling(self, driver, keyword, target_blogs):
        if target_blogs is None: target_blogs = []
        titles, blog_names = [], []
        sort_name = "관련도순"
        screenshot_filepaths = []
        target_folder_path = ""
        
        self.log(f"▶ '{keyword}' - 데이터 탐색 중... ({self.display_count}개 목표)")
        
        try:
            url = f"https://m.search.naver.com/search.naver?ssc=tab.m_blog.all&sm=mtb_jum&query={urllib.parse.quote(keyword)}"
            driver.get(url)
            
            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, ".lst_total > li.bx, div[data-template-id='ugcItem']"))
                )
            except TimeoutException:
                self.log(f"[안내] '{keyword}' - 요소를 찾지 못해 대기 시간이 초과되었습니다.")
                return pd.DataFrame({'블로그명': pd.Series(dtype='str'), '제목': pd.Series(dtype='str')})

            last_height = driver.execute_script("return document.body.scrollHeight")
            
            while True:
                js_count = """
                var els = document.querySelectorAll(".lst_total > li.bx");
                if (els.length === 0) els = document.querySelectorAll("div[data-template-id='ugcItem']");
                var organicCount = 0;
                for(var i=0; i<els.length; i++) {
                    if (!els[i].querySelector(".sp_ad, .ad_ico, a[data-heatmap-target='articleSourceJSX_adtag']")) {
                        organicCount++;
                    }
                }
                return organicCount;
                """
                current_count = driver.execute_script(js_count)
                
                if current_count >= self.display_count:
                    break
                    
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(0.5) 
                
                new_height = driver.execute_script("return document.body.scrollHeight")
                if new_height == last_height:
                    time.sleep(1.0) 
                    new_height = driver.execute_script("return document.body.scrollHeight")
                    if new_height == last_height:
                        break 
                last_height = new_height

            js_extract = """
            var results = [];
            var containers = document.querySelectorAll(".lst_total > li.bx");
            if (containers.length === 0) containers = document.querySelectorAll("div[data-template-id='ugcItem']");
            
            for(var i=0; i<containers.length; i++) {
                if (containers[i].querySelector(".sp_ad, .ad_ico, a[data-heatmap-target='articleSourceJSX_adtag']")) {
                    continue;
                }
                
                var t = containers[i].querySelector(".api_txt_lines.total_tit, .title_link, a[data-heatmap-target='.nblg']");
                var b = containers[i].querySelector(".name, .sub_txt, a[data-heatmap-target='articleSourceJSX_title']");
                
                var titleText = t ? t.innerText.trim() : "제목 추출 불가";
                var blogText = "";
                if (b) {
                    blogText = b.innerText.trim();
                } else {
                    blogText = containers[i].innerText.replace(/\\n/g, ' ').substring(0, 50);
                }
                
                results.push({title: titleText, blog: blogText});
            }
            return results;
            """
            extracted_data = driver.execute_script(js_extract)
            
            for data in extracted_data[:self.display_count]:
                titles.append(data['title'])
                blog_names.append(data['blog'])
                    
            if self.capture_screenshot and target_blogs and titles:
                js_highlight_and_count = """
                var targets = arguments[0];
                var containers = document.querySelectorAll(".lst_total > li.bx");
                if (containers.length === 0) containers = document.querySelectorAll("div[data-template-id='ugcItem']");
                
                var matchRanks = [];
                var matchCount = 0;
                var actualRank = 1;
                
                for (var i = 0; i < containers.length; i++) {
                    if (containers[i].querySelector(".sp_ad, .ad_ico, a[data-heatmap-target='articleSourceJSX_adtag']")) {
                        continue;
                    }
                    
                    var b = containers[i].querySelector(".name, .sub_txt, a[data-heatmap-target='articleSourceJSX_title']");
                    var blogNameText = b ? b.innerText.trim() : "";
                    
                    var isMatch = false;
                    for (var j = 0; j < targets.length; j++) {
                        if (targets[j].trim() !== "" && blogNameText.includes(targets[j].trim())) {
                            isMatch = true;
                            break;
                        }
                    }
                    
                    if (isMatch) {
                        var targetElement = b || containers[i];
                        targetElement.style.backgroundColor = 'yellow';
                        targetElement.style.color = 'black';
                        containers[i].setAttribute('data-target-match', matchCount);
                        matchRanks.push(actualRank);
                        matchCount++;
                    }
                    
                    actualRank++;
                }
                return matchRanks;
                """
                match_ranks = driver.execute_script(js_highlight_and_count, target_blogs)
                
                if match_ranks:
                    base_dir = os.path.dirname(os.path.abspath(__file__)) if not getattr(sys, 'frozen', False) else os.path.dirname(sys.executable)
                    workspace_dir = os.path.join(base_dir, "Workspace")
                    
                    safe_keyword = re.sub(r'[\\/*?:"<>|]', "", keyword)
                    representative_target = target_blogs[0]
                    safe_target = re.sub(r'[\\/*?:"<>|]', "", representative_target)
                    timestamp = time.strftime("%Y%m%d_%H%M%S")
                    
                    target_folder_path = os.path.join(workspace_dir, safe_target)
                    
                    if self.is_custom_excel:
                        target_folder_path = os.path.join(target_folder_path, safe_keyword)
                        
                    os.makedirs(target_folder_path, exist_ok=True)
                    
                    for i, rank in enumerate(match_ranks):
                        js_scroll = f"""
                        var el = document.querySelector("[data-target-match='{i}']");
                        if (el) {{ el.scrollIntoView({{behavior: 'instant', block: 'center'}}); }}
                        """
                        driver.execute_script(js_scroll)
                        time.sleep(1)
                        
                        screenshot_filename = f"{safe_keyword}_{sort_name}_{rank}위노출_{timestamp}.png"
                        screenshot_filepath = os.path.join(target_folder_path, screenshot_filename)
                        driver.save_screenshot(screenshot_filepath)
                        screenshot_filepaths.append((rank, screenshot_filepath))
                        
                        if self.is_custom_excel:
                            self.log(f"[캡처] 목표({representative_target}) {rank}위 저장 완료: '{safe_target}/{safe_keyword}'")
                        else:
                            self.log(f"[캡처] 목표({representative_target}) {rank}위 저장 완료: '{safe_target}'")
                else:
                    self.log(f"[안내] 목표 블로그 '{', '.join(target_blogs)}'가 존재하지 않아 캡처 생략.")
            
            if not titles:
                self.log(f"[안내] '{keyword}' - 데이터를 추출하지 못했습니다.")
                        
        except Exception as e:
            self.log(f"[오류] '{keyword}' - 크롤링 실패: {e}")
                
        df = pd.DataFrame({'블로그명': blog_names, '제목': titles})
        df['제목'] = df['제목'].apply(lambda x: x[:20] + '...' if len(x) > 20 else x)
        df.index = df.index + 1
        return df, screenshot_filepaths, target_folder_path

    def get_rank_string(self, df_result, target_list):
        if df_result.empty or not target_list:
            return "검색/분석 불가"
        
        valid_targets = [t.strip() for t in target_list if t.strip()]
        if not valid_targets:
            return "미노출 (범위 밖)"
            
        pattern = "|".join([re.escape(t) for t in valid_targets])
        mask = df_result['블로그명'].str.contains(pattern, na=False, regex=True)
        found_indices = df_result.index[mask].tolist()
        if found_indices:
            return ", ".join([f"{i}위" for i in found_indices])
        return "미노출 (범위 밖)"

    def run(self):
        try:
            df_input = pd.read_excel(self.excel_path, sheet_name=self.selected_sheet)
            required_columns = ['키워드', '목표블로그', '관련도순_순위', '최근조회일시']
            for col in required_columns:
                if col not in df_input.columns:
                    df_input[col] = None
                    
            df_input['관련도순_순위'] = df_input['관련도순_순위'].astype(object)
            total_rows = len(df_input)
            
            valid_keywords = df_input['키워드'].dropna().astype(str).str.strip()
            valid_keywords = valid_keywords[(valid_keywords.str.lower() != 'nan') & (valid_keywords != '') & (valid_keywords.str.lower() != 'none')]
            
            if total_rows == 0 or valid_keywords.empty:
                self.signals.error.emit(f"'{self.selected_sheet}' 탭의 엑셀에 유효한 키워드가 비어있습니다.")
                return

            self.log(f"\n총 {total_rows}개의 작업 목록 발견. 병렬 조회를 시작합니다.\n" + "-"*40)
            
            def process_task(index, row_dict):
                raw_keyword = str(row_dict['키워드']).strip()
                raw_target = str(row_dict['목표블로그']).strip()
                keyword = raw_keyword if pd.notna(row_dict['키워드']) and raw_keyword.lower() != 'nan' else ""
                target_blogs = [t.strip() for t in raw_target.split(',') if t.strip()] if pd.notna(row_dict['목표블로그']) and raw_target.lower() != 'nan' else []
                display_targets = ", ".join(target_blogs)
                
                if not keyword: return None
                    
                self.log(f"\n=== [{index + 1}/{total_rows}] 키워드: '{keyword}', 목표: '{display_targets}' ===")
                
                delay_time = random.uniform(2.5, 4.5)
                self.log(f"※ 봇 탐지 우회를 위해 {delay_time:.1f}초 대기 중...")
                time.sleep(delay_time)
                
                driver = self.get_thread_driver()
                df_sim, screenshots, folder_path = self.get_blog_data_by_crawling(driver, keyword, target_blogs)
                sim_rank = self.get_rank_string(df_sim, target_blogs)
                
                ranks = []
                if sim_rank and "위" in sim_rank:
                    for part in sim_rank.split(","):
                        if "위" in part:
                            try:
                                r_val = int(part.replace("위", "").strip())
                                ranks.append(r_val)
                            except:
                                pass
                                
                if ranks:
                    representative_target = target_blogs[0]
                    if not screenshots:
                        screenshots = [(r, "") for r in ranks]
                    self.signals.match_found.emit({
                        'company': representative_target,
                        'keyword': keyword,
                        'ranks': ranks,
                        'screenshots': screenshots,
                        'folder': folder_path
                    })
                
                result_text = f"<< '{keyword}' 검색 결과 >>\n"
                for idx, row in df_sim.iterrows():
                    result_text += f"[{idx}위] {row['블로그명']} | {row['제목']}\n"
                self.log(result_text)
                self.log(f"▶ '{keyword}' 기록 준비 완료: {sim_rank}")
                
                return {'keyword': keyword, 'target_blog': raw_target, 'sim_rank': sim_rank}

            results_buffer = [] 
            
            with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
                futures = []
                for index, row in df_input.iterrows():
                    futures.append(executor.submit(process_task, index, row.to_dict()))
                    
                for future in concurrent.futures.as_completed(futures):
                    try:
                        result = future.result() 
                        if result: results_buffer.append(result)
                    except Exception as e:
                        self.log(f"[치명적 오류] 작업 중단: {e}")
                            
            if results_buffer:
                self.log("\n모든 탐색 완료. 엑셀 파일에 결과를 저장합니다...")
                while True:
                    try:
                        wb = openpyxl.load_workbook(self.excel_path)
                        ws = wb[self.selected_sheet] if self.selected_sheet in wb.sheetnames else wb.active
                        
                        header_map = {cell.value: cell.column for cell in ws[1]}
                        for req_col in required_columns:
                            if req_col not in header_map:
                                new_col_idx = ws.max_column + 1
                                ws.cell(row=1, column=new_col_idx, value=req_col)
                                header_map[req_col] = new_col_idx
                                
                        keyword_col_idx = header_map['키워드']
                        target_col_idx = header_map['목표블로그']
                        
                        for data in results_buffer:
                            current_row = None
                            for r in range(2, ws.max_row + 1):
                                cell_keyword = str(ws.cell(row=r, column=keyword_col_idx).value or "").strip()
                                cell_target = str(ws.cell(row=r, column=target_col_idx).value or "").strip()
                                if cell_keyword == data['keyword'] and cell_target == data['target_blog']:
                                    current_row = r
                                    break
                                    
                            if current_row is None:
                                current_row = ws.max_row + 1
                                ws.cell(row=current_row, column=keyword_col_idx, value=data['keyword'])
                                ws.cell(row=current_row, column=target_col_idx, value=data['target_blog'])
                            
                            current_time = time.strftime("%Y-%m-%d %H:%M:%S")
                            ws.cell(row=current_row, column=header_map['관련도순_순위'], value=data['sim_rank'])
                            ws.cell(row=current_row, column=header_map['최근조회일시'], value=current_time)
                            
                        wb.save(self.excel_path)
                        wb.close()
                        self.log("엑셀 업데이트가 성공적으로 완료되었습니다.")
                        break 
                    except PermissionError:
                        self.log("[경고] 엑셀 파일이 열려있습니다. 창을 닫아주세요. 5초 뒤 재시도...")
                        time.sleep(5.0)
                    except Exception as e:
                        self.log(f"[오류] 엑셀 저장 실패: {e}")
                        break

        except Exception as e:
            self.log(f"[메인 오류] {e}")
            traceback.print_exc()
        finally:
            self.cleanup_drivers()
            self.signals.finished.emit()


def get_windows_accent_color():
    try:
        registry_key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows\DWM")
        value, regtype = winreg.QueryValueEx(registry_key, "AccentColor")
        winreg.CloseKey(registry_key)
        # Windows registry stores it as AABBGGRR (ABGR)
        # We extract R, G, B
        r = value & 0xff
        g = (value >> 8) & 0xff
        b = (value >> 16) & 0xff
        return f"#{r:02x}{g:02x}{b:02x}"
    except Exception:
        return None

# --- [업체 리스트 기능 추가] ---

def get_companies_file_path():
    base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(base_dir, 'Data')
    os.makedirs(data_dir, exist_ok=True)
    return os.path.join(data_dir, 'companies.json')

def load_companies():
    path = get_companies_file_path()
    if os.path.exists(path):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return []
    return [
        {
            "name": "예시 업체 1",
            "homepage": "https://www.naver.com",
            "place": "",
            "blog1": "https://blog.naver.com",
            "blog2": "",
            "instagram": "https://www.instagram.com"
        }
    ]

def save_companies(companies):
    path = get_companies_file_path()
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(companies, f, ensure_ascii=False, indent=4)
        return True
    except Exception:
        return False

def get_holiday_checks_file_path():
    base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(base_dir, 'Data')
    os.makedirs(data_dir, exist_ok=True)
    return os.path.join(data_dir, 'holiday_checks.json')

def load_holiday_checks():
    path = get_holiday_checks_file_path()
    if os.path.exists(path):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_holiday_checks(data):
    path = get_holiday_checks_file_path()
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        return True
    except Exception:
        return False

def get_korean_holidays(year):
    holidays = {
        f"{year}-01-01": "신정",
        f"{year}-03-01": "삼일절",
        f"{year}-05-05": "어린이날",
        f"{year}-06-06": "현충일",
        f"{year}-07-17": "제헌절",
        f"{year}-08-15": "광복절",
        f"{year}-10-03": "개천절",
        f"{year}-10-09": "한글날",
        f"{year}-12-25": "성탄절"
    }
    
    if year == 2024:
        holidays.update({
            "2024-02-09": "설날 연휴",
            "2024-02-10": "설날",
            "2024-02-11": "설날 연휴",
            "2024-02-12": "설날 대체공휴일",
            "2024-04-10": "제22대 국회의원 선거",
            "2024-05-06": "어린이날 대체공휴일",
            "2024-05-15": "부처님오신날",
            "2024-09-16": "추석 연휴",
            "2024-09-17": "추석",
            "2024-09-18": "추석 연휴"
        })
    elif year == 2025:
        holidays.update({
            "2025-01-27": "임시공휴일",
            "2025-01-28": "설날 연휴",
            "2025-01-29": "설날",
            "2025-01-30": "설날 연휴",
            "2025-03-03": "삼일절 대체공휴일",
            "2025-05-06": "어린이날 및 부처님오신날 대체공휴일",
            "2025-10-05": "추석 연휴",
            "2025-10-06": "추석",
            "2025-10-07": "추석 연휴",
            "2025-10-08": "추석 대체공휴일"
        })
    elif year == 2026:
        holidays.update({
            "2026-02-16": "설날 연휴",
            "2026-02-17": "설날",
            "2026-02-18": "설날 연휴",
            "2026-03-02": "삼일절 대체공휴일",
            "2026-05-24": "부처님오신날",
            "2026-05-25": "부처님오신날 대체공휴일",
            "2026-06-03": "제9회 전국동시지방선거",
            "2026-08-17": "광복절 대체공휴일",
            "2026-09-24": "추석 연휴",
            "2026-09-25": "추석",
            "2026-09-26": "추석 연휴",
            "2026-09-28": "추석 대체공휴일",
            "2026-10-05": "개천절 대체공휴일"
        })
    elif year == 2027:
        holidays.update({
            "2027-02-06": "설날 연휴",
            "2027-02-07": "설날",
            "2027-02-08": "설날 연휴",
            "2027-02-09": "설날 대체공휴일",
            "2027-05-13": "부처님오신날",
            "2027-07-19": "제헌절 대체공휴일",
            "2027-08-16": "광복절 대체공휴일",
            "2027-09-14": "추석 연휴",
            "2027-09-15": "추석",
            "2027-09-16": "추석 연휴",
            "2027-10-04": "개천절 대체공휴일",
            "2027-10-11": "한글날 대체공휴일",
            "2027-12-27": "성탄절 대체공휴일"
        })
    elif year == 2028:
        holidays.update({
            "2028-01-26": "설날 연휴",
            "2028-01-27": "설날",
            "2028-01-28": "설날 연휴",
            "2028-04-12": "제23대 국회의원 선거",
            "2028-05-02": "부처님오신날",
            "2028-09-30": "추석 연휴",
            "2028-10-01": "추석 연휴",
            "2028-10-02": "추석",
            "2028-10-04": "추석 대체공휴일",
            "2028-10-05": "대체공휴일"
        })
    return holidays

class CompanyCard(QFrame):
    def __init__(self, company_data, on_edit, on_delete, parent=None):
        super().__init__(parent)
        self.company_data = company_data
        self.on_edit = on_edit
        self.on_delete = on_delete
        self.setObjectName("CompanyCard")
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 20, 24, 20)
        layout.setSpacing(14)
        
        # 상단 영역 (업체명 + 수정/삭제 버튼)
        top_layout = QHBoxLayout()
        
        self.name_label = QLabel(company_data.get('name', '업체명 없음'), self)
        self.name_label.setFont(QFont("SUIT", 16, QFont.Bold))
        top_layout.addWidget(self.name_label)
        
        top_layout.addStretch(1)
        
        self.edit_btn = TransparentToolButton(FluentIcon.EDIT, self)
        self.edit_btn.setFixedSize(30, 30)
        self.edit_btn.clicked.connect(self.on_edit)
        top_layout.addWidget(self.edit_btn)
        
        self.delete_btn = TransparentToolButton(FluentIcon.DELETE, self)
        self.delete_btn.setFixedSize(30, 30)
        self.delete_btn.clicked.connect(self.on_delete)
        top_layout.addWidget(self.delete_btn)
        
        layout.addLayout(top_layout)
        
        # 하단 영역 (링크들)
        links_layout = QHBoxLayout()
        links_layout.setSpacing(24)
        
        self.link_configs = [
            ('homepage', '홈페이지'),
            ('place', '플레이스'),
            ('blog1', '블로그1'),
            ('blog2', '블로그2'),
            ('instagram', '인스타그램')
        ]
        
        self.buttons = []
        for key, display_name in self.link_configs:
            btn = PushButton(display_name, self)
            url = company_data.get(key, '').strip()
            if url:
                btn.setEnabled(True)
                btn.clicked.connect(lambda checked, u=url: self.open_link(u))
            else:
                btn.setEnabled(False)
                
            links_layout.addWidget(btn)
            self.buttons.append((btn, url))
            
        links_layout.addStretch(1)
        layout.addLayout(links_layout)
        
        self.update_style()
        qconfig.themeChanged.connect(self.update_style)
        
    def update_style(self):
        is_dark = isDarkTheme()
        bg_color = "#2C2C2C" if is_dark else "#F3F3F3"
        border_color = "#3A3A3A" if is_dark else "#E5E5E5"
        hover_bg = "#383838" if is_dark else "#EBEBEB"
        hover_border = "#4D4D4D" if is_dark else "#D8D8D8"
        text_color = "#FFFFFF" if is_dark else "#000000"
        
        link_color = "#60CDFF" if is_dark else "#0078D4"
        link_hover_color = "#A6E2FF" if is_dark else "#005A9E"
        link_disabled_color = "#666666" if is_dark else "#B0B0B0"
        
        self.setStyleSheet(f"""
            QFrame#CompanyCard {{
                background-color: {bg_color};
                border: 1px solid {border_color};
                border-radius: 16px;
            }}
            QFrame#CompanyCard:hover {{
                background-color: {hover_bg};
                border: 1px solid {hover_border};
            }}
        """)
        self.name_label.setStyleSheet(f"background: transparent; color: {text_color};")
        
        for btn, url in self.buttons:
            btn.setStyleSheet(f"""
                QPushButton {{
                    color: {link_color};
                    background: transparent;
                    border: none;
                    font-family: 'SUIT';
                    font-size: 14px;
                    font-weight: 500;
                    padding: 2px 6px;
                    text-align: left;
                }}
                QPushButton:hover {{
                    color: {link_hover_color};
                    text-decoration: underline;
                    background: transparent;
                }}
                QPushButton:disabled {{
                    color: {link_disabled_color};
                    background: transparent;
                }}
            """)
            
    def open_link(self, url):
        if not url:
            return
        url = url.strip()
        if not (url.startswith("http://") or url.startswith("https://")):
            url = "https://" + url
        QDesktopServices.openUrl(QUrl(url))
        
    def closeEvent(self, event):
        try:
            qconfig.themeChanged.disconnect(self.update_style)
        except Exception:
            pass
        super().closeEvent(event)

class CompanyDialog(QDialog):
    def __init__(self, company_data=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("업체 정보 입력")
        self.setFixedSize(450, 580)
        
        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(24, 24, 24, 24)
        self.layout.setSpacing(14)
        
        title_text = "업체 정보 수정" if company_data else "새 업체 추가"
        self.title_label = SubtitleLabel(title_text, self)
        self.layout.addWidget(self.title_label)
        
        # Form Fields
        self.inputs = {}
        self.labels = []
        fields = [
            ('name', '업체명 *'),
            ('homepage', '홈페이지 주소'),
            ('place', '플레이스 주소'),
            ('blog_id', '네이버 블로그 ID'),
            ('blog1', '블로그 1 주소'),
            ('blog2', '블로그 2 주소'),
            ('instagram', '인스타그램 주소')
        ]
        
        for key, label in fields:
            row_layout = QVBoxLayout()
            row_layout.setSpacing(4)
            
            lbl = QLabel(label, self)
            lbl.setFont(QFont("SUIT", 10, QFont.Bold))
            row_layout.addWidget(lbl)
            self.labels.append((lbl, key == 'name'))
            
            edit = LineEdit(self)
            edit.setPlaceholderText(f"{label} 입력")
            if company_data:
                edit.setText(company_data.get(key, ''))
                
            row_layout.addWidget(edit)
            self.layout.addLayout(row_layout)
            self.inputs[key] = edit
            
        self.layout.addStretch(1)
        
        # Action Buttons
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(12)
        btn_layout.addStretch(1)
        
        self.ok_btn = PushButton("저장", self)
        self.ok_btn.clicked.connect(self.validate_and_accept)
        
        self.cancel_btn = PushButton("취소", self)
        self.cancel_btn.clicked.connect(self.reject)
        
        btn_layout.addWidget(self.ok_btn)
        btn_layout.addWidget(self.cancel_btn)
        
        self.layout.addLayout(btn_layout)
        
        self.update_style()
        qconfig.themeChanged.connect(self.update_style)
        
    def update_style(self):
        is_dark = isDarkTheme()
        bg_color = "#202020" if is_dark else "#FFFFFF"
        text_color = "#FFFFFF" if is_dark else "#000000"
        label_color = "#AAAAAA" if is_dark else "#333333"
        required_color = "#FF6B6B" if is_dark else "#D83B01"
        
        self.setStyleSheet(f"""
            QDialog {{
                background-color: {bg_color};
            }}
        """)
        
        self.title_label.setStyleSheet(f"color: {text_color}; background: transparent;")
        
        for lbl, is_required in self.labels:
            color = required_color if is_required else label_color
            lbl.setStyleSheet(f"color: {color}; font-size: 13px; font-weight: bold; background: transparent;")
            
        # Style text inputs to make sure they are highly readable
        input_bg = "#161616" if is_dark else "#FFFFFF"
        input_fg = "#FFFFFF" if is_dark else "#000000"
        input_border = "#3A3A3A" if is_dark else "#CCCCCC"
        
        for edit in self.inputs.values():
            edit.setStyleSheet(f"""
                QLineEdit {{
                    background-color: {input_bg};
                    color: {input_fg};
                    border: 1px solid {input_border};
                    border-radius: 4px;
                    padding: 4px 8px;
                    font-family: 'SUIT';
                }}
                QLineEdit:focus {{
                    border: 1px solid #0078D4;
                }}
            """)
            
        # Action Buttons Style
        self.ok_btn.setStyleSheet("""
            QPushButton {
                background-color: #0078D4;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 16px;
                font-weight: bold;
                font-family: 'SUIT';
            }
            QPushButton:hover {
                background-color: #005A9E;
            }
        """)
        
        cancel_bg = "#1E1E1E" if is_dark else "#F3F3F3"
        cancel_fg = "#FFFFFF" if is_dark else "#000000"
        cancel_border = "#3A3A3A" if is_dark else "#CCCCCC"
        cancel_hover = "#2C2C2C" if is_dark else "#EBEBEB"
        
        self.cancel_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {cancel_bg};
                color: {cancel_fg};
                border: 1px solid {cancel_border};
                border-radius: 4px;
                padding: 6px 16px;
                font-family: 'SUIT';
            }}
            QPushButton:hover {{
                background-color: {cancel_hover};
            }}
        """)
        
    def validate_and_accept(self):
        name = self.inputs['name'].text().strip()
        if not name:
            QMessageBox.warning(self, "입력 오류", "업체명은 필수 입력 항목입니다.")
            return
        self.accept()
        
    def get_data(self):
        return {
            'name': self.inputs['name'].text().strip(),
            'homepage': self.inputs['homepage'].text().strip(),
            'place': self.inputs['place'].text().strip(),
            'blog_id': self.inputs['blog_id'].text().strip(),
            'blog1': self.inputs['blog1'].text().strip(),
            'blog2': self.inputs['blog2'].text().strip(),
            'instagram': self.inputs['instagram'].text().strip()
        }

    def closeEvent(self, event):
        try:
            qconfig.themeChanged.disconnect(self.update_style)
        except Exception:
            pass
        super().closeEvent(event)

class HolidayCheckCard(QFrame):
    def __init__(self, company_data, date_str, check_data, on_changed, parent=None):
        super().__init__(parent)
        self.company_data = company_data
        self.date_str = date_str
        self.check_data = check_data  # {"place_checked": bool, "place_reason": str, "popup_checked": bool, "popup_reason": str}
        self.on_changed = on_changed
        self.setObjectName("HolidayCheckCard")
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(10)
        
        # Row 1: Company Name
        self.name_label = QLabel(company_data.get('name', '업체명 없음'), self)
        self.name_label.setFont(QFont("SUIT", 12, QFont.Bold))
        layout.addWidget(self.name_label)
        
        # Row 2: Verification Controls
        controls_layout = QHBoxLayout()
        controls_layout.setSpacing(8)
        
        # 1) Place Section
        self.place_btn = QPushButton("플레이스", self)
        self.place_btn.setCursor(Qt.PointingHandCursor)
        place_url = company_data.get('place', '').strip()
        if place_url:
            self.place_btn.setEnabled(True)
            self.place_btn.clicked.connect(lambda checked, u=place_url: self.open_link(u))
        else:
            self.place_btn.setEnabled(False)
            self.place_btn.setCursor(Qt.ArrowCursor)
            
        controls_layout.addWidget(self.place_btn)
        
        self.place_input = LineEdit(self)
        self.place_input.setPlaceholderText("사유")
        self.place_input.setText(check_data.get('place_reason', ''))
        self.place_input.setFixedWidth(90)
        self.place_input.textEdited.connect(self.notify_change)
        controls_layout.addWidget(self.place_input)
        
        self.place_chk_btn = QPushButton(self)
        self.place_chk_btn.setFixedSize(24, 24)
        is_place_checked = check_data.get('place_checked', True)
        self.set_toggle_style(self.place_chk_btn, is_place_checked)
        self.place_chk_btn.clicked.connect(self.toggle_place_status)
        controls_layout.addWidget(self.place_chk_btn)
        
        # Spacer
        controls_layout.addSpacing(16)
        
        # 2) Popup Section
        self.popup_btn = QPushButton("팝업", self)
        self.popup_btn.setCursor(Qt.PointingHandCursor)
        popup_url = company_data.get('homepage', '').strip()
        if popup_url:
            self.popup_btn.setEnabled(True)
            self.popup_btn.clicked.connect(lambda checked, u=popup_url: self.open_link(u))
        else:
            self.popup_btn.setEnabled(False)
            self.popup_btn.setCursor(Qt.ArrowCursor)
            
        controls_layout.addWidget(self.popup_btn)
        
        self.popup_input = LineEdit(self)
        self.popup_input.setPlaceholderText("사유")
        self.popup_input.setText(check_data.get('popup_reason', ''))
        self.popup_input.setFixedWidth(90)
        self.popup_input.textEdited.connect(self.notify_change)
        controls_layout.addWidget(self.popup_input)
        
        self.popup_chk_btn = QPushButton(self)
        self.popup_chk_btn.setFixedSize(24, 24)
        is_popup_checked = check_data.get('popup_checked', True)
        self.set_toggle_style(self.popup_chk_btn, is_popup_checked)
        self.popup_chk_btn.clicked.connect(self.toggle_popup_status)
        controls_layout.addWidget(self.popup_chk_btn)
        
        controls_layout.addStretch(1)
        layout.addLayout(controls_layout)
        
        self.update_style()
        qconfig.themeChanged.connect(self.update_style)
        
    def set_toggle_style(self, btn, is_checked):
        btn.setProperty("checked_state", is_checked)
        if is_checked:
            btn.setText("✔")
            btn.setStyleSheet("""
                QPushButton {
                    background-color: #25A662;
                    color: white;
                    font-family: 'SUIT';
                    font-size: 12px;
                    font-weight: bold;
                    border: none;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #1D8B52;
                }
            """)
        else:
            btn.setText("X")
            btn.setStyleSheet("""
                QPushButton {
                    background-color: #E81123;
                    color: white;
                    font-family: 'SUIT';
                    font-size: 12px;
                    font-weight: bold;
                    border: none;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #D10F20;
                }
            """)
            
    def update_style(self):
        is_dark = isDarkTheme()
        bg_color = "#2C2C2C" if is_dark else "#F3F3F3"
        border_color = "#3A3A3A" if is_dark else "#E5E5E5"
        text_color = "#FFFFFF" if is_dark else "#000000"
        
        link_color = "#60CDFF" if is_dark else "#0078D4"
        link_hover_color = "#A6E2FF" if is_dark else "#005A9E"
        link_disabled_color = "#666666" if is_dark else "#B0B0B0"
        
        self.setStyleSheet(f"""
            QFrame#HolidayCheckCard {{
                background-color: {bg_color};
                border: 1px solid {border_color};
                border-radius: 12px;
            }}
        """)
        self.name_label.setStyleSheet(f"background: transparent; color: {text_color};")
        
        for btn in [self.place_btn, self.popup_btn]:
            btn.setStyleSheet(f"""
                QPushButton {{
                    color: {link_color};
                    background: transparent;
                    border: none;
                    font-family: 'SUIT';
                    font-size: 15px;
                    font-weight: bold;
                    padding: 2px 4px;
                    text-align: left;
                }}
                QPushButton:hover {{
                    color: {link_hover_color};
                    text-decoration: underline;
                    background: transparent;
                }}
                QPushButton:disabled {{
                    color: {link_disabled_color};
                    background: transparent;
                }}
            """)
            
        input_bg = "#161616" if is_dark else "#FFFFFF"
        input_fg = "#FFFFFF" if is_dark else "#000000"
        input_border = "#3A3A3A" if is_dark else "#CCCCCC"
        input_style = f"""
            QLineEdit {{
                background-color: {input_bg} !important;
                color: {input_fg} !important;
                border: 1px solid {input_border};
                border-radius: 4px;
                font-family: 'SUIT';
                font-size: 14px;
                font-weight: bold;
                padding: 2px 4px;
            }}
            QLineEdit:focus {{
                border: 1px solid {link_color};
            }}
        """
        self.place_input.setStyleSheet(input_style)
        self.popup_input.setStyleSheet(input_style)
            
    def toggle_place_status(self):
        current = self.place_chk_btn.property("checked_state")
        self.set_toggle_style(self.place_chk_btn, not current)
        self.notify_change()
        
    def toggle_popup_status(self):
        current = self.popup_chk_btn.property("checked_state")
        self.set_toggle_style(self.popup_chk_btn, not current)
        self.notify_change()
        
    def notify_change(self):
        data = {
            "place_checked": self.place_chk_btn.property("checked_state"),
            "place_reason": self.place_input.text().strip(),
            "popup_checked": self.popup_chk_btn.property("checked_state"),
            "popup_reason": self.popup_input.text().strip()
        }
        self.on_changed(self.company_data.get('name', ''), data)
        
    def open_link(self, url):
        if not url:
            return
        url = url.strip()
        if not (url.startswith("http://") or url.startswith("https://")):
            url = "https://" + url
        QDesktopServices.openUrl(QUrl(url))
        
    def closeEvent(self, event):
        try:
            qconfig.themeChanged.disconnect(self.update_style)
        except Exception:
            pass
        super().closeEvent(event)

class CompanyListInterface(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent=parent)
        self.setObjectName("CompanyListInterface")
        
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(36, 36, 36, 36)
        main_layout.setSpacing(16)
        
        # Header (Title + Buttons)
        header_layout = QHBoxLayout()
        self.title_label = TitleLabel("업체 리스트", self)
        header_layout.addWidget(self.title_label)
        
        header_layout.addStretch(1)
        
        self.add_btn = PushButton("업체 추가", self)
        self.add_btn.setIcon(FluentIcon.ADD)
        self.add_btn.clicked.connect(self.add_company)
        header_layout.addWidget(self.add_btn)
        
        # Holiday Check Button next to Add Company (reordered to the right)
        self.holiday_btn = PushButton("휴진 체크", self)
        self.holiday_btn.setIcon(FluentIcon.CALENDAR)
        self.holiday_btn.clicked.connect(self.toggle_holiday_panel)
        header_layout.addWidget(self.holiday_btn)
        
        main_layout.addLayout(header_layout)
        
        # Content layout (Split layout)
        self.content_layout = QHBoxLayout()
        self.content_layout.setSpacing(24)
        
        # Left Panel: Scroll Area for Card List
        self.scroll_area = ScrollArea(self)
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        
        self.scroll_content = QWidget()
        self.scroll_content.setObjectName("ScrollContent")
        self.scroll_content.setStyleSheet("QWidget#ScrollContent { background: transparent; }")
        self.scroll_layout = QVBoxLayout(self.scroll_content)
        self.scroll_layout.setContentsMargins(0, 0, 10, 0)
        self.scroll_layout.setSpacing(16)
        
        self.scroll_area.setWidget(self.scroll_content)
        self.content_layout.addWidget(self.scroll_area, 3) # Stretch 3
        
        # Right Panel: Holiday Check Panel
        self.holiday_panel = QWidget(self)
        self.holiday_panel.setObjectName("HolidayPanel")
        self.holiday_panel_layout = QVBoxLayout(self.holiday_panel)
        self.holiday_panel_layout.setContentsMargins(0, 0, 0, 0)
        self.holiday_panel_layout.setSpacing(16)
        
        # 1. Calendar Widget
        self.calendar = QCalendarWidget(self.holiday_panel)
        self.calendar.setGridVisible(True)
        self.calendar.setVerticalHeaderFormat(QCalendarWidget.NoVerticalHeader)
        self.calendar.selectionChanged.connect(self.render_holiday_checks)
        self.calendar.currentPageChanged.connect(self.update_calendar_holidays)
        self.holiday_panel_layout.addWidget(self.calendar)
        
        # Add "오늘" button to calendar navigation bar
        nav_bar = self.calendar.findChild(QWidget, "qt_calendar_navigationbar")
        if nav_bar:
            layout = nav_bar.layout()
            if layout:
                self.today_btn = QPushButton("오늘", nav_bar)
                self.today_btn.setFixedSize(45, 24)
                self.today_btn.setCursor(Qt.PointingHandCursor)
                self.today_btn.clicked.connect(self.go_to_today)
                layout.insertWidget(layout.count() - 1, self.today_btn)
        
        # 2. Holiday checklist title and export button (QHBoxLayout)
        title_layout = QHBoxLayout()
        self.holiday_list_title = SubtitleLabel("휴진 체크 리스트", self.holiday_panel)
        self.holiday_list_title.setFont(QFont("SUIT", 12, QFont.Bold))
        title_layout.addWidget(self.holiday_list_title)
        
        title_layout.addStretch(1)
        
        self.export_btn = PushButton("메모장 내보내기", self.holiday_panel)
        self.export_btn.setIcon(FluentIcon.SHARE)
        self.export_btn.setFixedSize(160, 28)
        self.export_btn.setFont(QFont("SUIT", 9, QFont.Bold))
        self.export_btn.clicked.connect(self.export_holiday_checks)
        title_layout.addWidget(self.export_btn)
        
        self.holiday_panel_layout.addLayout(title_layout)
        
        # 3. Checklist Scroll Area
        self.holiday_list_scroll = ScrollArea(self.holiday_panel)
        self.holiday_list_scroll.setWidgetResizable(True)
        
        self.holiday_list_content = QWidget()
        self.holiday_list_content.setObjectName("HolidayListContent")
        self.holiday_list_content.setStyleSheet("QWidget#HolidayListContent { background: transparent; }")
        self.holiday_list_layout = QVBoxLayout(self.holiday_list_content)
        self.holiday_list_layout.setContentsMargins(12, 12, 12, 12)
        self.holiday_list_layout.setSpacing(12)
        
        self.holiday_list_scroll.setWidget(self.holiday_list_content)
        self.holiday_panel_layout.addWidget(self.holiday_list_scroll)
        
        self.content_layout.addWidget(self.holiday_panel, 2) # Stretch 2
        
        main_layout.addLayout(self.content_layout)
        
        # Hide holiday panel by default
        self.holiday_panel.setVisible(False)
        
        # Load and render initial list
        self.companies = load_companies()
        self.render_list()
        
        # Styling & Holiday markings setup
        self.update_holiday_panel_style()
        self.update_calendar_style()
        self.update_calendar_holidays()
        qconfig.themeChanged.connect(self.update_holiday_panel_style)
        qconfig.themeChanged.connect(self.update_calendar_style)
        qconfig.themeChanged.connect(self.update_calendar_holidays)
        
    def render_list(self):
        # Clear existing layout
        while self.scroll_layout.count():
            child = self.scroll_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()
                
        # Create a card for each company
        for idx, company in enumerate(self.companies):
            card = CompanyCard(
                company, 
                on_edit=lambda checked, i=idx: self.edit_company(i),
                on_delete=lambda checked, i=idx: self.delete_company(i),
                parent=self.scroll_content
            )
            self.scroll_layout.addWidget(card)
            
        self.scroll_layout.addStretch(1)
        
        # Sync holiday check list if panel is visible
        if hasattr(self, 'holiday_panel') and self.holiday_panel.isVisible():
            self.render_holiday_checks()
            
    def toggle_holiday_panel(self):
        is_visible = not self.holiday_panel.isVisible()
        self.holiday_panel.setVisible(is_visible)
        if is_visible:
            self.render_holiday_checks()
            self.update_calendar_holidays()
            
    def render_holiday_checks(self):
        while self.holiday_list_layout.count():
            child = self.holiday_list_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()
                
        selected_date_str = self.calendar.selectedDate().toString("yyyy-MM-dd")
        all_checks = load_holiday_checks()
        date_checks = all_checks.get(selected_date_str, {})
        
        # Reload companies to reflect updates/deletes/additions
        self.companies = load_companies()
        
        for company in self.companies:
            comp_name = company.get('name', '')
            comp_checks = date_checks.get(comp_name, {
                "place_checked": True,
                "place_reason": "",
                "popup_checked": True,
                "popup_reason": ""
            })
            
            card = HolidayCheckCard(
                company_data=company,
                date_str=selected_date_str,
                check_data=comp_checks,
                on_changed=lambda name, data: self.save_company_check(selected_date_str, name, data),
                parent=self.holiday_list_content
            )
            self.holiday_list_layout.addWidget(card)
            
        self.holiday_list_layout.addStretch(1)
        
    def save_company_check(self, date_str, company_name, check_data):
        all_checks = load_holiday_checks()
        if date_str not in all_checks:
            all_checks[date_str] = {}
        all_checks[date_str][company_name] = check_data
        save_holiday_checks(all_checks)
        
    def export_holiday_checks(self):
        selected_date_str = self.calendar.selectedDate().toString("yyyy-MM-dd")
        all_checks = load_holiday_checks()
        date_checks = all_checks.get(selected_date_str, {})
        
        self.companies = load_companies()
        if not self.companies:
            InfoBar.warning("안내", "내보낼 업체 데이터가 없습니다.", duration=3000, parent=self)
            return
            
        lines = [f"[{selected_date_str} 휴진 체크 리스트]\n"]
        for company in self.companies:
            name = company.get('name', '업체명 없음')
            comp_checks = date_checks.get(name, {
                "place_checked": True,
                "place_reason": "",
                "popup_checked": True,
                "popup_reason": ""
            })
            
            lines.append(f"●{name}")
            
            # Place Status (O / X)
            place_ok = "O" if comp_checks.get("place_checked", True) else "X"
            place_reason = comp_checks.get("place_reason", "").strip()
            place_suffix = f" ({place_reason})" if place_reason else ""
            lines.append(f"플레이스 : {place_ok}{place_suffix}")
            
            # Popup Status (O / X)
            popup_ok = "O" if comp_checks.get("popup_checked", True) else "X"
            popup_reason = comp_checks.get("popup_reason", "").strip()
            popup_suffix = f" ({popup_reason})" if popup_reason else ""
            lines.append(f"팝업 : {popup_ok}{popup_suffix}\n")
            
        export_text = "\n".join(lines)
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "메모장 내보내기",
            f"휴진체크_{selected_date_str}.txt",
            "텍스트 파일 (*.txt)"
        )
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(export_text)
                InfoBar.success("성공", "텍스트 파일이 저장되었습니다.", duration=3000, parent=self)
            except Exception as e:
                InfoBar.error("오류", f"파일을 저장할 수 없습니다: {e}", duration=4000, parent=self)
                
    def go_to_today(self):
        self.calendar.setSelectedDate(QDate.currentDate())
        self.calendar.showToday()
        self.render_holiday_checks()
        
    def update_calendar_holidays(self):
        year = self.calendar.yearShown()
        holidays = get_korean_holidays(year)
        
        # Apply red color format to Korean holidays
        for date_str, holiday_name in holidays.items():
            y, m, d = map(int, date_str.split('-'))
            qdate = QDate(y, m, d)
            fmt = QTextCharFormat()
            fmt.setForeground(QBrush(QColor("#E81123")))
            fmt.setFontWeight(QFont.Bold)
            fmt.setToolTip(holiday_name)
            self.calendar.setDateTextFormat(qdate, fmt)
            
    def update_calendar_style(self):
        is_dark = isDarkTheme()
        bg_color = "#2C2C2C" if is_dark else "#FFFFFF"
        header_bg = "#202020" if is_dark else "#F9F9F9"
        text_color = "#FFFFFF" if is_dark else "#000000"
        border_color = "#3A3A3A" if is_dark else "#E5E5E5"
        select_bg = "#60CDFF" if is_dark else "#0078D4"
        select_fg = "#000000" if is_dark else "#FFFFFF"
        hover_bg = "#383838" if is_dark else "#EBEBEB"
        pressed_bg = "#4D4D4D" if is_dark else "#D8D8D8"
        disabled_color = "#666666" if is_dark else "#B0B0B0"
        
        # Style today button if it exists
        if hasattr(self, 'today_btn'):
            btn_color = "#60CDFF" if is_dark else "#0078D4"
            btn_hover = "rgba(96, 205, 255, 0.1)" if is_dark else "rgba(0, 120, 212, 0.1)"
            self.today_btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: transparent;
                    color: {btn_color};
                    font-family: 'SUIT';
                    font-size: 11px;
                    font-weight: bold;
                    border: 1px solid {btn_color};
                    border-radius: 4px;
                }}
                QPushButton:hover {{
                    background-color: {btn_hover};
                }}
            """)
        
        # Style navigation buttons directly in code
        from PyQt5.QtWidgets import QToolButton
        from PyQt5.QtGui import QIcon
        
        prev_btn = self.calendar.findChild(QToolButton, "qt_calendar_prevmonth")
        if prev_btn:
            prev_btn.setIcon(QIcon())
            prev_btn.setText("◀")
            
        next_btn = self.calendar.findChild(QToolButton, "qt_calendar_nextmonth")
        if next_btn:
            next_btn.setIcon(QIcon())
            next_btn.setText("▶")
            
        self.calendar.setStyleSheet(f"""
            QCalendarWidget {{
                background-color: {bg_color};
                border: 1px solid {border_color};
                border-radius: 12px;
            }}
            QCalendarWidget QWidget {{
                alternate-background-color: transparent;
                background-color: {bg_color};
                color: {text_color};
                font-family: 'SUIT';
                font-size: 13px;
            }}
            QCalendarWidget QWidget#qt_calendar_navigationbar {{
                background-color: {header_bg};
                border-top-left-radius: 12px;
                border-top-right-radius: 12px;
                border-bottom: 1px solid {border_color};
            }}
            QCalendarWidget QToolButton {{
                color: {text_color};
                background-color: transparent;
                font-weight: bold;
                border: none;
                border-radius: 6px;
                margin: 4px;
                padding: 4px 8px;
            }}
            QCalendarWidget QToolButton:hover {{
                background-color: {hover_bg};
            }}
            QCalendarWidget QToolButton:pressed {{
                background-color: {pressed_bg};
            }}
            QCalendarWidget QMenu {{
                background-color: {bg_color};
                color: {text_color};
                border: 1px solid {border_color};
            }}
            QCalendarWidget QSpinBox {{
                background-color: {bg_color};
                color: {text_color};
                border: 1px solid {border_color};
                border-radius: 4px;
                margin-right: 4px;
            }}
            QCalendarWidget QTableView {{
                background-color: {bg_color};
                border: none;
                border-bottom-left-radius: 12px;
                border-bottom-right-radius: 12px;
                selection-background-color: {select_bg};
                selection-color: {select_fg};
                gridline-color: transparent;
            }}
            QCalendarWidget QAbstractItemView:enabled {{
                color: {text_color};
            }}
            QCalendarWidget QAbstractItemView:disabled {{
                color: {disabled_color};
            }}
        """)
        
    def update_holiday_panel_style(self):
        is_dark = isDarkTheme()
        border_color = "#3A3A3A" if is_dark else "#E5E5E5"
        bg_color = "#202020" if is_dark else "#FFFFFF"
        self.holiday_list_scroll.setStyleSheet(f"QScrollArea {{ border: 1px solid {border_color}; border-radius: 8px; background-color: {bg_color}; }}")

    def add_company(self):
        dialog = CompanyDialog(parent=self)
        if dialog.exec_() == QDialog.Accepted:
            new_data = dialog.get_data()
            self.companies.append(new_data)
            save_companies(self.companies)
            self.render_list()
            InfoBar.success("성공", f"'{new_data['name']}' 업체가 추가되었습니다.", duration=3000, parent=self)
            
    def edit_company(self, index):
        if index < 0 or index >= len(self.companies):
            return
        company_data = self.companies[index]
        dialog = CompanyDialog(company_data, parent=self)
        if dialog.exec_() == QDialog.Accepted:
            updated_data = dialog.get_data()
            self.companies[index] = updated_data
            save_companies(self.companies)
            self.render_list()
            InfoBar.success("성공", f"'{updated_data['name']}' 업체 정보가 수정되었습니다.", duration=3000, parent=self)
            
    def delete_company(self, index):
        if index < 0 or index >= len(self.companies):
            return
        company_name = self.companies[index].get('name', '이름 없음')
        
        w = MessageBox('업체 삭제', f"'{company_name}' 업체를 삭제하시겠습니까?", self)
        if w.exec_():
            self.companies.pop(index)
            save_companies(self.companies)
            self.render_list()
            InfoBar.success("성공", f"'{company_name}' 업체가 삭제되었습니다.", duration=3000, parent=self)

class ExposureCard(QFrame):
    def __init__(self, match_data, parent=None):
        super().__init__(parent)
        self.match_data = match_data  # {'company': ..., 'keyword': ..., 'ranks': [...], 'screenshots': [...], 'folder': ...}
        self.setObjectName("ExposureCard")
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 14, 16, 14)
        layout.setSpacing(10)
        
        # Header (업체명 + 키워드 + 바로가기)
        header_layout = QHBoxLayout()
        header_layout.setSpacing(12)
        
        self.company_label = QLabel(match_data.get('company', '업체명 없음'), self)
        self.company_label.setFont(QFont("SUIT", 12, QFont.Bold))
        header_layout.addWidget(self.company_label)
        
        self.keyword_label = QLabel(match_data.get('keyword', '키워드 없음'), self)
        self.keyword_label.setFont(QFont("SUIT", 11, QFont.Bold))
        header_layout.addWidget(self.keyword_label)
        
        header_layout.addStretch(1)
        
        self.shortcut_btn = PushButton("스크린샷 바로가기", self)
        self.shortcut_btn.clicked.connect(self.open_folder)
        header_layout.addWidget(self.shortcut_btn)
        
        layout.addLayout(header_layout)
        
        # Ranks box
        self.ranks_box = QFrame(self)
        self.ranks_box.setObjectName("RanksBox")
        box_layout = QHBoxLayout(self.ranks_box)
        box_layout.setContentsMargins(12, 6, 12, 6)
        box_layout.setSpacing(16)
        
        # Add rank buttons
        self.rank_buttons = []
        for rank, path in match_data.get('screenshots', []):
            btn = PushButton(f"{rank}위", self.ranks_box)
            btn.setFont(QFont("SUIT", 11, QFont.Bold))
            btn.clicked.connect(lambda checked, p=path: self.open_screenshot(p))
            box_layout.addWidget(btn)
            self.rank_buttons.append((btn, path))
            
        box_layout.addStretch(1)
        layout.addWidget(self.ranks_box)
        
        self.update_style()
        qconfig.themeChanged.connect(self.update_style)
        
    def update_style(self):
        is_dark = isDarkTheme()
        
        # Main card colors
        bg_color = "#2C2C2C" if is_dark else "#F3F3F3"
        border_color = "#3A3A3A" if is_dark else "#E5E5E5"
        text_color = "#FFFFFF" if is_dark else "#000000"
        sub_text_color = "#CCCCCC" if is_dark else "#333333"
        
        # Ranks box colors
        box_bg = "#161616" if is_dark else "#E5E5E5"
        btn_fg = "#FFFFFF" if is_dark else "#000000"
        btn_hover_fg = "#E0E0E0" if is_dark else "#0078D4"
        
        self.setStyleSheet(f"""
            QFrame#ExposureCard {{
                background-color: {bg_color};
                border: 1px solid {border_color};
                border-radius: 12px;
            }}
        """)
        self.company_label.setStyleSheet(f"color: {text_color}; background: transparent; border: none;")
        self.keyword_label.setStyleSheet(f"color: {sub_text_color}; background: transparent; border: none;")
        
        # Shortcut button style
        link_color = "#60CDFF" if is_dark else "#0078D4"
        link_hover_color = "#A6E2FF" if is_dark else "#005A9E"
        self.shortcut_btn.setStyleSheet(f"""
            QPushButton {{
                color: {link_color};
                background: transparent;
                border: none;
                font-family: 'SUIT';
                font-size: 11px;
                font-weight: bold;
                padding: 0px;
            }}
            QPushButton:hover {{
                color: {link_hover_color};
                text-decoration: underline;
                background: transparent;
            }}
        """)
        
        # Ranks box stylesheet
        self.ranks_box.setStyleSheet(f"""
            QFrame#RanksBox {{
                background-color: {box_bg};
                border-radius: 8px;
                border: none;
            }}
        """)
        
        # Rank buttons style
        for btn, _ in self.rank_buttons:
            btn.setStyleSheet(f"""
                QPushButton {{
                    color: {btn_fg};
                    background: transparent;
                    border: none;
                    padding: 0px;
                }}
                QPushButton:hover {{
                    text-decoration: underline;
                    color: {btn_hover_fg};
                    background: transparent;
                }}
            """)
            
    def open_folder(self):
        folder = self.match_data.get('folder', '')
        if folder and os.path.exists(folder):
            os.startfile(folder)
        else:
            InfoBar.warning("안내", "스크린샷이 없습니다!", duration=3000, parent=self)
            
    def open_screenshot(self, filepath):
        if filepath and os.path.exists(filepath):
            subprocess.Popen(f'explorer /select,"{os.path.abspath(filepath)}"')
        else:
            InfoBar.warning("안내", "스크린샷이 없습니다!", duration=3000, parent=self)
            
    def closeEvent(self, event):
        try:
            qconfig.themeChanged.disconnect(self.update_style)
        except Exception:
            pass
        super().closeEvent(event)


class NewsWorker(QThread):
    finished = pyqtSignal(list)
    error = pyqtSignal(str)

    def run(self):
        url = "https://www.dmktnews.com/kwa-home"
        try:
            headers = {"User-Agent": "Mozilla/5.0"}
            response = requests.get(url, headers=headers, timeout=10)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, 'html.parser')
            
            news_items = []
            for a in soup.find_all('a'):
                href = a.get('href', '')
                title_elem = a.find('div', class_='tit') or a.find('h3') or a.find('strong')
                title = title_elem.text.strip() if title_elem else a.text.strip()
                
                if "뉴스202" in title:
                    title = title.split("뉴스202")[0].strip()
                    
                if href.startswith('/news/') and title and len(title) > 5:
                    full_url = f"https://www.dmktnews.com{href}"
                    if not any(item['url'] == full_url for item in news_items):
                        news_items.append({'title': title, 'url': full_url})
                    if len(news_items) >= 6:
                        break
                        
            self.finished.emit(news_items)
        except Exception as e:
            self.error.emit(str(e))


class FeatureCard(QFrame):
    clicked = pyqtSignal()

    def __init__(self, icon, title, content, parent=None):
        super().__init__(parent)
        self.icon_widget = IconWidget(icon)
        
        self.title_label = QLabel(title)
        self.title_label.setFont(QFont("SUIT", 14, QFont.Bold))
        self.title_label.setStyleSheet("color: #EAEAEA; background: transparent;")
        
        self.content_label = QLabel(content)
        self.content_label.setFont(QFont("SUIT", 10))
        self.content_label.setStyleSheet("color: rgba(255, 255, 255, 0.6); background: transparent;")
        
        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(24, 24, 24, 24)
        
        header_layout = QHBoxLayout()
        self.icon_widget.setFixedSize(28, 28)
        header_layout.addWidget(self.icon_widget)
        header_layout.addWidget(self.title_label)
        header_layout.addStretch(1)
        
        layout.addLayout(header_layout)
        layout.addWidget(self.content_label)
        layout.addStretch(1)
        
        self.setFixedSize(280, 150)
        self.setObjectName("FeatureCard")
        self.setStyleSheet("""
            #FeatureCard {
                background-color: rgba(30, 30, 35, 0.85);
                border: 1px solid rgba(255, 255, 255, 0.12);
                border-radius: 16px;
            }
            #FeatureCard:hover {
                background-color: rgba(45, 45, 50, 0.9);
                border: 1px solid rgba(255, 255, 255, 0.25);
            }
        """)
        self.setCursor(Qt.PointingHandCursor)

    def mousePressEvent(self, event):
        super().mousePressEvent(event)
        self.clicked.emit()


class HomeViewWidget(QWidget):
    def __init__(self, bg_path, parent=None):
        super().__init__(parent)
        self.bg_pixmap = QPixmap(bg_path)
        
    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.SmoothPixmapTransform)
        
        if self.parentWidget() and not self.bg_pixmap.isNull():
            viewport_size = self.parentWidget().size()
            scaled_pixmap = self.bg_pixmap.scaled(viewport_size, Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)
            
            x_offset = (viewport_size.width() - scaled_pixmap.width()) // 2
            y_offset = (viewport_size.height() - scaled_pixmap.height()) // 2
            
            scroll_x = -self.x()
            scroll_y = -self.y()
            
            painter.drawPixmap(scroll_x + x_offset, scroll_y + y_offset, scaled_pixmap)
        else:
            super().paintEvent(event)


class HomeInterface(ScrollArea):
    def __init__(self, main_window, parent=None):
        super().__init__(parent=parent)
        self.main_window = main_window
        self.setObjectName("HomeInterface")
        
        # Determine time of day
        hour = datetime.datetime.now().hour
        bg_image_name = "bg_night.png"
        if 5 <= hour < 12:
            greeting = "좋은 아침입니다"
            time_icon = "🌅"
            bg_image_name = "bg_morning.png"
        elif 12 <= hour < 17:
            greeting = "좋은 오후입니다"
            time_icon = "☀️"
            bg_image_name = "bg_afternoon.png"
        else:
            greeting = "좋은 저녁입니다"
            time_icon = "🌙"
            
        base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
        bg_path = os.path.join(base_dir, "assets", "images", bg_image_name).replace('\\', '/')
        
        self.view = HomeViewWidget(bg_path, self)
        self.view.setObjectName("HomeView")
        self.setWidget(self.view)
        self.setWidgetResizable(True)
        
        self.setStyleSheet("""
            QScrollArea { border: none; background: transparent; }
            #HomeInterface { background: transparent; }
        """)
        
        self.overlay = QFrame()
        self.overlay.setStyleSheet("background-color: rgba(10, 10, 15, 0.65);")
        
        self.base_layout = QVBoxLayout(self.view)
        self.base_layout.setContentsMargins(0, 0, 0, 0)
        self.base_layout.addWidget(self.overlay)
        
        self.main_vbox = QVBoxLayout(self.overlay)
        self.main_vbox.setContentsMargins(60, 60, 60, 60)
        self.main_vbox.setSpacing(50)
        
        # HEADER
        self.header_layout = QHBoxLayout()
        
        self.header_left_layout = QHBoxLayout()
        self.header_left_layout.setSpacing(20)
        
        self.icon_label = QLabel(time_icon)
        self.icon_label.setFont(QFont("Segoe UI Emoji", 48))
        self.icon_label.setStyleSheet("background: transparent;")
        
        self.greeting_vbox = QVBoxLayout()
        self.greeting_label = QLabel(greeting)
        self.greeting_label.setFont(QFont("SUIT", 42, QFont.Bold))
        self.greeting_label.setStyleSheet("color: #FFFFFF; background: transparent;")
        
        self.subtitle_label = QLabel("푸름애드 블로그 관리 프로그램에 오신 것을 환영합니다.")
        self.subtitle_label.setFont(QFont("SUIT", 16))
        self.subtitle_label.setStyleSheet("color: rgba(255, 255, 255, 0.8); background: transparent;")
        
        self.greeting_vbox.addWidget(self.greeting_label)
        self.greeting_vbox.addWidget(self.subtitle_label)
        
        self.header_left_layout.addWidget(self.icon_label)
        self.header_left_layout.addLayout(self.greeting_vbox)
        
        self.version_label = QLabel(f"Program ver : {__version__}")
        self.version_label.setFont(QFont("SUIT", 12))
        self.version_label.setStyleSheet("color: rgba(255, 255, 255, 0.6); background: transparent;")
        self.version_label.setAlignment(Qt.AlignRight | Qt.AlignTop)
        
        self.header_layout.addLayout(self.header_left_layout)
        self.header_layout.addStretch(1)
        self.header_layout.addWidget(self.version_label)
        
        self.main_vbox.addLayout(self.header_layout)
        
        # BODY
        self.body_layout = QHBoxLayout()
        self.body_layout.setSpacing(60)
        
        # Left Panel (Features)
        self.left_panel = QVBoxLayout()
        self.left_panel.setSpacing(20)
        self.feature_title = QLabel("빠른 실행")
        self.feature_title.setFont(QFont("SUIT", 18, QFont.Bold))
        self.feature_title.setStyleSheet("color: rgba(255, 255, 255, 0.9); background: transparent;")
        self.left_panel.addWidget(self.feature_title)
        
        self.grid_layout = QGridLayout()
        self.grid_layout.setSpacing(20)
        
        self.scraper_card = FeatureCard(FluentIcon.DOCUMENT, "블로그 순위 체크", "키워드 기반 순위 검색 및\\n데이터 추출")
        self.place_card = FeatureCard(getattr(FluentIcon, "POI", FluentIcon.SEARCH), "플레이스 순위 체크", "모바일 플레이스 순위 검색\\n(광고 제외)")
        self.company_card = FeatureCard(FluentIcon.PEOPLE, "업체 리스트", "저장된 관리 업체 목록 확인 및\\n수정")
        self.index_card = FeatureCard(getattr(FluentIcon, "PIE_SINGLE", FluentIcon.DOCUMENT), "지수 체크", "블로그 지수 분석 및 확인")
        self.spell_card = FeatureCard(FluentIcon.EDIT, "맞춤법 검사기", "네이버 맞춤법 검사기 기반\\n원고 교정")
        
        self.scraper_card.clicked.connect(lambda: self.main_window.switchTo(self.main_window.scraper_interface))
        self.place_card.clicked.connect(lambda: self.main_window.switchTo(self.main_window.place_scraper_interface))
        self.company_card.clicked.connect(lambda: self.main_window.switchTo(self.main_window.company_list_interface))
        self.index_card.clicked.connect(lambda: self.main_window.switchTo(self.main_window.index_check_interface))
        self.spell_card.clicked.connect(lambda: self.main_window.switchTo(self.main_window.spell_check_interface))
        
        self.grid_layout.addWidget(self.scraper_card, 0, 0)
        self.grid_layout.addWidget(self.place_card, 0, 1)
        self.grid_layout.addWidget(self.company_card, 1, 0)
        self.grid_layout.addWidget(self.index_card, 1, 1)
        self.grid_layout.addWidget(self.spell_card, 2, 0)
        
        self.left_panel.addLayout(self.grid_layout)
        self.left_panel.addStretch(1)
        
        # Right Panel (News)
        self.right_panel = QVBoxLayout()
        self.right_panel.setSpacing(20)
        self.news_title = QLabel("마케팅 주요 뉴스")
        self.news_title.setFont(QFont("SUIT", 18, QFont.Bold))
        self.news_title.setStyleSheet("color: rgba(255, 255, 255, 0.9); background: transparent;")
        self.right_panel.addWidget(self.news_title)
        
        self.news_container = QFrame()
        self.news_container.setObjectName("NewsContainer")
        self.news_container.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.news_container.setStyleSheet("""
            #NewsContainer {
                background-color: rgba(30, 30, 35, 0.85);
                border-radius: 16px;
                border: 1px solid rgba(255, 255, 255, 0.12);
            }
        """)
        self.news_layout = QVBoxLayout(self.news_container)
        self.news_layout.setContentsMargins(30, 30, 30, 30)
        self.news_layout.setSpacing(16)
        
        self.loading_label = QLabel("뉴스를 불러오는 중입니다...")
        self.loading_label.setFont(QFont("SUIT", 11))
        self.loading_label.setStyleSheet("color: rgba(255, 255, 255, 0.6); background: transparent;")
        self.news_layout.addWidget(self.loading_label)
        self.news_layout.addStretch(1)
        
        self.right_panel.addWidget(self.news_container)
        
        self.body_layout.addLayout(self.left_panel, 5)
        self.body_layout.addLayout(self.right_panel, 5)
        
        self.main_vbox.addLayout(self.body_layout)
        self.main_vbox.addStretch(1)
        
        # Fetch News
        self.worker = NewsWorker()
        self.worker.finished.connect(self.on_news_fetched)
        self.worker.error.connect(self.on_news_error)
        self.worker.start()
        
    def on_news_fetched(self, news_items):
        self.loading_label.hide()
        for i in reversed(range(self.news_layout.count())): 
            widget = self.news_layout.itemAt(i).widget()
            if widget is not None and widget != self.loading_label:
                widget.deleteLater()
                
        for item in news_items:
            link = HyperlinkButton(item['url'], item['title'])
            link.setFont(QFont("SUIT", 11))
            link.setStyleSheet("""
                HyperlinkButton {
                    color: #60CDFF;
                    background: transparent;
                    border: none;
                    text-align: left;
                    padding: 4px;
                }
                HyperlinkButton:hover {
                    color: #A6D8FF;
                    background-color: rgba(255, 255, 255, 0.05);
                    border-radius: 6px;
                }
            """)
            self.news_layout.insertWidget(self.news_layout.count() - 1, link)
            
    def on_news_error(self, err):
        self.loading_label.setText(f"뉴스를 불러오는데 실패했습니다: {err}")


class SettingInterface(ScrollArea):
    def __init__(self, parent=None):
        super().__init__(parent=parent)
        self.setObjectName("SettingInterface")
        
        # Scroll Area configuration
        self.setWidgetResizable(True)
        self.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        
        self.scroll_content = QWidget()
        self.scroll_content.setObjectName("ScrollContent")
        self.scroll_content.setStyleSheet("QWidget#ScrollContent { background: transparent; }")
        self.scroll_layout = QVBoxLayout(self.scroll_content)
        self.scroll_layout.setContentsMargins(36, 36, 36, 36)
        
        self.setWidget(self.scroll_content)
        
        # Title
        self.title_label = TitleLabel("설정", self.scroll_content)
        self.scroll_layout.addWidget(self.title_label)
        self.scroll_layout.addSpacing(20)
        
        # Update Card
        self.update_card = CardWidget(self.scroll_content)
        self.update_layout = QHBoxLayout(self.update_card)
        self.update_layout.setContentsMargins(20, 20, 20, 20)
        
        self.version_info = BodyLabel(f"현재 버전: v{__version__}", self.update_card)
        self.update_btn = PrimaryPushButton("업데이트 확인", self.update_card)
        self.update_btn.clicked.connect(lambda: check_for_updates(manual_check=True))
        
        self.update_layout.addWidget(self.version_info)
        self.update_layout.addStretch(1)
        self.update_layout.addWidget(self.update_btn)
        
        self.scroll_layout.addWidget(self.update_card)
        self.scroll_layout.addStretch(1)


class ScraperInterface(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent=parent)
        self.setObjectName("ScraperInterface")
        
        self.base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
        
        self.workspace_dir = os.path.join(self.base_dir, 'Workspace')
        os.makedirs(self.workspace_dir, exist_ok=True)
        
        self.excel_filename = "키워드_순위_작업표.xlsx"
        self.excel_path = os.path.join(self.workspace_dir, self.excel_filename)
        
        self.data_dir = os.path.join(self.base_dir, 'Data')
        self.config_file = os.path.join(self.data_dir, 'config.ini')
        
        self.config = configparser.ConfigParser()
        self.global_driver_path = ""
        
        self.init_ui()
        self.check_environment()

    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(36, 36, 36, 36)
        main_layout.setSpacing(16)

        self.title_label = TitleLabel("블로그 순위 체크")
        main_layout.addWidget(self.title_label)

        # Horizontal Split Layout
        split_layout = QHBoxLayout()
        split_layout.setSpacing(24)
        
        # Left Panel (Existing UI controls)
        self.left_panel = QWidget(self)
        left_layout = QVBoxLayout(self.left_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(16)

        control_layout = QHBoxLayout()
        
        self.sheet_combo = ComboBox()
        self.sheet_combo.setMinimumWidth(200)
        control_layout.addWidget(SubtitleLabel("작업 시트 선택:"))
        control_layout.addWidget(self.sheet_combo)
        
        self.count_spinbox = SpinBox()
        self.count_spinbox.setRange(1, 100)
        control_layout.addStretch(1)
        control_layout.addWidget(SubtitleLabel("탐색 목표 개수:"))
        control_layout.addWidget(self.count_spinbox)
        
        self.screenshot_switch = SwitchButton()
        self.screenshot_switch.setOnText("캡처 켜짐")
        self.screenshot_switch.setOffText("캡처 꺼짐")
        font = QFont("SUIT", 10, QFont.Bold)
        self.screenshot_switch.setFont(font)
        if hasattr(self.screenshot_switch, 'label'):
            self.screenshot_switch.label.setFont(font)
            self.screenshot_switch.label.setStyleSheet("font-family: 'SUIT'; font-weight: bold;")
        else:
            self.screenshot_switch.setStyleSheet("QLabel { font-family: 'SUIT'; font-weight: bold; }")
        control_layout.addStretch(1)
        control_layout.addWidget(self.screenshot_switch)
        
        left_layout.addLayout(control_layout)

        self.loading_container = QWidget()
        loading_layout = QHBoxLayout(self.loading_container)
        loading_layout.setAlignment(Qt.AlignCenter)
        loading_layout.setContentsMargins(0, 0, 0, 0)
        
        self.loading_ring = IndeterminateProgressRing()
        self.loading_ring.setFixedSize(25, 25)
        self.loading_label = SubtitleLabel("크롬 드라이버를 백그라운드에서 점검/동기화 중입니다...")
        
        loading_layout.addWidget(self.loading_ring)
        loading_layout.addWidget(self.loading_label)
        left_layout.addWidget(self.loading_container)
        self.loading_container.hide()

        button_layout = QHBoxLayout()
        
        self.start_btn = PushButton("순위 체크 시작")
        self.start_btn.setStyleSheet("""
            QPushButton {
                background-color: #3CA0F0;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-family: 'SUIT';
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #59B4FF;
            }
            QPushButton:pressed {
                background-color: #268CD9;
            }
            QPushButton:disabled {
                background-color: #2C2C2C;
                color: #666666;
            }
        """)
        self.start_btn.clicked.connect(self.start_scraping)
        
        self.open_excel_btn = PushButton("엑셀 파일 열기")
        self.open_excel_btn.setStyleSheet("""
            QPushButton {
                background-color: #107C41;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-family: 'SUIT';
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1F9A55;
            }
            QPushButton:pressed {
                background-color: #0E6233;
            }
            QPushButton:disabled {
                background-color: #2C2C2C;
                color: #666666;
            }
        """)
        self.open_excel_btn.clicked.connect(self.open_excel_file)
        
        button_layout.addWidget(self.start_btn)
        button_layout.addWidget(self.open_excel_btn)
        
        left_layout.addLayout(button_layout)

        self.console_output = TextEdit()
        self.console_output.setReadOnly(True)
        left_layout.addWidget(self.console_output)

        split_layout.addWidget(self.left_panel, 3) # stretch factor 3

        # Right Panel (Exposure Status Panel)
        self.right_panel = QWidget(self)
        right_layout = QVBoxLayout(self.right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(12)
        
        right_title = SubtitleLabel("실시간 노출 현황")
        right_title.setFont(QFont("SUIT", 12, QFont.Bold))
        right_layout.addWidget(right_title)
        
        self.exposure_scroll = ScrollArea(self.right_panel)
        self.exposure_scroll.setWidgetResizable(True)
        
        self.exposure_content = QWidget()
        self.exposure_content.setObjectName("ExposureContent")
        self.exposure_content.setStyleSheet("QWidget#ExposureContent { background-color: transparent; }")
        self.exposure_layout = QVBoxLayout(self.exposure_content)
        self.exposure_layout.setContentsMargins(12, 12, 12, 12)
        self.exposure_layout.setSpacing(12)
        self.exposure_layout.addStretch(1)
        
        self.exposure_scroll.setWidget(self.exposure_content)
        right_layout.addWidget(self.exposure_scroll)
        
        split_layout.addWidget(self.right_panel, 2) # stretch factor 2

        main_layout.addLayout(split_layout)

        # Connect to theme changed for right panel styling
        self.update_right_panel_style()
        qconfig.themeChanged.connect(self.update_right_panel_style)

    def update_right_panel_style(self):
        is_dark = isDarkTheme()
        border_color = "#3A3A3A" if is_dark else "#E5E5E5"
        bg_color = "#202020" if is_dark else "#FFFFFF"
        self.exposure_scroll.setStyleSheet(f"QScrollArea {{ border: 1px solid {border_color}; border-radius: 8px; background-color: {bg_color}; }}")

    def clear_exposure_cards(self):
        while self.exposure_layout.count():
            child = self.exposure_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()
        self.exposure_layout.addStretch(1)

    def add_exposure_card(self, match_data):
        card = ExposureCard(match_data, self.exposure_content)
        insert_idx = max(0, self.exposure_layout.count() - 1)
        self.exposure_layout.insertWidget(insert_idx, card)

    def open_excel_file(self):
        if not os.path.exists(self.excel_path):
            InfoBar.error("오류", "엑셀 파일이 아직 생성되지 않았습니다.", duration=3000, position=InfoBarPosition.TOP, parent=self)
            return
            
        msg_box = QMessageBox(self)
        msg_box.setIcon(QMessageBox.Warning)
        msg_box.setWindowTitle("엑셀 파일 수정 주의사항")
        msg_box.setText("작업 시트를 수정하신 후, 반드시 엑셀 프로그램에서 '저장(Ctrl+S)'을 누른 다음 창을 닫고 크롤링을 시작해주세요.")
        msg_box.setInformativeText("저장하지 않고 닫거나 엑셀 창을 켜둔 채로 크롤링을 시작하면, 이전 데이터가 불러와지거나 저장 시 권한 오류가 발생합니다.")
        msg_box.setStandardButtons(QMessageBox.Ok)
        msg_box.exec_()
            
        try:
            if os.name == 'nt':
                os.startfile(self.excel_path)
            elif sys.platform == 'darwin':
                subprocess.call(['open', self.excel_path])
            else:
                subprocess.call(['xdg-open', self.excel_path])
            self.append_log("[안내] 엑셀 파일을 열었습니다. 작업 후 반드시 저장하고 엑셀 창을 닫아주세요.")
        except Exception as e:
            self.append_log(f"[오류] 엑셀 파일을 열 수 없습니다: {e}")

    def append_log(self, text):
        self.console_output.append(text)
        scrollbar = self.console_output.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    def check_environment(self):
        self.append_log("[시스템] 초기 환경을 점검합니다...")
        
        if os.path.exists(self.config_file):
            self.config.read(self.config_file, encoding='utf-8')
            display_count = int(self.config['SETTINGS'].get('DISPLAY_COUNT', 10))
            capture_str = self.config['SETTINGS'].get('CAPTURE_SCREENSHOT', 'False').strip().lower()
            capture_option = capture_str in ['true', '1', 'y', 'yes', 't']
        else:
            display_count = 10
            capture_option = False
            self.config['SETTINGS'] = {'DISPLAY_COUNT': '10', 'CAPTURE_SCREENSHOT': 'False'}
            os.makedirs(self.data_dir, exist_ok=True)
            with open(self.config_file, 'w', encoding='utf-8') as f:
                self.config.write(f)
                
        self.count_spinbox.setValue(display_count)
        self.screenshot_switch.setChecked(capture_option)

        if not os.path.exists(self.excel_path):
            df_template = pd.DataFrame(columns=['키워드', '목표블로그', '관련도순_순위', '최근조회일시'])
            with pd.ExcelWriter(self.excel_path, engine='openpyxl') as writer:
                df_template.to_excel(writer, index=False, sheet_name="기본작업표")
            self.format_excel_file(self.excel_path, "기본작업표")
            self.append_log(f"[시스템] 'Workspace/{self.excel_filename}' 템플릿 파일이 생성되었습니다.")
        
        self.load_excel_sheets()
        self.backup_excel()

        self.start_btn.setEnabled(False)
        self.open_excel_btn.setEnabled(False)
        self.loading_container.show()
        self.loading_ring.start()
        
        self.driver_worker = DriverInitWorker()
        self.driver_worker.finished.connect(self.on_driver_ready)
        self.driver_worker.error.connect(self.on_driver_error)
        self.driver_worker.start()

    def on_driver_ready(self, path):
        self.global_driver_path = path
        self.loading_ring.stop()
        self.loading_container.hide()
        self.append_log("[시스템] 크롬 드라이버 동기화 및 준비 완료.")
        self.start_btn.setEnabled(True)
        self.open_excel_btn.setEnabled(True)

    def on_driver_error(self, err_msg):
        self.loading_ring.stop()
        self.loading_container.hide()
        self.append_log(f"[치명적 오류] 크롬 드라이버 설치 실패: {err_msg}")
        InfoBar.error("오류", "크롬 드라이버를 설치할 수 없습니다.", duration=5000, position=InfoBarPosition.TOP, parent=self)

    def load_excel_sheets(self):
        try:
            wb = openpyxl.load_workbook(self.excel_path, read_only=True)
            for name in wb.sheetnames:
                self.sheet_combo.addItem(name)
            wb.close()
        except Exception as e:
            self.append_log(f"[오류] 엑셀 파일을 읽을 수 없습니다 (파일이 열려있는지 확인): {e}")

    def format_excel_file(self, file_path, sheet_name):
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 15
        wb.save(file_path)

    def backup_excel(self):
        try:
            backup_dir = os.path.join(self.workspace_dir, '.backup')
            os.makedirs(backup_dir, exist_ok=True)
            if os.name == 'nt': os.system(f'attrib +h "{backup_dir}"')
            
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            backup_path = os.path.join(backup_dir, f"키워드_순위_작업표_backup_{timestamp}.xlsx")
            shutil.copy2(self.excel_path, backup_path)
            
            now = time.time()
            deleted = 0
            for file_name in os.listdir(backup_dir):
                file_path = os.path.join(backup_dir, file_name)
                if os.path.isfile(file_path) and file_name.endswith('.xlsx'):
                    if now - os.path.getmtime(file_path) > 30 * 24 * 60 * 60:
                        os.remove(file_path)
                        deleted += 1
            if deleted > 0: self.append_log(f"[시스템] 30일 경과 백업 파일 {deleted}개 삭제 완료.")
        except Exception:
            pass

    def start_scraping(self):
        selected_sheet = self.sheet_combo.currentText()
        if not selected_sheet:
            InfoBar.error("오류", "작업할 시트를 선택해주세요.", duration=3000, position=InfoBarPosition.TOP, parent=self)
            return

        if os.path.exists(self.excel_path):
            try:
                with open(self.excel_path, 'a') as f:
                    pass
            except PermissionError:
                self.append_log("[오류] 엑셀 파일이 열려있어 작업을 시작할 수 없습니다. 창을 닫아주세요.")
                InfoBar.error("작업 거부", "엑셀 파일이 열려있습니다. 창을 닫은 후 다시 시도해주세요.", duration=5000, position=InfoBarPosition.TOP, parent=self)
                return
            except Exception as e:
                self.append_log(f"[오류] 엑셀 파일 접근 검사 실패: {e}")
                return

        self.start_btn.setEnabled(False)
        self.open_excel_btn.setEnabled(False)
        self.sheet_combo.setEnabled(False)
        self.console_output.clear()
        self.clear_exposure_cards()
        
        display_count = self.count_spinbox.value()
        capture_screenshot = self.screenshot_switch.isChecked()
        is_custom_excel = selected_sheet != "기본작업표"

        self.config['SETTINGS'] = {'DISPLAY_COUNT': str(display_count), 'CAPTURE_SCREENSHOT': str(capture_screenshot)}
        with open(self.config_file, 'w', encoding='utf-8') as f:
            self.config.write(f)

        self.worker = ScraperWorker(
            self.excel_path, selected_sheet, display_count, 
            capture_screenshot, self.global_driver_path, is_custom_excel
        )
        self.worker.signals.log.connect(self.append_log)
        self.worker.signals.error.connect(self.show_error)
        self.worker.signals.match_found.connect(self.add_exposure_card)
        self.worker.signals.finished.connect(self.on_scraping_finished)
        self.worker.start()

    def show_error(self, err_msg):
        InfoBar.error("작업 중단", err_msg, duration=5000, position=InfoBarPosition.TOP, parent=self)
        self.on_scraping_finished()

    def on_scraping_finished(self):
        self.start_btn.setEnabled(True)
        self.open_excel_btn.setEnabled(True)
        self.sheet_combo.setEnabled(True)
        self.append_log("\n[안내] 모든 루틴이 종료되었습니다.")
        InfoBar.success("완료", "크롤링 작업이 성공적으로 종료되었습니다.", duration=4000, position=InfoBarPosition.TOP, parent=self)

# --- [신설] 네이버 블로그 통계 자동수집 워커 ---
class StatsScraperWorker(QThread):
    finished = pyqtSignal(dict)  # Emits parsed stats dict
    error = pyqtSignal(str)      # Emits error message
    status = pyqtSignal(str)     # Emits status log message

    def __init__(self, blog_id, company_name, global_driver_path):
        super().__init__()
        self.blog_id = blog_id
        self.company_name = company_name
        self.global_driver_path = global_driver_path
        self.driver = None

    def run(self):
        try:
            self.status.emit("크롬 드라이버 준비 중...")
            if not self.global_driver_path:
                try:
                    self.global_driver_path = ChromeDriverManager().install()
                except Exception as e:
                    self.error.emit(f"드라이버 설치 실패: {e}")
                    return
            
            self.status.emit("크롬 브라우저를 구동 중입니다...")
            options = Options()
            
            # Setup Chrome User Data Profile Directory for session persistence
            base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
            profile_dir = os.path.join(base_dir, "Data", "ChromeProfiles", self.blog_id)
            os.makedirs(profile_dir, exist_ok=True)
            options.add_argument(f"--user-data-dir={os.path.abspath(profile_dir)}")
            
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--window-size=1024,768')
            options.add_argument('--disable-blink-features=AutomationControlled')
            
            service = Service(self.global_driver_path)
            self.driver = webdriver.Chrome(service=service, options=options)
            
            self.status.emit("네이버 블로그 통계 페이지로 이동 중...")
            target_url = f"https://blog.stat.naver.com/blog/visitor/daily?blogId={self.blog_id}"
            self.driver.get(target_url)
            
            # Detect if we need to login
            login_wait_start = time.time()
            logged_in = False
            
            while time.time() - login_wait_start < 120:  # Wait up to 2 minutes
                try:
                    curr_url = self.driver.current_url
                except Exception:
                    # Browser closed by user
                    self.error.emit("사용자가 브라우저를 종료했거나 통신이 끊겼습니다.")
                    return
                
                if "nid.naver.com" in curr_url:
                    self.status.emit("로그인 및 2차 인증 대기 중... (브라우저에서 로그인해 주세요)")
                elif "blog.stat.naver.com" in curr_url:
                    # Logged in, wait for page loaded
                    try:
                        WebDriverWait(self.driver, 5).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "body"))
                        )
                        logged_in = True
                        break
                    except Exception:
                        pass
                
                time.sleep(2)
                
            if not logged_in:
                self.error.emit("로그인 대기 시간이 초과되었습니다.")
                self.cleanup()
                return
                
            self.status.emit("로그인 완료 감지! 블로그 통계 데이터를 수집 중...")
            
            # Fetch visitor, views, inflow data using JS fetch on the same origin
            end_date = QDate.currentDate().toString("yyyy-MM-dd")
            start_date = QDate.currentDate().addDays(-30).toString("yyyy-MM-dd")
            
            # 1. Daily Visitors
            js_visitor = f"""
            var callback = arguments[arguments.length - 1];
            fetch('/api/visitor/daily?blogId={self.blog_id}&startDate={start_date}&endDate={end_date}')
                .then(res => res.json())
                .then(data => callback(data))
                .catch(err => callback(null));
            """
            self.driver.set_script_timeout(10)
            res_visitor = self.driver.execute_async_script(js_visitor)
            
            # 2. Daily Views
            js_views = f"""
            var callback = arguments[arguments.length - 1];
            fetch('/api/views/daily?blogId={self.blog_id}&startDate={start_date}&endDate={end_date}')
                .then(res => res.json())
                .then(data => callback(data))
                .catch(err => callback(null));
            """
            res_views = self.driver.execute_async_script(js_views)
            
            # 3. Referrer / Inflow
            js_inflow = f"""
            var callback = arguments[arguments.length - 1];
            fetch('/api/inflow/daily?blogId={self.blog_id}&startDate={start_date}&endDate={end_date}')
                .then(res => res.json())
                .then(data => callback(data))
                .catch(err => callback(null));
            """
            res_inflow = self.driver.execute_async_script(js_inflow)
            
            # Parsing daily stats
            dates = []
            visitor_cnts = []
            views_cnts = []
            inflow_paths = []
            
            # Process visitor
            try:
                raw_list = res_visitor.get('result', {}).get('dailyVisitor', [])
                for item in raw_list:
                    dt = str(item.get('date', ''))
                    if len(dt) == 8:
                        formatted_dt = f"{dt[4:6]}/{dt[6:8]}"
                    else:
                        formatted_dt = dt
                    dates.append(formatted_dt)
                    visitor_cnts.append(item.get('visitor', 0))
            except Exception:
                pass
                
            # Process views
            try:
                raw_list = res_views.get('result', {}).get('dailyViews', [])
                for item in raw_list:
                    views_cnts.append(item.get('views', 0))
            except Exception:
                pass
                
            # Adjust arrays size to match
            min_len = min(len(dates), len(visitor_cnts), len(views_cnts))
            if min_len > 0:
                dates = dates[:min_len]
                visitor_cnts = visitor_cnts[:min_len]
                views_cnts = views_cnts[:min_len]
            
            # Process inflow paths
            try:
                raw_list = res_inflow.get('result', {}).get('inflow', [])
                for item in raw_list[:5]:  # Top 5
                    inflow_paths.append({
                        'name': item.get('source', '알 수 없음'),
                        'value': item.get('count', 0)
                    })
            except Exception:
                pass
                
            if not dates:
                self.error.emit("데이터를 파싱하지 못했거나 권한이 없습니다.")
                self.cleanup()
                return
                
            stats_data = {
                'last_updated': time.strftime("%Y-%m-%d %H:%M:%S"),
                'dates': dates,
                'visitor': visitor_cnts,
                'views': views_cnts,
                'inflow': inflow_paths
            }
            
            # Save to JSON file
            data_dir = os.path.join(base_dir, "Data")
            stats_path = os.path.join(data_dir, f"stats_{self.company_name}.json")
            with open(stats_path, 'w', encoding='utf-8') as f:
                json.dump(stats_data, f, ensure_ascii=False, indent=4)
                
            self.finished.emit(stats_data)
            
        except Exception as e:
            self.error.emit(f"오류가 발생했습니다: {e}")
        finally:
            self.cleanup()

    def cleanup(self):
        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass
            self.driver = None


class IndexCheckInterface(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent=parent)
        self.setObjectName("IndexCheckInterface")
        
        self.base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
        self.companies = []
        self.worker = None
        
        # Main layout
        layout = QVBoxLayout(self)
        layout.setContentsMargins(36, 36, 36, 36)
        layout.setSpacing(20)
        
        # Title
        title_label = TitleLabel("지수 체크 (블로그 통계 대시보드)", self)
        layout.addWidget(title_label)
        
        # Top Control Bar
        control_card = CardWidget(self)
        control_layout = QHBoxLayout(control_card)
        control_layout.setContentsMargins(20, 16, 20, 16)
        control_layout.setSpacing(16)
        
        self.combo_lbl = QLabel("업체 선택:", control_card)
        self.combo_lbl.setFont(QFont("SUIT", 10, QFont.Bold))
        is_dark = isDarkTheme()
        self.combo_lbl.setStyleSheet(f"color: {'#FFFFFF' if is_dark else '#000000'}; background: transparent;")
        control_layout.addWidget(self.combo_lbl)
        
        self.company_combo = ComboBox(control_card)
        self.company_combo.setMinimumWidth(200)
        self.company_combo.currentIndexChanged.connect(self.on_company_changed)
        control_layout.addWidget(self.company_combo)
        
        self.update_btn = PushButton("통계 데이터 갱신", control_card)
        self.update_btn.setMinimumWidth(150)
        self.update_btn.clicked.connect(self.update_stats)
        control_layout.addWidget(self.update_btn)
        
        self.loading_ring = IndeterminateProgressRing(control_card)
        self.loading_ring.setFixedSize(24, 24)
        self.loading_ring.hide()
        control_layout.addWidget(self.loading_ring)
        
        self.status_label = QLabel(control_card)
        self.status_label.setFont(QFont("SUIT", 9))
        self.status_label.setStyleSheet("color: #AAAAAA; background: transparent;")
        control_layout.addWidget(self.status_label)
        
        control_layout.addStretch(1)
        
        self.time_label = QLabel("업데이트 기록 없음", control_card)
        self.time_label.setFont(QFont("SUIT", 9))
        self.time_label.setStyleSheet("color: #888888; background: transparent;")
        control_layout.addWidget(self.time_label)
        
        layout.addWidget(control_card)
        
        if QWebEngineView is None:
            error_lbl = SubtitleLabel("PyQtWebEngine 모듈이 설치되지 않아 차트를 렌더링할 수 없습니다.\n앱을 재시작하면 자동으로 설치됩니다.", self)
            layout.addWidget(error_lbl)
            layout.addStretch(1)
            return
            
        # WebEngineView for ECharts
        self.chart_view = QWebEngineView(self)
        self.chart_view.setStyleSheet("background: transparent; border-radius: 12px;")
        
        self.init_chart()
        layout.addWidget(self.chart_view, 1) # stretch 1
        
        qconfig.themeChanged.connect(self.update_chart_theme)
        
    def showEvent(self, event):
        super().showEvent(event)
        self.load_company_list()
        
    def load_company_list(self):
        prev_idx = self.company_combo.currentIndex()
        prev_text = self.company_combo.currentText()
        
        self.company_combo.blockSignals(True)
        self.company_combo.clear()
        self.companies = load_companies()
        
        for comp in self.companies:
            name = comp.get('name', '이름 없음')
            self.company_combo.addItem(name)
            
        self.company_combo.blockSignals(False)
        
        # Restore previous selection if possible
        if prev_text:
            idx = self.company_combo.findText(prev_text)
            if idx >= 0:
                self.company_combo.setCurrentIndex(idx)
            else:
                self.company_combo.setCurrentIndex(0 if self.companies else -1)
        else:
            self.company_combo.setCurrentIndex(0 if self.companies else -1)
            
        # Force refresh data trigger
        self.on_company_changed(self.company_combo.currentIndex())
        
    def on_company_changed(self, index):
        if index < 0 or index >= len(self.companies):
            self.clear_chart()
            self.time_label.setText("업체 정보가 없습니다.")
            return
            
        company_data = self.companies[index]
        company_name = company_data.get('name', '')
        
        # Load local stats JSON file
        stats_path = os.path.join(self.base_dir, "Data", f"stats_{company_name}.json")
        if os.path.exists(stats_path):
            try:
                with open(stats_path, 'r', encoding='utf-8') as f:
                    stats_data = json.load(f)
                self.time_label.setText(f"최근 업데이트: {stats_data.get('last_updated', '알 수 없음')}")
                self.render_stats(stats_data)
            except Exception:
                self.time_label.setText("데이터 로드 오류")
                self.clear_chart()
        else:
            self.time_label.setText("통계 데이터가 없습니다. 업데이트를 진행해 주세요.")
            self.clear_chart()
            
    def render_stats(self, stats_data):
        if not hasattr(self, 'chart_view'): return
        
        dates = stats_data.get('dates', [])
        visitor = stats_data.get('visitor', [])
        views = stats_data.get('views', [])
        inflow = stats_data.get('inflow', [])
        
        js_code = f"updateData({json.dumps(dates)}, {json.dumps(visitor)}, {json.dumps(views)}, {json.dumps(inflow)});"
        self.chart_view.page().runJavaScript(js_code)
        
    def clear_chart(self):
        if not hasattr(self, 'chart_view'): return
        js_code = "updateData([], [], [], []);"
        self.chart_view.page().runJavaScript(js_code)
        
    def update_stats(self):
        index = self.company_combo.currentIndex()
        if index < 0 or index >= len(self.companies):
            return
            
        company_data = self.companies[index]
        company_name = company_data.get('name', '')
        blog_id = company_data.get('blog_id', '').strip()
        
        if not blog_id:
            MessageBox("알림", "해당 업체의 네이버 블로그 ID가 설정되어 있지 않습니다.\n'업체 리스트' 탭에서 [수정]을 통해 블로그 ID를 먼저 등록해 주세요.", self).exec_()
            return
            
        # Get global driver path from scraper interface if available
        driver_path = ""
        if hasattr(self.parent(), 'scraper_interface'):
            driver_path = self.parent().scraper_interface.global_driver_path
            
        # UI update to loading state
        self.company_combo.setEnabled(False)
        self.update_btn.setEnabled(False)
        self.loading_ring.show()
        self.loading_ring.start()
        self.status_label.setText("크롬 드라이버 준비 중...")
        
        self.worker = StatsScraperWorker(blog_id, company_name, driver_path)
        self.worker.status.connect(self.on_worker_status)
        self.worker.finished.connect(self.on_worker_finished)
        self.worker.error.connect(self.on_worker_error)
        self.worker.start()
        
    def on_worker_status(self, msg):
        self.status_label.setText(msg)
        
    def on_worker_finished(self, stats_data):
        self.company_combo.setEnabled(True)
        self.update_btn.setEnabled(True)
        self.loading_ring.stop()
        self.loading_ring.hide()
        self.status_label.setText("데이터 업데이트 성공!")
        self.time_label.setText(f"최근 업데이트: {stats_data.get('last_updated', '')}")
        self.render_stats(stats_data)
        InfoBar.success("성공", f"'{self.company_combo.currentText()}' 블로그 통계 수집이 완료되었습니다.", duration=3000, parent=self)
        
    def on_worker_error(self, err_msg):
        self.company_combo.setEnabled(True)
        self.update_btn.setEnabled(True)
        self.loading_ring.stop()
        self.loading_ring.hide()
        self.status_label.setText("실패")
        MessageBox("오류 발생", f"통계 데이터 수집 중 오류가 발생했습니다:\n{err_msg}", self).exec_()
        
    def init_chart(self):
        is_dark = isDarkTheme()
        theme_str = 'dark' if is_dark else 'light'
        bg_color = '#202020' if is_dark else '#F3F3F3'
        card_bg = '#2C2C2C' if is_dark else '#FFFFFF'
        border_color = '#3A3A3A' if is_dark else '#E5E5E5'
        text_color = '#FFFFFF' if is_dark else '#000000'
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <script src="https://cdn.jsdelivr.net/npm/echarts@5.5.0/dist/echarts.min.js"></script>
            <style>
                html, body {{
                    margin: 0;
                    padding: 0;
                    width: 100%;
                    height: 100%;
                    background-color: {bg_color};
                    overflow: hidden;
                    border-radius: 12px;
                }}
                .container {{
                    display: flex;
                    width: 100%;
                    height: 100%;
                    box-sizing: border-box;
                    padding: 20px;
                    gap: 24px;
                }}
                #main {{
                    flex: 2;
                    height: 100%;
                    background-color: {card_bg};
                    border-radius: 12px;
                    border: 1px solid {border_color};
                }}
                #inflow {{
                    flex: 1;
                    height: 100%;
                    background-color: {card_bg};
                    border-radius: 12px;
                    border: 1px solid {border_color};
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div id="main"></div>
                <div id="inflow"></div>
            </div>
            <script type="text/javascript">
                var currentTheme = '{theme_str}';
                var textCol = '{text_color}';
                var mainChart = echarts.init(document.getElementById('main'), currentTheme);
                var inflowChart = echarts.init(document.getElementById('inflow'), currentTheme);
                
                var mainOption = {{
                    backgroundColor: 'transparent',
                    title: {{
                        text: '방문자수 및 조회수 추이 (최근 30일)',
                        textStyle: {{ color: textCol, fontFamily: 'SUIT, sans-serif', fontSize: 16 }},
                        left: '5%',
                        top: 20
                    }},
                    tooltip: {{
                        trigger: 'axis',
                        axisPointer: {{ type: 'cross' }}
                    }},
                    legend: {{
                        data: ['방문자수', '조회수'],
                        top: 20,
                        right: '5%',
                        textStyle: {{ color: textCol, fontFamily: 'SUIT, sans-serif' }}
                    }},
                    grid: {{
                        left: '5%',
                        right: '5%',
                        bottom: '10%',
                        top: '20%',
                        containLabel: true
                    }},
                    xAxis: [{{
                        type: 'category',
                        boundaryGap: false,
                        data: [],
                        axisLine: {{ lineStyle: {{ color: textCol, opacity: 0.5 }} }},
                        axisLabel: {{ fontFamily: 'SUIT, sans-serif', color: textCol }}
                    }}],
                    yAxis: [{{
                        type: 'value',
                        axisLine: {{ lineStyle: {{ color: textCol, opacity: 0.5 }} }},
                        splitLine: {{ lineStyle: {{ color: textCol, opacity: 0.1 }} }},
                        axisLabel: {{ fontFamily: 'SUIT, sans-serif', color: textCol }}
                    }}],
                    series: [
                        {{
                            name: '방문자수',
                            type: 'line',
                            smooth: true,
                            lineStyle: {{ width: 4, color: '#0078D4' }},
                            itemStyle: {{ color: '#0078D4' }},
                            areaStyle: {{
                                color: new echarts.graphic.LinearGradient(0, 0, 0, 1, [
                                    {{ offset: 0, color: 'rgba(0, 120, 212, 0.3)' }},
                                    {{ offset: 1, color: 'rgba(0, 120, 212, 0.0)' }}
                                    ])
                            }},
                            data: []
                        }},
                        {{
                            name: '조회수',
                            type: 'line',
                            smooth: true,
                            lineStyle: {{ width: 4, color: '#25A662' }},
                            itemStyle: {{ color: '#25A662' }},
                            areaStyle: {{
                                color: new echarts.graphic.LinearGradient(0, 0, 0, 1, [
                                    {{ offset: 0, color: 'rgba(37, 166, 98, 0.3)' }},
                                    {{ offset: 1, color: 'rgba(37, 166, 98, 0.0)' }}
                                    ])
                            }},
                            data: []
                        }}
                    ]
                }};
                
                var inflowOption = {{
                    backgroundColor: 'transparent',
                    title: {{
                        text: '유입 경로 분석 (Top 5)',
                        textStyle: {{ color: textCol, fontFamily: 'SUIT, sans-serif', fontSize: 16 }},
                        left: 'center',
                        top: 20
                    }},
                    tooltip: {{
                        trigger: 'item',
                        formatter: '{{b}}: {{c}} ({{d}}%)'
                    }},
                    legend: {{
                        bottom: 15,
                        left: 'center',
                        textStyle: {{ color: textCol, fontFamily: 'SUIT, sans-serif', fontSize: 11 }}
                    }},
                    series: [{{
                        name: '유입 수',
                        type: 'pie',
                        radius: ['35%', '60%'],
                        center: ['50%', '45%'],
                        avoidLabelOverlap: false,
                        itemStyle: {{
                            borderRadius: 8,
                            borderColor: '{card_bg}',
                            borderWidth: 2
                        }},
                        label: {{ show: false }},
                        data: []
                    }}]
                }};
                
                mainChart.setOption(mainOption);
                inflowChart.setOption(inflowOption);
                
                window.onresize = function() {{
                    mainChart.resize();
                    inflowChart.resize();
                }};
                
                function updateData(dates, visitor, views, inflow) {{
                    mainOption.xAxis[0].data = dates;
                    mainOption.series[0].data = visitor;
                    mainOption.series[1].data = views;
                    mainChart.setOption(mainOption);
                    
                    inflowOption.series[0].data = inflow;
                    inflowChart.setOption(inflowOption);
                }}
                
                function updateTheme(newTheme, newBgColor, newCardBg, newBorderColor, newTextColor) {{
                    mainChart.dispose();
                    inflowChart.dispose();
                    
                    document.body.style.backgroundColor = newBgColor;
                    
                    var divs = document.querySelectorAll('#main, #inflow');
                    divs.forEach(function(d) {{
                        d.style.backgroundColor = newCardBg;
                        d.style.borderColor = newBorderColor;
                    }});
                    
                    currentTheme = newTheme;
                    textCol = newTextColor;
                    
                    mainChart = echarts.init(document.getElementById('main'), currentTheme);
                    inflowChart = echarts.init(document.getElementById('inflow'), currentTheme);
                    
                    mainOption.title.textStyle.color = textCol;
                    mainOption.legend.textStyle.color = textCol;
                    mainOption.xAxis[0].axisLine.lineStyle.color = textCol;
                    mainOption.xAxis[0].axisLabel.color = textCol;
                    mainOption.yAxis[0].axisLine.lineStyle.color = textCol;
                    mainOption.yAxis[0].splitLine.lineStyle.color = textCol;
                    mainOption.yAxis[0].axisLabel.color = textCol;
                    mainChart.setOption(mainOption);
                    
                    inflowOption.title.textStyle.color = textCol;
                    inflowOption.legend.textStyle.color = textCol;
                    inflowOption.series[0].itemStyle.borderColor = newCardBg;
                    inflowChart.setOption(inflowOption);
                }}
            </script>
        </body>
        </html>
        """
        self.chart_view.setHtml(html)
        
    def update_chart_theme(self):
        if not hasattr(self, 'chart_view'): return
        is_dark = isDarkTheme()
        theme_str = 'dark' if is_dark else 'light'
        bg_color = '#202020' if is_dark else '#F3F3F3'
        card_bg = '#2C2C2C' if is_dark else '#FFFFFF'
        border_color = '#3A3A3A' if is_dark else '#E5E5E5'
        text_color = '#FFFFFF' if is_dark else '#000000'
        
        js_code = f"updateTheme('{theme_str}', '{bg_color}', '{card_bg}', '{border_color}', '{text_color}');"
        self.chart_view.page().runJavaScript(js_code)
        
        if hasattr(self, 'combo_lbl'):
            self.combo_lbl.setStyleSheet(f"color: {text_color}; background: transparent;")

class SpellCheckWorker(QThread):
    finished = pyqtSignal(dict)
    error = pyqtSignal(str)

    def __init__(self, text, parent=None):
        super().__init__(parent)
        self.text = text

    def run(self):
        try:
            result = check_text(self.text)
            self.finished.emit(result)
        except Exception as e:
            self.error.emit(str(e))

class SpellCheckIssueCard(QFrame):
    def __init__(self, issue, parent=None):
        super().__init__(parent)
        self.issue = issue  # SpellCheckIssue
        self.setObjectName("SpellCheckIssueCard")
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(8)
        
        top_layout = QHBoxLayout()
        
        org_label = QLabel(issue.original, self)
        org_label.setFont(QFont("SUIT", 12, QFont.Bold))
        org_label.setStyleSheet("color: #FF6B6B; text-decoration: line-through; background: transparent; border: none;")
        top_layout.addWidget(org_label)
        
        arrow_label = QLabel("→", self)
        arrow_label.setFont(QFont("SUIT", 12, QFont.Bold))
        arrow_label.setStyleSheet("color: #A0A0A0; background: transparent; border: none;")
        top_layout.addWidget(arrow_label)
        
        sug_text = issue.suggestions[0] if issue.suggestions else "(없음)"
        sug_label = QLabel(sug_text, self)
        sug_label.setFont(QFont("SUIT", 12, QFont.Bold))
        sug_label.setStyleSheet("color: #40C463; background: transparent; border: none;")
        top_layout.addWidget(sug_label)
        
        top_layout.addStretch(1)
        layout.addLayout(top_layout)
        
        if issue.reason:
            reason_label = QLabel(issue.reason, self)
            reason_label.setFont(QFont("SUIT", 11))
            reason_label.setWordWrap(True)
            reason_label.setStyleSheet("color: #E5E5E5; background: transparent; border: none; line-height: 1.4;")
            layout.addWidget(reason_label)
            
        self.update_style()
        qconfig.themeChanged.connect(self.update_style)
        
    def update_style(self):
        is_dark = isDarkTheme()
        bg_color = "#2C2C2C" if is_dark else "#FFFFFF"
        border_color = "#3A3A3A" if is_dark else "#E5E5E5"
        self.setStyleSheet(f"""
            QFrame#SpellCheckIssueCard {{
                background-color: {bg_color};
                border: 1px solid {border_color};
                border-radius: 8px;
            }}
        """)

class SpellCheckInterface(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("SpellCheckInterface")
        self.worker = None
        
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(36, 22, 36, 22)
        main_layout.setSpacing(20)
        
        # Title
        title_layout = QHBoxLayout()
        title_lbl = TitleLabel("맞춤법 검사기", self)
        title_layout.addWidget(title_lbl)
        
        self.status_lbl = QLabel("", self)
        self.status_lbl.setFont(QFont("SUIT", 11))
        self.status_lbl.setStyleSheet("color: #A0A0A0; background: transparent; border: none;")
        title_layout.addWidget(self.status_lbl)
        
        title_layout.addStretch(1)
        
        self.progress_ring = IndeterminateProgressRing(self)
        self.progress_ring.setFixedSize(20, 20)
        self.progress_ring.hide()
        title_layout.addWidget(self.progress_ring)
        
        main_layout.addLayout(title_layout)
        
        # Split layout (Left: Input, Right: Output)
        split_layout = QHBoxLayout()
        split_layout.setSpacing(24)
        
        # Left Panel (Input)
        left_panel = QWidget(self)
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(12)
        
        left_layout.addWidget(SubtitleLabel("검사할 텍스트"))
        
        self.input_edit = TextEdit(self)
        self.input_edit.setPlaceholderText("여기에 검사할 텍스트를 입력하거나 붙여넣으세요...")
        self.input_edit.setFont(QFont("SUIT", 12))
        left_layout.addWidget(self.input_edit)
        
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)
        
        self.check_btn = PrimaryPushButton("검사 시작", self)
        self.check_btn.setFixedWidth(120)
        self.check_btn.clicked.connect(self.start_check)
        
        self.clear_btn = PushButton("비우기", self)
        self.clear_btn.setFixedWidth(100)
        self.clear_btn.clicked.connect(self.clear_text)
        
        btn_layout.addWidget(self.check_btn)
        btn_layout.addWidget(self.clear_btn)
        btn_layout.addStretch(1)
        left_layout.addLayout(btn_layout)
        
        split_layout.addWidget(left_panel, 3)
        
        # Right Panel (Output & Details)
        right_panel = QWidget(self)
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(12)
        
        right_layout.addWidget(SubtitleLabel("교정 완료 결과"))
        
        self.output_edit = TextEdit(self)
        self.output_edit.setReadOnly(True)
        self.output_edit.setFont(QFont("SUIT", 12))
        right_layout.addWidget(self.output_edit, 2)
        
        right_layout.addWidget(SubtitleLabel("상세 교정 내역"))
        
        self.scroll_area = ScrollArea(self)
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        
        self.scroll_content = QWidget()
        self.scroll_content.setObjectName("ScrollContent")
        self.scroll_content.setStyleSheet("QWidget#ScrollContent { background: transparent; }")
        self.scroll_layout = QVBoxLayout(self.scroll_content)
        self.scroll_layout.setContentsMargins(0, 0, 10, 0)
        self.scroll_layout.setSpacing(10)
        self.scroll_layout.addStretch(1)
        
        self.scroll_area.setWidget(self.scroll_content)
        right_layout.addWidget(self.scroll_area, 3)
        
        split_layout.addWidget(right_panel, 2)
        
        main_layout.addLayout(split_layout)
        
        self.update_style()
        qconfig.themeChanged.connect(self.update_style)
        
    def update_style(self):
        is_dark = isDarkTheme()
        bg_color = "#202020" if is_dark else "#FFFFFF"
        border_color = "#3A3A3A" if is_dark else "#E5E5E5"
        self.scroll_area.setStyleSheet(f"QScrollArea {{ border: 1px solid {border_color}; border-radius: 8px; background-color: {bg_color}; }}")
        
    def clear_text(self):
        self.input_edit.clear()
        self.output_edit.clear()
        self.clear_cards()
        self.status_lbl.setText("")
        
    def clear_cards(self):
        while self.scroll_layout.count():
            child = self.scroll_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()
        self.scroll_layout.addStretch(1)
        
    def start_check(self):
        text = self.input_edit.toPlainText().strip()
        if not text:
            InfoBar.warning("안내", "검사할 텍스트를 입력해주세요.", duration=3000, parent=self)
            return
            
        self.check_btn.setEnabled(False)
        self.clear_btn.setEnabled(False)
        self.progress_ring.show()
        self.progress_ring.start()
        self.status_lbl.setText("맞춤법 검사 중...")
        self.output_edit.clear()
        self.clear_cards()
        
        self.worker = SpellCheckWorker(text, self)
        self.worker.finished.connect(self.on_check_finished)
        self.worker.error.connect(self.on_check_error)
        self.worker.start()
        
    def on_check_finished(self, result):
        self.check_btn.setEnabled(True)
        self.clear_btn.setEnabled(True)
        self.progress_ring.stop()
        self.progress_ring.hide()
        self.status_lbl.setText("검사 완료")
        
        corrected_text = result.get("corrected_text", "")
        issues = result.get("issues", [])
        
        rendered_html = corrected_text.replace("\n", "<br>")
        
        unique_suggestions = sorted(list(set(issue.suggestions[0] for issue in issues if issue.suggestions)), key=len, reverse=True)
        for sug in unique_suggestions:
            if sug:
                highlight_tag = f'<span style="background-color: rgba(10, 132, 255, 0.28); color: #60CDFF; font-weight: 800; border-radius: 4px; padding: 2px 4px;">{sug}</span>'
                escaped_sug = re.escape(sug)
                rendered_html = re.sub(rf'(?<![0-9a-zA-Z가-힣]){escaped_sug}(?![0-9a-zA-Z가-힣])(?![^<]*>)', highlight_tag, rendered_html)
                
        self.output_edit.setHtml(f"""
            <html>
            <head>
                <style>
                    body {{
                        font-family: 'SUIT', sans-serif;
                        font-size: 16px;
                        color: #E5E5E5;
                        line-height: 1.8;
                    }}
                </style>
            </head>
            <body>
                {rendered_html}
            </body>
            </html>
        """)
        
        if issues:
            for issue in issues:
                card = SpellCheckIssueCard(issue, self.scroll_content)
                insert_idx = max(0, self.scroll_layout.count() - 1)
                self.scroll_layout.insertWidget(insert_idx, card)
        else:
            no_issue_lbl = QLabel("검출된 맞춤법 오류가 없습니다. 완벽한 문장입니다!", self.scroll_content)
            no_issue_lbl.setFont(QFont("SUIT", 12, QFont.Bold))
            no_issue_lbl.setAlignment(Qt.AlignCenter)
            no_issue_lbl.setStyleSheet("color: #40C463; padding: 20px; background: transparent; border: none;")
            insert_idx = max(0, self.scroll_layout.count() - 1)
            self.scroll_layout.insertWidget(insert_idx, no_issue_lbl)
            
    def on_check_error(self, err_msg):
        self.check_btn.setEnabled(True)
        self.clear_btn.setEnabled(True)
        self.progress_ring.stop()
        self.progress_ring.hide()
        self.status_lbl.setText("오류 발생")
        
        MessageBox("맞춤법 검사 실패", f"맞춤법 검사 중 오류가 발생했습니다:\n{err_msg}", self).exec_()

class PlaceScraperWorker(QThread):
    def __init__(self, keyword, target_company, display_count, global_driver_path):
        super().__init__()
        self.keyword = keyword
        self.target_company = target_company
        self.display_count = display_count
        self.global_driver_path = global_driver_path
        self.signals = WorkerSignals()
        self.driver = None

    def log(self, message):
        self.signals.log.emit(message)

    def cleanup_drivers(self):
        if self.driver:
            try:
                self.driver.quit()
            except:
                pass
            self.driver = None

    def get_place_data_by_crawling(self, driver, keyword, target_company):
        titles = []
        
        self.log(f"▶ '{keyword}' - 데이터 탐색 중... ({self.display_count}개 목표)")
        
        try:
            departments = [
                "이비인후과", "마취통증의학과", "정신건강의학과", "소아청소년과", 
                "재활의학과", "가정의학과", "성형외과", "정형외과", "신경외과", 
                "피부과", "산부인과", "비뇨기과", "통증의학과", "신경과", 
                "안과", "치과", "내과", "외과", "소아과", "한의원", "의원", "한방병원", "병원"
            ]
            dept = keyword
            for d in departments:
                if d in keyword:
                    dept = d
                    break
            
            q = urllib.parse.quote(keyword)
            d_q = urllib.parse.quote(dept)
            url = f"https://m.place.naver.com/hospital/list?query={q}&department={d_q}&x=126.9783882&y=37.5666103&level=top&entry=pll&originalQuery={q}"
            driver.get(url)
            
            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "li"))
                )
            except TimeoutException:
                self.log(f"[안내] '{keyword}' - 요소를 찾지 못해 대기 시간이 초과되었습니다.")
                return pd.DataFrame({'제목': pd.Series(dtype='str')})

            last_height = driver.execute_script("return document.body.scrollHeight")
            
            while True:
                js_count = """
                var els = document.querySelectorAll("li");
                var organicCount = 0;
                for(var i=0; i<els.length; i++) {
                    var isAd = false;
                    
                    var adLinks = els[i].querySelectorAll("a");
                    for(var j=0; j<adLinks.length; j++) {
                        var href = adLinks[j].getAttribute("href") || "";
                        if(href.indexOf("ader.naver.com") !== -1 || href.indexOf("adcr.naver.com") !== -1 || href.indexOf("NSPM_70") !== -1) {
                            isAd = true;
                            break;
                        }
                    }
                    if(!isAd) {
                        var svgs = els[i].querySelectorAll("svg");
                        for(var j=0; j<svgs.length; j++) {
                            var vb = svgs[j].getAttribute("viewBox");
                            if(vb === "0 0 39 16" || vb === "0 0 38 16" || vb === "0 0 40 16") {
                                isAd = true;
                                break;
                            }
                        }
                    }
                    if(!isAd) {
                        var blinds = els[i].querySelectorAll(".place_blind");
                        for(var j=0; j<blinds.length; j++) {
                            if(blinds[j].textContent.trim() === "광고") {
                                isAd = true;
                                break;
                            }
                        }
                    }
                    
                    if (!isAd) {
                        var links = els[i].querySelectorAll("a");
                        var foundTitle = false;
                        for(var k=0; k<links.length; k++) {
                            var a_text = links[k].innerText || "";
                            var first_line = a_text.trim().split('\\n')[0].trim();
                            if(first_line.length > 0 && first_line.indexOf("이미지 수") === -1 && isNaN(first_line)) {
                                foundTitle = true;
                                break;
                            }
                        }
                        if(foundTitle) {
                            organicCount++;
                        }
                    }
                }
                return organicCount;
                """
                current_count = driver.execute_script(js_count)
                
                if current_count >= self.display_count:
                    break
                    
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(0.5) 
                
                new_height = driver.execute_script("return document.body.scrollHeight")
                if new_height == last_height:
                    time.sleep(1.0) 
                    new_height = driver.execute_script("return document.body.scrollHeight")
                    if new_height == last_height:
                        break 
                last_height = new_height

            js_extract = """
            var results = [];
            var els = document.querySelectorAll("li");
            for(var i=0; i<els.length; i++) {
                var isAd = false;
                
                var adLinks = els[i].querySelectorAll("a");
                for(var j=0; j<adLinks.length; j++) {
                    var href = adLinks[j].getAttribute("href") || "";
                    if(href.indexOf("ader.naver.com") !== -1 || href.indexOf("adcr.naver.com") !== -1 || href.indexOf("NSPM_70") !== -1) {
                        isAd = true;
                        break;
                    }
                }
                if(!isAd) {
                    var svgs = els[i].querySelectorAll("svg");
                    for(var j=0; j<svgs.length; j++) {
                        var vb = svgs[j].getAttribute("viewBox");
                        if(vb === "0 0 39 16" || vb === "0 0 38 16" || vb === "0 0 40 16") {
                            isAd = true;
                            break;
                        }
                    }
                }
                if(!isAd) {
                    var blinds = els[i].querySelectorAll(".place_blind");
                    for(var j=0; j<blinds.length; j++) {
                        if(blinds[j].textContent.trim() === "광고") {
                            isAd = true;
                            break;
                        }
                    }
                }
                
                if(isAd) continue;
                
                var links = els[i].querySelectorAll("a");
                var title = "";
                for(var k=0; k<links.length; k++) {
                    var a_text = links[k].innerText || "";
                    var first_line = a_text.trim().split('\\n')[0].trim();
                    if(first_line.length > 0 && first_line.indexOf("이미지 수") === -1 && isNaN(first_line)) {
                        title = first_line;
                        break;
                    }
                }
                if(title.length > 0) {
                    results.push(title);
                }
            }
            return results;
            """
            extracted_data = driver.execute_script(js_extract)
            
            for data in extracted_data[:self.display_count]:
                titles.append(data)
                
            if not titles:
                self.log(f"[안내] '{keyword}' - 데이터를 추출하지 못했습니다.")
                        
        except Exception as e:
            self.log(f"[오류] '{keyword}' - 크롤링 실패: {e}")
                
        df = pd.DataFrame({'제목': titles})
        df.index = df.index + 1
        return df

    def get_rank_string(self, df_result, target_company):
        if df_result.empty or not target_company:
            return "검색/분석 불가"
        
        valid_target = target_company.strip()
        if not valid_target:
            return "미노출 (범위 밖)"
            
        pattern = re.escape(valid_target)
        mask = df_result['제목'].str.contains(pattern, na=False, regex=True)
        found_indices = df_result.index[mask].tolist()
        if found_indices:
            return ", ".join([f"{i}위" for i in found_indices])
        return "미노출 (범위 밖)"

    def run(self):
        try:
            if not self.keyword.strip() or not self.target_company.strip():
                self.signals.error.emit("키워드와 목표 업체명을 모두 입력해주세요.")
                return

            self.log(f"=== 키워드: '{self.keyword}', 목표 업체명: '{self.target_company}' ===")
            
            delay_time = random.uniform(2.5, 4.5)
            self.log(f"※ 봇 탐지 우회를 위해 {delay_time:.1f}초 대기 중...")
            time.sleep(delay_time)
            
            options = Options()
            options.page_load_strategy = 'eager'
            options.add_argument('--headless=new')
            options.add_argument('--disable-gpu')
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--window-size=390,844')
            options.add_argument('user-agent=Mozilla/5.0 (iPhone; CPU iPhone OS 16_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Mobile/15E148 Safari/604.1')
            options.add_argument('--disable-blink-features=AutomationControlled')
            
            service = Service(self.global_driver_path)
            self.driver = webdriver.Chrome(service=service, options=options)
            
            df_sim = self.get_place_data_by_crawling(self.driver, self.keyword, self.target_company)
            sim_rank = self.get_rank_string(df_sim, self.target_company)
            
            ranks = []
            if sim_rank and "위" in sim_rank:
                for part in sim_rank.split(","):
                    if "위" in part:
                        try:
                            r_val = int(part.replace("위", "").strip())
                            ranks.append(r_val)
                        except:
                            pass
                            
            if ranks:
                self.signals.match_found.emit({
                    'company': self.target_company,
                    'keyword': self.keyword,
                    'ranks': ranks,
                    'screenshots': [],
                    'folder': ''
                })
            
            result_text = f"<< '{self.keyword}' 검색 결과 >>\n"
            for idx, row in df_sim.iterrows():
                result_text += f"[{idx}위] {row['제목']}\n"
            self.log(result_text)
            self.log(f"▶ 기록 준비 완료: {sim_rank}")
            
            self.signals.finished.emit()

        except Exception as e:
            self.log(f"[메인 오류] {e}")
            traceback.print_exc()
            self.signals.error.emit(str(e))
        finally:
            self.cleanup_drivers()

class PlaceExposureCard(QFrame):
    def __init__(self, match_data, parent=None):
        super().__init__(parent)
        self.match_data = match_data  
        self.setObjectName("PlaceExposureCard")
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 14, 16, 14)
        layout.setSpacing(10)
        
        header_layout = QHBoxLayout()
        header_layout.setSpacing(12)
        
        self.company_label = QLabel(match_data.get('company', ''), self)
        self.company_label.setFont(QFont("SUIT", 12, QFont.Bold))
        header_layout.addWidget(self.company_label)
        
        self.keyword_label = QLabel(match_data.get('keyword', ''), self)
        self.keyword_label.setFont(QFont("SUIT", 11, QFont.Bold))
        header_layout.addWidget(self.keyword_label)
        
        header_layout.addStretch(1)
        layout.addLayout(header_layout)
        
        self.ranks_box = QFrame(self)
        self.ranks_box.setObjectName("RanksBox")
        box_layout = QHBoxLayout(self.ranks_box)
        box_layout.setContentsMargins(12, 6, 12, 6)
        box_layout.setSpacing(16)
        
        for rank in match_data.get('ranks', []):
            lbl = QLabel(f"{rank}위", self.ranks_box)
            lbl.setFont(QFont("SUIT", 11, QFont.Bold))
            lbl.setStyleSheet("color: #60CDFF; background: transparent; border: none;")
            box_layout.addWidget(lbl)
            
        box_layout.addStretch(1)
        layout.addWidget(self.ranks_box)
        
        self.update_style()
        qconfig.themeChanged.connect(self.update_style)
        
    def update_style(self):
        is_dark = isDarkTheme()
        bg_color = "#2C2C2C" if is_dark else "#F3F3F3"
        border_color = "#3A3A3A" if is_dark else "#E5E5E5"
        text_color = "#FFFFFF" if is_dark else "#000000"
        sub_text_color = "#CCCCCC" if is_dark else "#333333"
        box_bg = "#161616" if is_dark else "#E5E5E5"
        
        self.setStyleSheet(f"""
            QFrame#PlaceExposureCard {{
                background-color: {bg_color};
                border: 1px solid {border_color};
                border-radius: 12px;
            }}
            QFrame#RanksBox {{
                background-color: {box_bg};
                border-radius: 8px;
            }}
        """)
        self.company_label.setStyleSheet(f"color: {text_color}; background: transparent; border: none;")
        self.keyword_label.setStyleSheet(f"color: {sub_text_color}; background: transparent; border: none;")

class PlaceScraperInterface(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent=parent)
        self.setObjectName("PlaceScraperInterface")
        self.global_driver_path = ""
        
        self.init_ui()
        self.check_environment()

    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(36, 36, 36, 36)
        main_layout.setSpacing(16)

        self.title_label = TitleLabel("플레이스 순위 체크")
        main_layout.addWidget(self.title_label)

        split_layout = QHBoxLayout()
        split_layout.setSpacing(24)
        
        self.left_panel = QWidget(self)
        left_layout = QVBoxLayout(self.left_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(16)

        input_layout = QGridLayout()
        input_layout.setSpacing(12)
        
        self.keyword_input = LineEdit()
        self.keyword_input.setPlaceholderText("예: 부산 성형외과")
        input_layout.addWidget(SubtitleLabel("검색 키워드:"), 0, 0)
        input_layout.addWidget(self.keyword_input, 0, 1)
        
        self.company_input = LineEdit()
        self.company_input.setPlaceholderText("예: 푸름애드 의원")
        input_layout.addWidget(SubtitleLabel("목표 업체명:"), 1, 0)
        input_layout.addWidget(self.company_input, 1, 1)
        
        self.count_spinbox = SpinBox()
        self.count_spinbox.setRange(1, 150)
        self.count_spinbox.setValue(50)
        input_layout.addWidget(SubtitleLabel("탐색 목표 개수:"), 2, 0)
        input_layout.addWidget(self.count_spinbox, 2, 1)
        
        left_layout.addLayout(input_layout)

        self.loading_container = QWidget()
        loading_layout = QHBoxLayout(self.loading_container)
        loading_layout.setAlignment(Qt.AlignCenter)
        loading_layout.setContentsMargins(0, 0, 0, 0)
        
        self.loading_ring = IndeterminateProgressRing()
        self.loading_ring.setFixedSize(25, 25)
        self.loading_label = SubtitleLabel("크롬 드라이버를 점검/동기화 중입니다...")
        
        loading_layout.addWidget(self.loading_ring)
        loading_layout.addWidget(self.loading_label)
        left_layout.addWidget(self.loading_container)
        self.loading_container.hide()

        button_layout = QHBoxLayout()
        
        self.start_btn = PushButton("순위 체크 시작")
        self.start_btn.setStyleSheet("""
            QPushButton {
                background-color: #3CA0F0;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-family: 'SUIT';
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #59B4FF;
            }
            QPushButton:pressed {
                background-color: #268CD9;
            }
            QPushButton:disabled {
                background-color: #2C2C2C;
                color: #666666;
            }
        """)
        self.start_btn.clicked.connect(self.start_scraping)
        button_layout.addWidget(self.start_btn)
        button_layout.addStretch(1)
        
        left_layout.addLayout(button_layout)

        self.console_output = TextEdit()
        self.console_output.setReadOnly(True)
        left_layout.addWidget(self.console_output)

        split_layout.addWidget(self.left_panel, 3) 

        self.right_panel = QWidget(self)
        right_layout = QVBoxLayout(self.right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(12)
        
        right_title = SubtitleLabel("실시간 노출 현황")
        right_title.setFont(QFont("SUIT", 12, QFont.Bold))
        right_layout.addWidget(right_title)
        
        self.exposure_scroll = ScrollArea(self.right_panel)
        self.exposure_scroll.setWidgetResizable(True)
        
        self.exposure_content = QWidget()
        self.exposure_content.setObjectName("ExposureContent")
        self.exposure_content.setStyleSheet("QWidget#ExposureContent { background-color: transparent; }")
        self.exposure_layout = QVBoxLayout(self.exposure_content)
        self.exposure_layout.setContentsMargins(12, 12, 12, 12)
        self.exposure_layout.setSpacing(12)
        self.exposure_layout.addStretch(1)
        
        self.exposure_scroll.setWidget(self.exposure_content)
        right_layout.addWidget(self.exposure_scroll)
        
        split_layout.addWidget(self.right_panel, 2) 

        main_layout.addLayout(split_layout)

        self.update_right_panel_style()
        qconfig.themeChanged.connect(self.update_right_panel_style)

    def update_right_panel_style(self):
        is_dark = isDarkTheme()
        border_color = "#3A3A3A" if is_dark else "#E5E5E5"
        bg_color = "#202020" if is_dark else "#FFFFFF"
        self.exposure_scroll.setStyleSheet(f"QScrollArea {{ border: 1px solid {border_color}; border-radius: 8px; background-color: {bg_color}; }}")

    def clear_exposure_cards(self):
        while self.exposure_layout.count():
            child = self.exposure_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()
        self.exposure_layout.addStretch(1)

    def add_exposure_card(self, match_data):
        card = PlaceExposureCard(match_data, self.exposure_content)
        self.exposure_layout.insertWidget(0, card)
        
        while self.exposure_layout.count() > 7:
            idx_to_remove = self.exposure_layout.count() - 2
            item = self.exposure_layout.takeAt(idx_to_remove)
            if item.widget():
                item.widget().deleteLater()

    def append_log(self, text):
        self.console_output.append(text)
        scrollbar = self.console_output.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    def check_environment(self):
        self.append_log("[시스템] 초기 환경을 점검합니다...")
        self.start_btn.setEnabled(False)
        self.loading_container.show()
        self.loading_ring.start()
        
        self.driver_worker = DriverInitWorker()
        self.driver_worker.finished.connect(self.on_driver_ready)
        self.driver_worker.error.connect(self.on_driver_error)
        self.driver_worker.start()

    def on_driver_ready(self, path):
        self.global_driver_path = path
        self.loading_ring.stop()
        self.loading_container.hide()
        self.append_log("[시스템] 크롬 드라이버 동기화 및 준비 완료.")
        self.start_btn.setEnabled(True)

    def on_driver_error(self, err_msg):
        self.loading_ring.stop()
        self.loading_container.hide()
        self.append_log(f"[치명적 오류] 크롬 드라이버 설치 실패: {err_msg}")
        InfoBar.error("오류", "크롬 드라이버를 설치할 수 없습니다.", duration=5000, position=InfoBarPosition.TOP, parent=self)

    def start_scraping(self):
        keyword = self.keyword_input.text()
        company = self.company_input.text()
        
        if not keyword.strip() or not company.strip():
            InfoBar.error("입력 오류", "검색 키워드와 목표 업체명을 모두 입력해주세요.", duration=3000, position=InfoBarPosition.TOP, parent=self)
            return

        self.start_btn.setEnabled(False)
        self.keyword_input.setEnabled(False)
        self.company_input.setEnabled(False)
        self.count_spinbox.setEnabled(False)
        self.console_output.clear()
        
        display_count = self.count_spinbox.value()

        self.worker = PlaceScraperWorker(
            keyword, company, display_count, self.global_driver_path
        )
        self.worker.signals.log.connect(self.append_log)
        self.worker.signals.error.connect(self.show_error)
        self.worker.signals.match_found.connect(self.add_exposure_card)
        self.worker.signals.finished.connect(self.on_scraping_finished)
        self.worker.start()

    def show_error(self, err_msg):
        InfoBar.error("작업 중단", err_msg, duration=5000, position=InfoBarPosition.TOP, parent=self)
        self.on_scraping_finished()

    def on_scraping_finished(self):
        self.start_btn.setEnabled(True)
        self.keyword_input.setEnabled(True)
        self.company_input.setEnabled(True)
        self.count_spinbox.setEnabled(True)
        self.append_log("\n[안내] 순위 체크가 종료되었습니다.")
        InfoBar.success("완료", "플레이스 순위 체크 작업이 성공적으로 종료되었습니다.", duration=4000, position=InfoBarPosition.TOP, parent=self)

def parse_version(v):
    return tuple(map(int, re.findall(r'\d+', v)))

def perform_update(release_info):
    try:
        assets = release_info.get("assets", [])
        exe_url = None
        sha_url = None
        
        for asset in assets:
            if asset["name"].endswith(".exe"):
                exe_url = asset["browser_download_url"]
            elif asset["name"].endswith(".txt") or asset["name"].endswith(".sha256"):
                sha_url = asset["browser_download_url"]
                
        if not exe_url:
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.critical(None, "오류", "업데이트 실행 파일을 찾을 수 없습니다.")
            return
            
        temp_dir = tempfile.gettempdir()
        new_exe_path = os.path.join(temp_dir, "new_app_temp.exe")
        
        # Download new exe with progress
        from PyQt5.QtWidgets import QProgressDialog, QApplication
        from PyQt5.QtCore import Qt
        
        response = requests.get(exe_url, stream=True)
        response.raise_for_status()
        total_size = int(response.headers.get('content-length', 0))
        
        progress = QProgressDialog("업데이트 파일을 다운로드 중입니다...", "취소", 0, 100, None)
        progress.setWindowTitle("업데이트 진행 중")
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)
        
        downloaded = 0
        with open(new_exe_path, "wb") as f:
            for chunk in response.iter_content(chunk_size=8192):
                if progress.wasCanceled():
                    try: os.remove(new_exe_path)
                    except: pass
                    return
                if chunk:
                    f.write(chunk)
                    downloaded += len(chunk)
                    if total_size > 0:
                        progress.setValue(int((downloaded / total_size) * 100))
                    QApplication.processEvents()
                    
        progress.setLabelText("파일 무결성을 검증 중입니다...")
        QApplication.processEvents()
        # Download and check hash
        if sha_url:
            sha_data = requests.get(sha_url).text
            expected_hash = sha_data.split()[0].strip().lower().replace("sha256:", "")
            
            sha256_hash = hashlib.sha256()
            with open(new_exe_path, "rb") as f:
                for byte_block in iter(lambda: f.read(4096), b""):
                    sha256_hash.update(byte_block)
            actual_hash = sha256_hash.hexdigest().lower()
            
            if expected_hash != actual_hash:
                from PyQt5.QtWidgets import QMessageBox
                QMessageBox.critical(None, "보안 경고", "다운로드된 파일의 무결성 검증에 실패하여 업데이트를 취소합니다.")
                try: os.remove(new_exe_path)
                except: pass
                return
                
        # Perform replacement
        current_exe = sys.executable
        if getattr(sys, 'frozen', False):
            cmd_chain = f'/c timeout /t 2 >nul & move /y "{new_exe_path}" "{current_exe}" & start "" "{current_exe}"'
            ctypes.windll.shell32.ShellExecuteW(None, "runas", "cmd.exe", cmd_chain, None, 0)
            sys.exit(0)
        else:
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.information(None, "알림", "개발 환경에서는 자동 업데이트 파일 교체를 생략합니다.")
            
    except Exception as e:
        from PyQt5.QtWidgets import QMessageBox
        QMessageBox.critical(None, "오류", f"업데이트 중 오류가 발생했습니다:\n{str(e)}")

def check_for_updates(manual_check=False):
    try:
        response = requests.get("https://api.github.com/repos/ks02149-star/test-2/releases/latest", timeout=5)
        response.raise_for_status()
        latest_release = response.json()
        latest_tag = latest_release.get("tag_name", "")
        
        if parse_version(latest_tag) > parse_version(__version__):
            from PyQt5.QtWidgets import QMessageBox
            msg_box = QMessageBox()
            msg_box.setIcon(QMessageBox.Information)
            msg_box.setWindowTitle("업데이트 알림")
            msg_box.setText("새로운 버전이 확인되었습니다. 업데이트 하시겠습니까?")
            msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            if msg_box.exec_() == QMessageBox.Yes:
                perform_update(latest_release)
        else:
            if manual_check:
                from PyQt5.QtWidgets import QMessageBox
                QMessageBox.information(None, "업데이트", "현재 최신 버전을 사용 중입니다.")
    except Exception as e:
        if manual_check:
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.critical(None, "오류", f"업데이트 확인 중 오류가 발생했습니다:\n{str(e)}")

class MainWindow(FluentWindow):
    def __init__(self):
        super().__init__()
        
        # Force Always Dark Mode
        setTheme(Theme.DARK)
        setThemeColor('#0078D4')
        
        self.init_window()
        
        self.home_interface = HomeInterface(self)
        self.scraper_interface = ScraperInterface(self)
        self.place_scraper_interface = PlaceScraperInterface(self)
        self.company_list_interface = CompanyListInterface(self)
        self.index_check_interface = IndexCheckInterface(self)
        self.spell_check_interface = SpellCheckInterface(self)
        self.settings_interface = SettingInterface(self)
        
        self.addSubInterface(self.home_interface, FluentIcon.HOME, '홈')
        self.addSubInterface(self.scraper_interface, FluentIcon.DOCUMENT, '블로그 순위 체크')
        self.addSubInterface(self.place_scraper_interface, getattr(FluentIcon, "POI", FluentIcon.SEARCH), '플레이스 순위 체크')
        self.addSubInterface(self.company_list_interface, FluentIcon.PEOPLE, '업체 리스트')
        self.addSubInterface(self.index_check_interface, getattr(FluentIcon, "PIE_SINGLE", FluentIcon.DOCUMENT), '지수 체크')
        self.addSubInterface(self.spell_check_interface, FluentIcon.EDIT, '맞춤법 검사기')
        self.addSubInterface(self.settings_interface, FluentIcon.SETTING, '설정', position=NavigationItemPosition.BOTTOM)
        
        qconfig.themeChanged.connect(self.update_theme_style)
        self.update_theme_style()
        
    def update_theme_style(self):
        is_dark = isDarkTheme()
        if is_dark:
            self.setStyleSheet("""
                MainWindow {
                    background-color: #161616;
                }
                NavigationPanel[menu=true], NavigationPanel[menu=false], NavigationPanel {
                    background-color: #161616 !important;
                    border: none !important;
                }
                StackedWidget {
                    background-color: #202020 !important;
                    border-top-left-radius: 10px !important;
                    border: 1px solid #2A2A2A !important;
                    border-right: none !important;
                    border-bottom: none !important;
                }
                QScrollArea, #ScrollContent, #scrollWidget {
                    border: none !important;
                    background-color: transparent !important;
                }
                CardWidget {
                    background-color: #2C2C2C !important;
                    border: 1px solid #3A3A3A !important;
                    border-radius: 10px !important;
                }
                TextEdit {
                    background-color: #161616 !important;
                    border: 1px solid #2C2C2C !important;
                    border-radius: 8px !important;
                }
            """)
        else:
            self.setStyleSheet("MainWindow { background-color: #F3F3F3; }")
        
    def init_window(self):
        self.setWindowTitle("푸름애드 블로그 관리 프로그램")
        self.resize(1400, 800)
        
        desktop = QApplication.desktop().availableGeometry()
        w, h = desktop.width(), desktop.height()
        self.move(w//2 - self.width()//2, h//2 - self.height()//2)
 
    def closeEvent(self, event):
        if hasattr(self.scraper_interface, 'driver_worker') and self.scraper_interface.driver_worker.isRunning():
            self.scraper_interface.driver_worker.terminate()
            self.scraper_interface.driver_worker.wait()
            
        if hasattr(self.scraper_interface, 'worker') and self.scraper_interface.worker.isRunning():
            self.scraper_interface.worker.cleanup_drivers()
            self.scraper_interface.worker.terminate()
            self.scraper_interface.worker.wait()
            
        if hasattr(self.index_check_interface, 'worker') and self.index_check_interface.worker and self.index_check_interface.worker.isRunning():
            self.index_check_interface.worker.cleanup()
            self.index_check_interface.worker.terminate()
            self.index_check_interface.worker.wait()
            
        if hasattr(self, 'spell_check_interface') and hasattr(self.spell_check_interface, 'worker') and self.spell_check_interface.worker and self.spell_check_interface.worker.isRunning():
            self.spell_check_interface.worker.terminate()
            self.spell_check_interface.worker.wait()
            
        if hasattr(self, 'place_scraper_interface') and hasattr(self.place_scraper_interface, 'worker') and self.place_scraper_interface.worker and self.place_scraper_interface.worker.isRunning():
            self.place_scraper_interface.worker.cleanup_drivers()
            self.place_scraper_interface.worker.terminate()
            self.place_scraper_interface.worker.wait()
            
        event.accept()

if __name__ == '__main__':
    try:
        QApplication.setHighDpiScaleFactorRoundingPolicy(Qt.HighDpiScaleFactorRoundingPolicy.PassThrough)
    except AttributeError:
        pass 
        
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps)

    app = QApplication(sys.argv)
    
    base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
    font_dir = os.path.join(base_dir, "Font")
    
    loaded_family = "SUIT"
    if os.path.exists(font_dir):
        # Only load highly readable weights to prevent thin/light fallback issues
        allowed_suffixes = ("regular.otf", "medium.otf", "semibold.otf", "bold.otf",
                            "regular.ttf", "medium.ttf", "semibold.ttf", "bold.ttf")
        for file_name in os.listdir(font_dir):
            if file_name.lower().endswith(('.otf', '.ttf')):
                if any(suffix in file_name.lower() for suffix in allowed_suffixes):
                    font_path = os.path.join(font_dir, file_name)
                    font_id = QFontDatabase.addApplicationFont(font_path)
                    if font_id != -1:
                        families = QFontDatabase.applicationFontFamilies(font_id)
                        if families and loaded_family != "SUIT":
                            loaded_family = families[0]
                            
    app_font = QFont(loaded_family, 10, QFont.Normal)
    app.setFont(app_font)

    check_for_updates()

    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
